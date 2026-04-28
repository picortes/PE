# -*- coding: utf-8 -*-
# ====================================================================================
# IMPORTACIONES Y CONFIGURACIÓN
# ====================================================================================
import os
import re
import sys
import time

# 🆕 Configurar stdout para UTF-8 (soportar emojis en terminal)
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import timedelta
import pyodbc
from functools import wraps
import json
import socket
from flask_cors import CORS, cross_origin
from datetime import datetime
import hashlib
from io import BytesIO

# Importaciones para generación de PDF
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.pdfgen import canvas
    PDF_DISPONIBLE = True
except ImportError:
    PDF_DISPONIBLE = False
    print("⚠️ ADVERTENCIA: reportlab no está instalado. La generación de PDF no estará disponible.")
    print("   Instalar con: pip install reportlab")

# Importaciones para generación de Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("⚠️ ADVERTENCIA: openpyxl no está instalado. La generación de Excel no estará disponible.")
    print("   Instalar con: pip install openpyxl")

# ====================================================================================
# FUNCIONES DE ENCRIPTACIÓN DE CONTRASEÑAS CON WERKZEUG
# ====================================================================================
def hash_password(password):
    """
    Encripta una contraseña usando werkzeug.security (Scrypt).
    Método recomendado y sin dependencias externas.
    """
    return generate_password_hash(password)
 
def verify_password(password, hashed_password):
    """
    Verifica una contraseña contra su hash.
    Soporta werkzeug (principal), SHA-256 con salt (legacy) y texto plano (legacy).
    Nota: Hashes legacy ya no son soportados - requieren migración.
    """
    if not hashed_password:
        return False
   
    try:
        # Verificar si es hash werkzeug (Scrypt o PBKDF2) - Método principal
        if hashed_password.startswith('scrypt:') or hashed_password.startswith('pbkdf2:'):
            return check_password_hash(hashed_password, password)
       
        # Hashes legacy no soportados - migrar a werkzeug
        if hashed_password.startswith('$2b$') or hashed_password.startswith('$2a$'):
            print("⚠️ Hash legacy encontrado - requiere migración a werkzeug")
            return False
       
        # Verificar si es hash SHA-256 con salt (legacy)
        if hashed_password.startswith('sha256$'):
            parts = hashed_password.split('$')
            if len(parts) == 3:
                salt = parts[1]
                stored_hash = parts[2]
                hash_obj = hashlib.sha256((password + salt).encode('utf-8'))
                return hash_obj.hexdigest() == stored_hash
       
        # Fallback: comparación directa (contraseñas legacy en texto plano)
        return password == hashed_password
       
    except Exception as e:
        print(f"Error en verificación de contraseña: {e}")
        return False
 
def is_password_hashed(password):
    """
    Determina si una contraseña ya está encriptada.
    Soporta formatos: werkzeug (scrypt/pbkdf2), SHA-256 legacy.
    """
    if not password:
        return False
   
    # Verificar si es werkzeug (scrypt o pbkdf2)
    if password.startswith('scrypt:') or password.startswith('pbkdf2:'):
        return True
   
    # Verificar si es hash legacy (ya no soportado - requiere migración)
    if password.startswith('$2b$') or password.startswith('$2a$'):
        return True
   
    # Verificar si es SHA-256 con salt (legacy)
    if password.startswith('sha256$'):
        return True
   
    return False
 
def migrate_password_if_needed(password):
    """
    Migra una contraseña legacy a formato werkzeug si es necesario.
    Prioriza werkzeug sobre otros formatos.
    """
    if password.startswith('scrypt:') or password.startswith('pbkdf2:'):
        return password  # Ya está en formato werkzeug
   
    # La contraseña necesita migración a werkzeug (texto plano, hash legacy, o sha256)
    return hash_password(password)

# Configuración directa de usuario y contraseña de la base de datos
DB_USER = 'sa'
DB_PASSWORD = 'Coppersink10EMESA'

SECRET_KEY = None
SHUTDOWN_SECRET_KEY = None

# Inicializar Flask (SOLO UNA VEZ)
app = Flask(__name__)
app.secret_key = 'emesa_checklist_secret_key_2025'  # Para sesiones
CORS(app, supports_credentials=True, resources={r"/*": {"origins": "*"}}, methods=["GET", "POST", "PUT", "OPTIONS"], allow_headers=["Content-Type"])

# 🔒 Configuración HTTPS con certificado firmado por CA interna EMESA
def generate_self_signed_cert():
    """
    Localiza los certificados firmados por la CA interna de EMESA.
    Orden de búsqueda:
      1. Mismo directorio que app.py  (cert.pem / key.pem copiados al desplegar)
      2. Carpeta Secretos de skills   (solo en desarrollo)
    No genera certificados auto-firmados: los certificados son emitidos por la CA
    raiz EMESA y distribuidos por GPO, por lo que no provocan advertencias en el navegador.
    """
    # Directorio base: junto a app.py (también funciona empaquetado con PyInstaller)
    if getattr(sys, 'frozen', False):
        certs_dir = os.path.dirname(sys.executable)
    else:
        certs_dir = os.path.dirname(os.path.abspath(__file__))

    cert_file = os.path.join(certs_dir, 'cert.pem')
    key_file  = os.path.join(certs_dir, 'key.pem')

    if os.path.exists(cert_file) and os.path.exists(key_file):
        print(f"✅ Certificado CA-EMESA encontrado en: {certs_dir}")
        return cert_file, key_file

    # Fallback: carpeta Secretos de skills (solo entorno de desarrollo)
    secretos_dir = os.path.join(
        os.path.expanduser("~"), ".codex", "skills", "Secretos",
        "emesa-internal-https-ca", "servers", "192.168.253.9"
    )
    cert_file_s = os.path.join(secretos_dir, 'cert.pem')
    key_file_s  = os.path.join(secretos_dir, 'key.pem')

    if os.path.exists(cert_file_s) and os.path.exists(key_file_s):
        print(f"✅ Certificado CA-EMESA encontrado en Secretos: {secretos_dir}")
        return cert_file_s, key_file_s

    print("❌ No se encontraron certificados CA-EMESA.")
    print(f"   Copia cert.pem y key.pem en: {certs_dir}")
    return None, None

# Middleware para agregar headers de seguridad y permisos de cámara
@app.after_request
def add_security_headers(response):
    """Agregar headers de seguridad y permisos de cámara para QR Scanner"""
    # Permitir acceso a cámara, micrófono y geolocalización desde cualquier origen
    response.headers['Permissions-Policy'] = 'camera=*, microphone=*, geolocation=*'
    # Permitir acceso desde cualquier origen (CORS)
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS, PATCH'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# Obtener directorio del proyecto de manera dinámica basado en la ubicación de app.py
# Cuando se ejecuta como binario (PyInstaller), los recursos viven en sys._MEIPASS
if getattr(sys, 'frozen', False):
    BASE_DIR = getattr(sys, '_MEIPASS', os.getcwd())  # type: ignore[attr-defined]
    CURRENT_DIR = os.path.join(BASE_DIR, "api")
    if not os.path.isdir(CURRENT_DIR):  # Fallback si la carpeta api no existe en el paquete
        CURRENT_DIR = BASE_DIR
else:
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))  # Directorio donde está app.py
    BASE_DIR = os.path.dirname(CURRENT_DIR)  # Directorio padre (raíz del proyecto)
RUTA_TEMPLATES = os.path.join(BASE_DIR, "Templates")
RUTA_ASSETS = os.path.join(BASE_DIR, "assets")
RUTA_IMAGENES = os.path.join(BASE_DIR, "IMAGENES")
RUTA_PDFS = os.path.join(BASE_DIR, "PDFs_Generados")  # 🆕 Carpeta para PDFs guardados
RUTA_PDFS_COMPARTIDA = r"\\servidor\unidad compartida\Checklist Calidad\PDFs"  # 🆕 Carpeta compartida
RUTA_FOTOS_COMPARTIDA = r"\\servidor\unidad compartida\Checklist Calidad\Fotos"  # 🆕 Carpeta compartida para FOTOS

# Crear carpeta de PDFs si no existe
if not os.path.exists(RUTA_PDFS):
    try:
        os.makedirs(RUTA_PDFS)
        print("[OK] Carpeta de PDFs creada: {}".format(RUTA_PDFS))
    except Exception as e:
        print("[WARN] No se pudo crear carpeta de PDFs: {}".format(e))

# Verificar acceso a carpeta compartida
if os.path.exists(RUTA_PDFS_COMPARTIDA):
    print("[OK] Carpeta compartida accesible: {}".format(RUTA_PDFS_COMPARTIDA))
else:
    print("[WARN] Carpeta compartida no accesible: {}".format(RUTA_PDFS_COMPARTIDA))

# ====================================================================================
# CLASE DE CONEXIÓN ODBC
# ====================================================================================
class ConexionODBC:
    def __init__(self, database=None, servidor='EMEBIDWH'):
        # Usa un driver más actual
        self.driver = 'ODBC Driver 17 for SQL Server'
        self.server = servidor
        self.database = database
        self.conn = None

    def __enter__(self):
        try:
            if not DB_USER or not DB_PASSWORD:
                raise Exception('Faltan credenciales de base de datos')

            conn_str = (
                f"DRIVER={{{self.driver}}};"
                f"SERVER={self.server};"
                f"DATABASE={self.database};"
                f"UID={DB_USER};PWD={DB_PASSWORD};"
                "TrustServerCertificate=yes;"
            )

            # Forzamos timeout para que no se quede colgado eternamente
            self.conn = pyodbc.connect(conn_str, timeout=5, autocommit=True)
            return self.conn

        except Exception as e:
            print(f"Error CRÍTICO al conectar con la base de datos: {e}")
            return None

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            self.conn.close()

# ====================================================================================
# FUNCIONES AUXILIARES PARA TURNOS Y FECHAS
# ====================================================================================
def calcular_fecha_y_turno():
    """Calcula la fecha y turno según las reglas del negocio"""
    now = datetime.now()
    hora_actual = now.hour
    
    # Determinar el turno según la hora
    if 6 <= hora_actual < 14:
        turno = "TM"
        fecha = now.date()
    elif 14 <= hora_actual < 22:
        turno = "TT" 
        fecha = now.date()
    else:  # 22:00 a 6:00 (turno nocturno)
        turno = "TN"
        # Si es entre 22:00 y 00:00, la fecha es hoy + 1
        if hora_actual >= 22:
            from datetime import timedelta
            fecha = now.date() + timedelta(days=1)
        else:  # Entre 00:00 y 6:00, mantener la fecha actual
            fecha = now.date()
    
    return fecha, turno

def obtener_usuario_sesion():
    """Obtiene el número de operario del usuario logueado de la sesión actual"""
    try:
        if 'user_data' in session and session['user_data']:
            user_data = session['user_data']
            # Devolver num_operario en lugar de nombre
            num_operario = user_data.get('num_operario')
            nombre = user_data.get('nombre', 'Usuario')
            
            if num_operario:
                return num_operario
            else:
                return nombre
        else:
            # Si no hay sesión activa, devolver un valor por defecto
            return 'Sistema'
    except Exception as e:
        print(f"Error obteniendo usuario de sesión: {e}")
        return 'Sistema'

def obtener_id_usuario_sesion():
    """Obtiene el ID interno del usuario autenticado desde la sesión actual"""
    try:
        if 'user_data' in session and session['user_data']:
            return session['user_data'].get('id')
        return None
    except Exception as e:
        print(f"Error obteniendo ID de usuario de sesión: {e}")
        return None

def obtener_id_usuario_request():
    """Obtiene el ID de usuario desde sesión o, como fallback, desde cabecera enviada por el frontend."""
    usuario_id = obtener_id_usuario_sesion()
    if usuario_id:
        return usuario_id

    try:
        header_user_id = request.headers.get('X-User-Id')
        if header_user_id:
            return int(header_user_id)
    except Exception as e:
        print(f"Error obteniendo ID de usuario desde headers: {e}")

    return None

# ====================================================================================
# RUTAS PARA SERVIR ARCHIVOS ESTÁTICOS
# ====================================================================================
@app.route('/')
def index():
    """Página principal - servir la SPA principal."""
    return send_from_directory(os.path.join(RUTA_TEMPLATES, "generales"), "ChecklistSPA.html")

@app.route('/Templates/<path:nombre_archivo>')
def servir_pantalla(nombre_archivo):
    ruta_completa = os.path.join(RUTA_TEMPLATES, nombre_archivo)
    if not os.path.exists(ruta_completa):
        return f"Archivo no encontrado: {nombre_archivo}", 404
    return send_from_directory(RUTA_TEMPLATES, nombre_archivo)

@app.route('/templates/<path:nombre_archivo>')
def servir_templates_minuscula(nombre_archivo):
    """Ruta alternativa con minúsculas"""
    ruta_completa = os.path.join(RUTA_TEMPLATES, nombre_archivo)
    if not os.path.exists(ruta_completa):
        return f"Archivo no encontrado: {nombre_archivo}", 404
    return send_from_directory(RUTA_TEMPLATES, nombre_archivo)

@app.route('/assets/<path:nombre_archivo>')
def servir_assets(nombre_archivo):
    ruta_completa = os.path.join(RUTA_ASSETS, nombre_archivo)
    if not os.path.exists(ruta_completa):
        return f"Archivo no encontrado: {nombre_archivo}", 404
    return send_from_directory(RUTA_ASSETS, nombre_archivo)

@app.route('/IMAGENES/<path:nombre_archivo>')
def servir_imagenes(nombre_archivo):
    ruta_completa = os.path.join(RUTA_IMAGENES, nombre_archivo)
    if not os.path.exists(ruta_completa):
        return f"Archivo no encontrado: {nombre_archivo}", 404
    return send_from_directory(RUTA_IMAGENES, nombre_archivo)

# ====================================================================================
# ENDPOINT DE LOGIN CON SOPORTE PARA CONTRASEÑAS ENCRIPTADAS
# ====================================================================================
@app.route('/api/login', methods=['POST'])
def login():
    """Endpoint de login mejorado compatible con el sistema EMESA existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Datos inválidos'}), 400
       
        usuario = data.get('usuario', '').strip()
        password = data.get('password', '').strip()
       
        if not usuario or not password:
            return jsonify({'success': False, 'message': 'Usuario y contraseña requeridos'}), 400
       
        # Buscar usuario en la base de datos usando la estructura EMESA
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
           
            cursor = conn.cursor()
            cursor.execute("""
                SELECT Id_Usuario, Num_Operario, Nombre, Nivel_Permisos, Roles, Contrasena
                FROM General.Usuarios
                WHERE Num_Operario = ?
            """, (usuario,))
           
            result = cursor.fetchone()
            if not result:
                return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 401
           
            # Verificar contraseña usando el sistema robusto
            stored_password = result[5]
            if not verify_password(password, stored_password):
                return jsonify({'success': False, 'message': 'Contraseña incorrecta'}), 401
           
            # Si la contraseña estaba en texto plano, migrarla a encriptada
            if not is_password_hashed(stored_password):
                new_hashed_password = hash_password(password)
                cursor.execute(
                    "UPDATE General.Usuarios SET Contrasena=? WHERE Id_Usuario=?",
                    (new_hashed_password, result[0])
                )
                conn.commit()
                print(f"✅ Contraseña migrada a hash para usuario {usuario}")
           
            # Crear sesión (mejora del sistema original)
            user_data = {
                'id': result[0],
                'num_operario': result[1],
                'nombre': result[2],
                'nivel': result[3],
                'rol': result[4]
            }
           
            session['user_id'] = result[0]
            session['user_data'] = user_data
            session.permanent = True
           
            # Respuesta compatible con el formato existente
            return jsonify({
                'success': True,
                'id': result[0],
                'num_operario': result[1],
                'nombre': result[2],
                'nivel': result[3],
                'rol': result[4]
            })
           
    except Exception as e:
        print(f"Error en /api/login: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ====================================================================================
# ENDPOINTS DE SESIÓN Y AUTENTICACIÓN
# ====================================================================================
@app.route('/api/logout', methods=['POST'])
def logout():
    """Endpoint de logout"""
    try:
        session.clear()
        return jsonify({'success': True, 'message': 'Sesión cerrada exitosamente'})
    except Exception as e:
        print(f"Error en logout: {e}")
        return jsonify({'success': False, 'message': 'Error del servidor'}), 500
 
@app.route('/api/verify_session', methods=['GET'])
def verify_session():
    """Verificar sesión actual"""
    try:
        if 'user_id' in session and 'user_data' in session:
            return jsonify({
                'success': True,
                'authenticated': True,
                'user': session['user_data']
            })
       
        return jsonify({
            'success': True,
            'authenticated': False,
            'user': None
        })
       
    except Exception as e:
        print(f"Error verificando sesión: {e}")
        return jsonify({'success': False, 'message': 'Error del servidor'}), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PERMISOS DE USUARIO
# ====================================================================================
@app.route('/api/obtener-permisos-usuario', methods=['GET'])
def obtener_permisos_usuario():
    """Obtiene los permisos del usuario autenticado"""
    try:
        if 'user_data' not in session:
            return jsonify({'success': False, 'message': 'Usuario no autenticado'}), 401
        
        user_data = session['user_data']
        nivel_permisos = user_data.get('nivel', 0)
        
        return jsonify({
            'success': True,
            'nivel_permisos': nivel_permisos,
            'user_data': user_data
        })
        
    except Exception as e:
        print(f"Error obteniendo permisos de usuario: {e}")
        return jsonify({'success': False, 'message': 'Error del servidor'}), 500

# ====================================================================================
# ENDPOINTS PARA PREFERENCIAS DE VISTA POR USUARIO
# ====================================================================================
@app.route('/api/preferencias-vista/puestos', methods=['GET'])
def get_preferencias_vista_puestos():
    """Obtiene el layout guardado de la vista Puestos para el usuario autenticado"""
    try:
        usuario_id = obtener_id_usuario_request()
        if not usuario_id:
            return jsonify({'success': False, 'message': 'Usuario no autenticado'}), 401

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                SELECT TOP 1 ConfiguracionJson, AutoAplicar, Fecha_Modificacion
                FROM [Digitalizacion].[PE].[PreferenciasVista]
                WHERE Id_Usuario = ? AND Vista = ?
            """, (usuario_id, 'puestos-layout'))

            row = cursor.fetchone()
            if not row:
                return jsonify({
                    'success': True,
                    'has_preferences': False,
                    'preferences': None
                })

            configuracion = json.loads(row[0]) if row[0] else {}

            return jsonify({
                'success': True,
                'has_preferences': True,
                'preferences': {
                    'puestos': configuracion.get('puestos', []),
                    'pautas': configuracion.get('pautas', []),
                    'auto_aplicar': bool(row[1]),
                    'fecha_modificacion': row[2].isoformat() if row[2] else None
                }
            })

    except Exception as e:
        print(f"💥 Error obteniendo preferencias de vista Puestos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/preferencias-vista/puestos', methods=['POST'])
def save_preferencias_vista_puestos():
    """Guarda o actualiza el layout preferido de la vista Puestos para el usuario autenticado"""
    try:
        usuario_id = obtener_id_usuario_request()
        if not usuario_id:
            return jsonify({'success': False, 'message': 'Usuario no autenticado'}), 401

        data = request.get_json() or {}
        puestos = data.get('puestos', [])
        pautas = data.get('pautas', [])
        auto_aplicar = bool(data.get('auto_aplicar', True))

        if not isinstance(puestos, list) or not isinstance(pautas, list):
            return jsonify({'success': False, 'message': 'Formato inválido para puestos o pautas'}), 400

        configuracion_json = json.dumps({
            'puestos': puestos,
            'pautas': pautas
        }, ensure_ascii=False)

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[PreferenciasVista]
                WHERE Id_Usuario = ? AND Vista = ?
            """, (usuario_id, 'puestos-layout'))

            existe = cursor.fetchone()[0] > 0

            if existe:
                cursor.execute("""
                    UPDATE [Digitalizacion].[PE].[PreferenciasVista]
                    SET ConfiguracionJson = ?,
                        AutoAplicar = ?,
                        Fecha_Modificacion = SYSDATETIME()
                    WHERE Id_Usuario = ? AND Vista = ?
                """, (configuracion_json, 1 if auto_aplicar else 0, usuario_id, 'puestos-layout'))
            else:
                cursor.execute("""
                    INSERT INTO [Digitalizacion].[PE].[PreferenciasVista]
                    (Id_Usuario, Vista, ConfiguracionJson, AutoAplicar)
                    VALUES (?, ?, ?, ?)
                """, (usuario_id, 'puestos-layout', configuracion_json, 1 if auto_aplicar else 0))

            return jsonify({
                'success': True,
                'message': 'Layout de Puestos guardado correctamente',
                'preferences': {
                    'puestos': puestos,
                    'pautas': pautas,
                    'auto_aplicar': auto_aplicar
                }
            })

    except Exception as e:
        print(f"💥 Error guardando preferencias de vista Puestos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/preferencias-vista/puestos', methods=['DELETE'])
def delete_preferencias_vista_puestos():
    """Elimina el layout guardado de la vista Puestos para el usuario autenticado"""
    try:
        usuario_id = obtener_id_usuario_request()
        if not usuario_id:
            return jsonify({'success': False, 'message': 'Usuario no autenticado'}), 401

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[PreferenciasVista]
                WHERE Id_Usuario = ? AND Vista = ?
            """, (usuario_id, 'puestos-layout'))

            return jsonify({
                'success': True,
                'message': 'Layout de Puestos eliminado correctamente'
            })

    except Exception as e:
        print(f"💥 Error eliminando preferencias de vista Puestos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA MOSTRAR PUESTOSCAB.HTML (PARA VISUALIZACIÓN)
# ====================================================================================
@app.route('/puestos-cab')
def mostrar_puestos_cab():
    """Mostrar PuestosCAB.html para visualización"""
    return send_from_directory(os.path.join(RUTA_TEMPLATES, "generales"), "PuestosCAB.html")

# ====================================================================================
# ENDPOINTS PARA CARGAR DATOS DE LA TABLA PE.CONTROLES
# ====================================================================================

@app.route('/api/get-columnas-controles', methods=['GET'])
def get_columnas_controles():
    """Obtiene todas las columnas disponibles de la vista PE.Columnas_Controles"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener todas las columnas de la vista PE.Columnas_Controles
            cursor.execute("""
                SELECT Columnas_Controles
                FROM [Digitalizacion].[PE].[Columnas_Controles]
                ORDER BY Columnas_Controles
            """)
            
            resultados = cursor.fetchall()
            
            # Convertir resultados a lista simple
            columnas = [row[0] for row in resultados if row[0]]  # Filtrar valores nulos si los hay
            
            print(f"📊 Columnas obtenidas: {len(columnas)} registros")
            
            return jsonify({
                'success': True,
                'columnas': columnas,
                'message': f'Se obtuvieron {len(columnas)} columnas exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo columnas controles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/create-puesto', methods=['POST'])
def create_puesto():
    """
    Crear un nuevo puesto con sus columnas configuradas.
    
    Inserta en [PE].[Puesto] (1 fila) y [PE].[Columnas_Puesto] (N filas).
    Usa transacción: rollback si falla en cualquier paso.
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_puesto = data.get('nombre', '').strip()
        columnas = data.get('columnas', [])
        puesto_compartido = data.get('puesto_compartido', 0)
        
        # Validaciones
        if not nombre_puesto:
            return jsonify({'success': False, 'message': 'El nombre del puesto es requerido'}), 400
        
        if not columnas or len(columnas) == 0:
            return jsonify({'success': False, 'message': 'Debe añadir al menos una columna'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            try:
                # 🔴 Paso 1: Verificar duplicado en tabla [PE].[Puesto]
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM [Digitalizacion].[PE].[Puesto] 
                    WHERE Nombre_Puesto = ?
                """, (nombre_puesto,))
                
                if cursor.fetchone()[0] > 0:
                    return jsonify({
                        'success': False, 
                        'message': f'Ya existe un puesto con el nombre "{nombre_puesto}"'
                    }), 409
                
                # 🔴 Paso 2: Insertar en [PE].[Puesto] y obtener ID_Puesto
                print(f"   ➕ Insertando puesto: '{nombre_puesto}' (compartido={puesto_compartido})")
                cursor.execute("""
                    INSERT INTO [Digitalizacion].[PE].[Puesto] (Nombre_Puesto, puesto_compartido)
                    OUTPUT INSERTED.ID_Puesto
                    VALUES (?, ?)
                """, (nombre_puesto, puesto_compartido))
                
                id_puesto_row = cursor.fetchone()
                if not id_puesto_row:
                    conn.rollback()
                    return jsonify({'success': False, 'message': 'Error al obtener ID del puesto'}), 500
                
                id_puesto = id_puesto_row[0]
                print(f"   ✅ Puesto creado con ID_Puesto={id_puesto}")
                
                # 🔴 Paso 3: Insertar columnas en [PE].[Columnas_Puesto]
                columnas_insertadas = 0
                for idx, columna in enumerate(columnas):
                    original_name = columna.get('originalName', '').strip()
                    display_name = columna.get('displayName', original_name).strip()
                    order = columna.get('order', 0)
                    
                    if not original_name:  # Solo si hay nombre original
                        continue
                    
                    if not display_name:
                        display_name = original_name
                    
                    print(f"   📌 Columna {idx+1}: '{original_name}' → '{display_name}' (orden={order})")
                    
                    cursor.execute("""
                        INSERT INTO [Digitalizacion].[PE].[Columnas_Puesto] 
                        (ID_Puesto, Columna, Nombre_Columna, Orden_Columna)
                        VALUES (?, ?, ?, ?)
                    """, (id_puesto, original_name, display_name, order))
                    
                    columnas_insertadas += 1
                
                # Confirmar transacción
                conn.commit()
                print(f"✅ Puesto '{nombre_puesto}' creado: {columnas_insertadas} columnas (ID={id_puesto})")
                
                return jsonify({
                    'success': True,
                    'message': f'Puesto "{nombre_puesto}" creado exitosamente con {columnas_insertadas} columnas',
                    'nombre_puesto': nombre_puesto,
                    'columnas_insertadas': columnas_insertadas,
                    'puesto_compartido': puesto_compartido,
                    'id_puesto': id_puesto
                })
                
            except Exception as e:
                conn.rollback()
                error_msg = str(e).lower()
                
                # Detectar violación de constraint único por nombre
                if 'unique' in error_msg and 'nombre_puesto' in error_msg:
                    return jsonify({
                        'success': False,
                        'message': f'Ya existe un puesto con el nombre "{nombre_puesto}"'
                    }), 409
                
                # Detectar violación de constraint único en columnas
                if 'unique' in error_msg and 'columna' in error_msg:
                    return jsonify({
                        'success': False,
                        'message': 'Una o más columnas están duplicadas en este puesto'
                    }), 409
                
                print(f"💥 Error en transacción: {e}")
                import traceback
                traceback.print_exc()
                
                return jsonify({
                    'success': False,
                    'message': f'Error del servidor: {str(e)}'
                }), 500
            
    except Exception as e:
        print(f"💥 Error creando puesto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-puestos-configurados', methods=['GET'])
def get_puestos_configurados():
    """Obtiene todos los puestos configurados con sus columnas"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener todos los registros de la tabla PE.Puestos
            cursor.execute("""
                SELECT ID_Puesto, Nombre_Puesto, Columna, Nombre_Columna, Orden_Columna
                FROM [Digitalizacion].[PE].[Puestos]
                ORDER BY Nombre_Puesto, Orden_Columna
            """)
            
            resultados = cursor.fetchall()
            
            # Convertir resultados a lista de diccionarios
            puestos = []
            for row in resultados:
                puestos.append({
                    'id_puesto': row[0],
                    'nombre_puesto': row[1],
                    'columna': row[2],
                    'nombre_columna': row[3],
                    'orden_columna': row[4]
                })
            
            print(f"📊 Puestos configurados obtenidos: {len(puestos)} registros")
            
            return jsonify({
                'success': True,
                'puestos': puestos,
                'message': f'Se obtuvieron {len(puestos)} registros exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo puestos configurados: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-puestos-config', methods=['GET'])
def get_puestos_config():
    """Obtiene lista de puestos únicos que tienen configuración"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener puestos únicos
            cursor.execute("""
                SELECT DISTINCT Nombre_Puesto
                FROM [Digitalizacion].[PE].[Puestos]
                ORDER BY Nombre_Puesto
            """)
            
            resultados = cursor.fetchall()
            
            # Convertir resultados a lista
            puestos = [row[0] for row in resultados]
            
            print(f"📋 Puestos únicos obtenidos para edición: {len(puestos)}")
            
            return jsonify({
                'success': True,
                'puestos': [{'nombre_puesto': p} for p in puestos]
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo puestos para edición: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-puesto-config/<nombre_puesto>', methods=['GET'])
def get_puesto_config(nombre_puesto):
    """Obtiene la configuración completa de un puesto específico"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener configuración del puesto incluyendo puesto_compartido
            cursor.execute("""
                SELECT ID_Puesto, Nombre_Puesto, Columna, Nombre_Columna, Orden_Columna, puesto_compartido
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE Nombre_Puesto = ?
                ORDER BY Orden_Columna
            """, (nombre_puesto,))
            
            resultados = cursor.fetchall()
            
            if not resultados:
                return jsonify({
                    'success': False,
                    'message': f'No se encontró configuración para el puesto "{nombre_puesto}"'
                }), 404
            
            # Usar el ID_Puesto y puesto_compartido del primer resultado
            id_puesto = resultados[0][0]
            puesto_compartido = resultados[0][5]  # 🆕 Obtener puesto_compartido
            
            # Convertir resultados a estructura de columnas
            columnas = []
            for row in resultados:
                columnas.append({
                    'nombre_columna': row[2],  # Columna
                    'nombre_columna_display': row[3],  # Nombre_Columna
                    'orden_columna': row[4]  # Orden_Columna
                })
            
            print(f"✅ Configuración del puesto '{nombre_puesto}' obtenida:")
            print(f"   📋 Columnas: {len(columnas)}")
            print(f"   🆕 puesto_compartido: {puesto_compartido}")
            
            return jsonify({
                'success': True,
                'puesto': {
                    'id_puesto': id_puesto,
                    'nombre_puesto': nombre_puesto,
                    'puesto_compartido': puesto_compartido,  # 🆕 Incluir en respuesta
                    'columnas': columnas
                }
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo configuración del puesto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-puesto/<int:id_puesto>', methods=['PUT'])
def update_puesto(id_puesto):
    """
    Actualiza un puesto completo y todas sus columnas manteniendo el ID_Puesto.
    
    Usa transacción: UPDATE [PE].[Puesto] y DELETE/INSERT [PE].[Columnas_Puesto].
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_puesto_nuevo = data.get('nombre', '').strip()
        columnas = data.get('columnas', [])
        puesto_compartido = data.get('puesto_compartido', 0)
        
        print(f"\n🔄 UPDATE PUESTO REQUEST")
        print(f"   📍 ID: {id_puesto}")
        print(f"   📝 Nombre nuevo: {nombre_puesto_nuevo}")
        print(f"   📊 puesto_compartido: {puesto_compartido}")
        print(f"   📋 Columnas: {len(columnas)}")
        
        if not nombre_puesto_nuevo:
            return jsonify({'success': False, 'message': 'El nombre del puesto es requerido'}), 400
        
        if not columnas or len(columnas) == 0:
            return jsonify({'success': False, 'message': 'Debe añadir al menos una columna'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            try:
                # 🔴 Paso 1: Obtener nombre actual desde [PE].[Puesto]
                cursor.execute("""
                    SELECT Nombre_Puesto FROM [Digitalizacion].[PE].[Puesto] WHERE ID_Puesto = ?
                """, (id_puesto,))
                
                resultado = cursor.fetchone()
                if not resultado:
                    return jsonify({'success': False, 'message': 'Puesto no encontrado'}), 404
                
                nombre_puesto_actual = resultado[0]
                print(f"   ✅ Puesto encontrado: {nombre_puesto_actual}")
                
                # 🔴 Paso 2: Actualizar [PE].[Puesto]
                print(f"   🔄 Actualizando nombre y compartido en [PE].[Puesto]")
                cursor.execute("""
                    UPDATE [Digitalizacion].[PE].[Puesto]
                    SET Nombre_Puesto = ?, puesto_compartido = ?
                    WHERE ID_Puesto = ?
                """, (nombre_puesto_nuevo, puesto_compartido, id_puesto))
                
                # 🔴 Paso 3: Eliminar todas las columnas antiguas
                print(f"   🗑️  Eliminando columnas antiguas")
                cursor.execute("""
                    DELETE FROM [Digitalizacion].[PE].[Columnas_Puesto]
                    WHERE ID_Puesto = ?
                """, (id_puesto,))
                
                columnas_eliminadas = cursor.rowcount
                print(f"   🗑️  {columnas_eliminadas} columnas eliminadas")
                
                # 🔴 Paso 4: Insertar nuevas columnas
                print(f"   ➕ Insertando {len(columnas)} nuevas columnas")
                columnas_insertadas = 0
                
                for i, columna in enumerate(columnas):
                    original_name = columna.get('originalName', '').strip()
                    display_name = columna.get('displayName', original_name).strip()
                    order = columna.get('order', 0)
                    
                    if not original_name:
                        continue
                    
                    if not display_name:
                        display_name = original_name
                    
                    print(f"   📌 Columna {i+1}: '{original_name}' → '{display_name}'")
                    
                    cursor.execute("""
                        INSERT INTO [Digitalizacion].[PE].[Columnas_Puesto]
                        (ID_Puesto, Columna, Nombre_Columna, Orden_Columna)
                        VALUES (?, ?, ?, ?)
                    """, (id_puesto, original_name, display_name, order))
                    
                    columnas_insertadas += 1
                
                # Confirmar transacción
                conn.commit()
                
                print(f"\n✅ RESULTADO:")
                print(f"   ID_Puesto: {id_puesto} (MANTENIDO)")
                print(f"   Nombre: '{nombre_puesto_actual}' → '{nombre_puesto_nuevo}'")
                print(f"   Columnas eliminadas: {columnas_eliminadas}")
                print(f"   Columnas insertadas: {columnas_insertadas}")
                print(f"   puesto_compartido: {puesto_compartido}")
                
                return jsonify({
                    'success': True,
                    'message': f'Puesto "{nombre_puesto_actual}" actualizado a "{nombre_puesto_nuevo}" con {columnas_insertadas} columnas',
                    'id_puesto': id_puesto,
                    'columnas_actualizadas': columnas_insertadas,
                    'puesto_compartido': puesto_compartido
                })
                
            except Exception as e:
                conn.rollback()
                print(f"💥 Error en transacción: {e}")
                import traceback
                traceback.print_exc()
                
                return jsonify({
                    'success': False,
                    'message': f'Error del servidor: {str(e)}'
                }), 500
            
    except Exception as e:
        print(f"💥 Error actualizando puesto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/delete-puesto/<int:id_puesto>', methods=['DELETE'])
def delete_puesto(id_puesto):
    """
    Elimina un puesto completo (eliminará en cascada las columnas).
    
    DELETE FROM [PE].[Puesto] eliminará automáticamente las filas de [PE].[Columnas_Puesto]
    si la FK tiene ON DELETE CASCADE.
    """
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            try:
                # 🔴 Paso 1: Obtener nombre del puesto desde [PE].[Puesto]
                cursor.execute("""
                    SELECT Nombre_Puesto FROM [Digitalizacion].[PE].[Puesto] WHERE ID_Puesto = ?
                """, (id_puesto,))
                
                resultado = cursor.fetchone()
                if not resultado:
                    return jsonify({'success': False, 'message': 'Puesto no encontrado'}), 404
                
                nombre_puesto = resultado[0]
                print(f"   ✅ Puesto encontrado: {nombre_puesto}")
                
                # 🔴 Paso 2: Eliminar columnas manualmente desde [PE].[Columnas_Puesto]
                print(f"   🗑️  Eliminando columnas desde [PE].[Columnas_Puesto]")
                cursor.execute("""
                    DELETE FROM [Digitalizacion].[PE].[Columnas_Puesto]
                    WHERE ID_Puesto = ?
                """, (id_puesto,))
                
                columnas_eliminadas = cursor.rowcount
                print(f"   🗑️  {columnas_eliminadas} columnas eliminadas")
                
                # 🔴 Paso 3: Eliminar puesto desde [PE].[Puesto]
                print(f"   🗑️  Eliminando puesto desde [PE].[Puesto]")
                cursor.execute("""
                    DELETE FROM [Digitalizacion].[PE].[Puesto]
                    WHERE ID_Puesto = ?
                """, (id_puesto,))
                
                conn.commit()
                
                print(f"🗑️ Puesto '{nombre_puesto}' eliminado completamente ({columnas_eliminadas} columnas)")
                
                return jsonify({
                    'success': True,
                    'message': f'Puesto "{nombre_puesto}" eliminado exitosamente con {columnas_eliminadas} columnas'
                })
                
            except Exception as e:
                conn.rollback()
                print(f"💥 Error en transacción: {e}")
                import traceback
                traceback.print_exc()
                
                return jsonify({
                    'success': False,
                    'message': f'Error del servidor: {str(e)}'
                }), 500
            
    except Exception as e:
        print(f"💥 Error eliminando puesto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-registro-puesto/<int:id_puesto>', methods=['PUT'])
def update_registro_puesto(id_puesto):
    """
    Actualiza un registro específico de columna (Nombre_Columna y Orden_Columna).
    
    id_puesto aquí es realmente el ID de la columna en [PE].[Columnas_Puesto].
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_columna = data.get('nombre_columna', '').strip()
        orden_columna = data.get('orden_columna')
        
        if not nombre_columna:
            return jsonify({'success': False, 'message': 'El nombre de la columna es requerido'}), 400
        
        if orden_columna is None or orden_columna < 0:
            return jsonify({'success': False, 'message': 'El orden de la columna debe ser >= 0'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que la columna existe en [PE].[Columnas_Puesto]
            cursor.execute("""
                SELECT COUNT(*) 
                FROM [Digitalizacion].[PE].[Columnas_Puesto] 
                WHERE ID_Columna_Puesto = ?
            """, (id_puesto,))
            
            if cursor.fetchone()[0] == 0:
                return jsonify({'success': False, 'message': 'Columna no encontrada'}), 404
            
            # Actualizar en [PE].[Columnas_Puesto]
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Columnas_Puesto]
                SET Nombre_Columna = ?, Orden_Columna = ?
                WHERE ID_Columna_Puesto = ?
            """, (nombre_columna, orden_columna, id_puesto))
            
            conn.commit()
            
            print(f"✅ Columna {id_puesto} actualizada: '{nombre_columna}', orden {orden_columna}")
            
            return jsonify({
                'success': True,
                'message': 'Columna actualizada exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error actualizando columna: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/delete-registro-puesto/<int:id_puesto>', methods=['DELETE'])
def delete_registro_puesto(id_puesto):
    """
    Elimina un registro específico de columna.
    
    id_puesto aquí es realmente el ID de la columna en [PE].[Columnas_Puesto].
    """
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que la columna existe en [PE].[Columnas_Puesto]
            cursor.execute("""
                SELECT ID_Puesto, Columna, Nombre_Columna
                FROM [Digitalizacion].[PE].[Columnas_Puesto] 
                WHERE ID_Columna_Puesto = ?
            """, (id_puesto,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Columna no encontrada'}), 404
            
            id_puesto_fk, columna, nombre_columna = resultado
            
            # Obtener nombre del puesto para el log
            cursor.execute("""
                SELECT Nombre_Puesto FROM [Digitalizacion].[PE].[Puesto]
                WHERE ID_Puesto = ?
            """, (id_puesto_fk,))
            
            puesto_row = cursor.fetchone()
            nombre_puesto = puesto_row[0] if puesto_row else "DESCONOCIDO"
            
            # Eliminar la columna desde [PE].[Columnas_Puesto]
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Columnas_Puesto]
                WHERE ID_Columna_Puesto = ?
            """, (id_puesto,))
            
            conn.commit()
            
            print(f"🗑️ Columna eliminada: {nombre_puesto} - {nombre_columna}")
            
            return jsonify({
                'success': True,
                'message': f'Columna eliminada exitosamente: {nombre_columna} de {nombre_puesto}'
            })
            
    except Exception as e:
        print(f"💥 Error eliminando columna: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# 🆕 ENDPOINT PARA GUARDAR TOGGLE "COMPARTIR PUESTO" (PASO 3)
# ====================================================================================
@app.route('/api/puesto/<nombre_puesto>/compartido', methods=['PUT'])
def actualizar_puesto_compartido(nombre_puesto):
    """
    Actualiza el toggle 'puesto_compartido' en la tabla [PE].[Puesto].
    
    Valida automáticamente que exista un control común de tipo "Listado Armarios"
    entre 2+ pautas antes de permitir activar el toggle.
    
    Body esperado:
    {
        "puesto_compartido": 1,  // 0 para desactivar, 1 para activar
        "validar": true          // Si es true, valida primero
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        puesto_compartido = data.get('puesto_compartido', 0)
        validar = data.get('validar', True)
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a BD'}), 500
            
            cursor = conn.cursor()
            
            # 🔴 Obtener ID del puesto desde [PE].[Puesto]
            cursor.execute("""
                SELECT [ID_Puesto]
                FROM [Digitalizacion].[PE].[Puesto]
                WHERE [Nombre_Puesto] = ?
            """, (nombre_puesto,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Puesto no encontrado'}), 404
            
            id_puesto = resultado[0]
            
            print(f"\n🔄 ACTUALIZANDO TOGGLE PUESTO_COMPARTIDO")
            print(f"   📍 Puesto: {nombre_puesto} (ID: {id_puesto})")
            print(f"   ✅/❌ Compartido: {puesto_compartido}")
            
            # Si se desactiva (compartido=0), no necesita validación
            if puesto_compartido == 0:
                cursor.execute("""
                    UPDATE [Digitalizacion].[PE].[Puesto]
                    SET [puesto_compartido] = 0
                    WHERE [ID_Puesto] = ?
                """, (id_puesto,))
                conn.commit()
                
                print(f"   ✅ Toggle desactivado")
                return jsonify({
                    'success': True,
                    'compartible': False,
                    'puesto_compartido': 0,
                    'nombre_puesto': nombre_puesto,
                    'message': f'Puesto {nombre_puesto} ya NO es compartido'
                }), 200
            
            # Si se activa (compartido=1) Y validar=true, validar antes
            if puesto_compartido == 1 and validar:
                print(f"   🔍 Validando compatibilidad de pautas...")
                
                # Obtener pautas del puesto usando ID_Puesto con JOIN a PE.Controles
                cursor.execute("""
                    SELECT DISTINCT p.[Nombre_Pauta]
                    FROM [Digitalizacion].[PE].[Pautas] p
                    INNER JOIN [Digitalizacion].[PE].[Controles] c ON p.[ID_Control] = c.[Id_Control]
                    WHERE c.[ID_Puesto] = ? AND p.[Activo] = 1
                    ORDER BY p.[Nombre_Pauta]
                """, (id_puesto,))
                
                pautas = [row[0] for row in cursor.fetchall()]
                
                if len(pautas) < 2:
                    return jsonify({
                        'success': False,
                        'compartible': False,
                        'message': f'No se puede activar: Se necesitan 2+ pautas. Encontradas: {len(pautas)}'
                    }), 400
                
                # Obtener controles "Listado Armarios" por pauta
                pautas_con_listado = {}
                for pauta in pautas:
                    cursor.execute("""
                        SELECT DISTINCT [Id_Control]
                        FROM [Digitalizacion].[PE].[Checklist]
                        WHERE [Puesto] = ? 
                        AND [Nombre_Pauta] = ? 
                        AND [TipoReg] = 'Listado Armarios'
                    """, (nombre_puesto, pauta))
                    
                    controles = cursor.fetchall()
                    if controles:
                        pautas_con_listado[pauta] = controles[0][0]
                
                if len(pautas_con_listado) < 2:
                    return jsonify({
                        'success': False,
                        'compartible': False,
                        'message': f'No se puede activar: No hay suficientes pautas con "Listado Armarios"'
                    }), 400
                
                # Detectar control compartido
                controles_por_id = {}
                for pauta, id_control in pautas_con_listado.items():
                    if id_control not in controles_por_id:
                        controles_por_id[id_control] = []
                    controles_por_id[id_control].append(pauta)
                
                controles_compartidos = {
                    id_ctrl: pautas_list 
                    for id_ctrl, pautas_list in controles_por_id.items() 
                    if len(pautas_list) >= 2
                }
                
                if not controles_compartidos:
                    return jsonify({
                        'success': False,
                        'compartible': False,
                        'pautas': pautas,
                        'pautas_con_listado': list(pautas_con_listado.keys()),
                        'message': 'No se puede activar: Las pautas no comparten el MISMO control "Listado Armarios"'
                    }), 400
                
                pautas_agrupadas = list(controles_compartidos.values())[0]
                print(f"   ✅ Validación exitosa: {len(pautas_agrupadas)} pautas pueden agruparse")
            
            # 🔴 Actualizar el toggle en [PE].[Puesto]
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Puesto]
                SET [puesto_compartido] = ?
                WHERE [ID_Puesto] = ?
            """, (puesto_compartido, id_puesto))
            
            conn.commit()
            
            print(f"   ✅ Toggle activado para {nombre_puesto}")
            
            return jsonify({
                'success': True,
                'compartible': True,
                'puesto_compartido': 1,
                'nombre_puesto': nombre_puesto,
                'pautas': pautas_agrupadas if 'pautas_agrupadas' in locals() else pautas,
                'message': f'✅ Puesto {nombre_puesto} ahora es COMPARTIDO'
            }), 200
            
    except Exception as e:
        print(f"💥 Error actualizando puesto_compartido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# 🆕 ENDPOINT PARA SUBIDA DE FOTOS
# ====================================================================================
@app.route('/api/upload_photo', methods=['POST'])
def upload_photo():
    """Recibe una foto y la guarda en la carpeta compartida"""
    try:
        if 'photo' not in request.files:
            return jsonify({'success': False, 'message': 'No se recibió ningún archivo'}), 400
        
        file = request.files['photo']
        control_id = request.form.get('controlId', 'unknown')
        id_pedido = (request.form.get('idPedido') or '').strip()
        armario = (request.form.get('armario') or '').strip()
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Nombre de archivo vacío'}), 400

        if file:
            def limpiar_segmento_nombre(valor, fallback):
                valor_limpio = re.sub(r'[^A-Za-z0-9_-]+', '-', str(valor or '').strip())
                valor_limpio = re.sub(r'-{2,}', '-', valor_limpio).strip('-_')
                return valor_limpio or fallback

            if id_pedido and not armario:
                try:
                    with ConexionODBC('Digitalizacion') as conn:
                        if conn:
                            cursor = conn.cursor()
                            cursor.execute("""
                                SELECT TOP 1 [Armario]
                                FROM [Digitalizacion].[PE].[Pedido]
                                WHERE [ID_Pedido] = ?
                            """, (id_pedido,))
                            pedido_row = cursor.fetchone()
                            if pedido_row:
                                if not armario:
                                    armario = (pedido_row[0] or '').strip()
                except Exception as lookup_error:
                    print(f"⚠️ No se pudo completar Armario para foto del pedido {id_pedido}: {lookup_error}")

            # Generar nombre único: Fecha_Hora_ControlID_Armario.jpg
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{timestamp}_Control_{control_id}_{limpiar_segmento_nombre(armario, 'SIN_ARMARIO')}.jpg"
            
            # Asegurar que el directorio existe
            if not os.path.exists(RUTA_FOTOS_COMPARTIDA):
                try:
                    os.makedirs(RUTA_FOTOS_COMPARTIDA)
                    print(f"✅ Carpeta de fotos creada: {RUTA_FOTOS_COMPARTIDA}")
                except Exception as e:
                    print(f"❌ Error creando carpeta de fotos: {e}")
                    return jsonify({'success': False, 'message': f'Error creando carpeta de destino: {str(e)}'}), 500
            
            # Ruta completa
            filepath = os.path.join(RUTA_FOTOS_COMPARTIDA, filename)
            
            # Guardar archivo
            file.save(filepath)
            print(f"📸 Foto guardada: {filepath}")
            
            return jsonify({
                'success': True, 
                'message': 'Foto guardada correctamente',
                'filename': filename,
                'path': filepath
            })
            
    except Exception as e:
        print(f"💥 Error subiendo foto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/save-controles', methods=['POST'])
def save_controles():
    """Guarda los controles temporales en la tabla PE.Controles usando ID_Puesto (FK)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        controles = data.get('controles', [])
        
        if not controles or len(controles) == 0:
            return jsonify({'success': False, 'message': 'No hay controles para guardar'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Insertar cada control en la base de datos
            controles_insertados = 0
            errores = []
            
            for control in controles:
                try:
                    # El frontend ahora envía ID_Puesto (int) en lugar de nombre del puesto
                    id_puesto = control.get('id_puesto')
                    punto_inspeccion = control.get('puntoInspeccion', '').strip()
                    proceso = control.get('proceso', '').strip()
                    carac_inspeccion = control.get('caracInspeccion', '').strip()
                    tipo_reg = control.get('tipoReg', '').strip()
                    descripcion_control = control.get('descripcionControl', '').strip()
                    metodo = control.get('metodo', '').strip()
                    ruta_foto = control.get('rutaFoto', '').strip()
                    visible_pdf = control.get('visiblePDF', 1)
                    
                    # Validar campos obligatorios
                    if not all([id_puesto, punto_inspeccion, proceso, carac_inspeccion, tipo_reg, descripcion_control]):
                        errores.append(f"Control con ID_Puesto '{id_puesto}' tiene campos obligatorios vacíos")
                        continue
                    
                    # Insertar el control en [PE].[Controles] con FK ID_Puesto
                    cursor.execute("""
                        INSERT INTO [Digitalizacion].[PE].[Controles] 
                        (ID_Puesto, PuntoInspección, Proceso, CaracInspeccion, TipoReg, DescripcionControl, Metodo, Ruta_foto_mostrar, VisiblePDF)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (id_puesto, punto_inspeccion, proceso, carac_inspeccion, tipo_reg, descripcion_control, metodo, ruta_foto, visible_pdf))
                    
                    controles_insertados += 1
                    
                except Exception as e:
                    errores.append(f"Error insertando control con ID_Puesto '{control.get('id_puesto', 'N/A')}': {str(e)}")
                    continue
            
            if controles_insertados > 0:
                # Confirmar la transacción solo si hay insertados exitosos
                conn.commit()
                print(f"✅ {controles_insertados} controles guardados en PE.Controles")
            
            # Preparar respuesta
            if controles_insertados == len(controles):
                return jsonify({
                    'success': True,
                    'message': f'Todos los controles guardados exitosamente ({controles_insertados} controles)',
                    'controles_insertados': controles_insertados
                })
            elif controles_insertados > 0:
                return jsonify({
                    'success': True,
                    'message': f'{controles_insertados} de {len(controles)} controles guardados. Algunos tuvieron errores.',
                    'controles_insertados': controles_insertados,
                    'errores': errores
                }), 206  # Partial Content
            else:
                return jsonify({
                    'success': False,
                    'message': 'No se pudieron guardar controles',
                    'errores': errores
                }), 400
            
    except Exception as e:
        print(f"💥 Error guardando controles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-puestos-controles', methods=['GET'])
def get_puestos_controles():
    """Obtiene ID_Puesto y Nombre_Puesto desde PE.Puesto para el dropdown de controles (nuevo modelo relacional)"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # 🆕 CAMBIO: Consultar directamente PE.Puesto en lugar de PE.Puestos
            # Retornar tanto ID como Nombre para usar como FK
            cursor.execute("""
                SELECT ID_Puesto, Nombre_Puesto
                FROM [Digitalizacion].[PE].[Puesto]
                WHERE Nombre_Puesto IS NOT NULL AND Nombre_Puesto != ''
                ORDER BY Nombre_Puesto
            """)
            
            resultados = cursor.fetchall()
            
            # Convertir resultados a lista de diccionarios con id y nombre
            puestos = []
            for row in resultados:
                puestos.append({
                    'id': row[0],           # ID_Puesto (entero)
                    'nombre': row[1]        # Nombre_Puesto (string)
                })
            
            print(f"🏭 Puestos únicos obtenidos desde PE.Puesto: {len(puestos)} registros")
            for p in puestos:
                print(f"   📍 ID={p['id']}, Nombre={p['nombre']}")
            
            return jsonify({
                'success': True,
                'puestos': puestos,
                'message': f'Se obtuvieron {len(puestos)} puestos únicos exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo puestos para controles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-controles-existentes', methods=['GET'])
def get_controles_existentes():
    """Obtiene todos los controles existentes con LEFT JOIN a PE.Puesto para manejar puestos NULL"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # 🆕 LEFT JOIN para ver TODOS los controles, incluso sin puesto asignado
            # Renombrar PuntoInspección a PuntoInspeccion para evitar problemas de encoding con pyodbc
            cursor.execute("""
                SELECT c.Id_Control, 
                       c.ID_Puesto,
                       p.Nombre_Puesto,
                       c.PuntoInspección AS PuntoInspeccion, 
                       c.Proceso, 
                       c.CaracInspeccion, 
                       c.TipoReg, 
                       c.DescripcionControl, 
                       c.Metodo, 
                       c.Ruta_foto_mostrar,
                       c.VisiblePDF
                FROM [Digitalizacion].[PE].[Controles] c
                LEFT JOIN [Digitalizacion].[PE].[Puesto] p ON c.ID_Puesto = p.ID_Puesto
                ORDER BY ISNULL(p.Nombre_Puesto, 'Sin Puesto'), c.PuntoInspección
            """)
            
            resultados = cursor.fetchall()
            print(f"📋 Query ejecutada, filas obtenidas: {len(resultados)}")
            
            # Convertir resultados a lista de diccionarios
            controles = []
            for row in resultados:
                # 🆕 Manejo robusto de nulos con validaciones
                id_control = row[0] if row[0] is not None else 0
                id_puesto = row[1] if row[1] is not None else None
                nombre_puesto = row[2] if row[2] is not None else 'Sin Puesto'
                punto_inspeccion = row[3] if row[3] is not None else ''
                proceso = row[4] if row[4] is not None else ''
                carac_inspeccion = row[5] if row[5] is not None else ''
                tipo_reg = row[6] if row[6] is not None else ''
                descripcion_control = row[7] if row[7] is not None else ''
                metodo = row[8] if row[8] is not None else ''
                ruta_foto = row[9] if row[9] is not None else ''
                visible_pdf = row[10] if row[10] is not None else 1
                
                controles.append({
                    'Id_Control': id_control,
                    'ID_Puesto': id_puesto,  # 🆕 FK numérica
                    'nombre_puesto': nombre_puesto,  # Nombre para referencia interna
                    'puesto': nombre_puesto,  # 🆕 Mantener para compatibilidad con frontend (HTML tabla)
                    'puntoInspeccion': punto_inspeccion,
                    'proceso': proceso,
                    'caracInspeccion': carac_inspeccion,
                    'tipoReg': tipo_reg,
                    'descripcionControl': descripcion_control,
                    'metodo': metodo,
                    'rutaFoto': ruta_foto,
                    'VisiblePDF': bool(visible_pdf)
                })
                
                print(f"   ✅ Control {id_control}: Puesto={nombre_puesto}, Punto={punto_inspeccion}")
            
            print(f"🔧 Controles existentes obtenidos: {len(controles)} registros")
            
            return jsonify({
                'success': True,
                'controles': controles,
                'message': f'Se obtuvieron {len(controles)} controles exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo controles existentes: {e}")
        import traceback
        traceback.print_exc()  # 🆕 Mostrar stack trace completo en terminal
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-control/<int:control_id>', methods=['PUT'])
def update_control(control_id):
    """Actualiza un control específico usando ID_Puesto (FK)"""
    try:
        print(f"🔄 Iniciando actualización de control ID: {control_id}")
        
        data = request.get_json()
        if not data:
            print("❌ No se recibieron datos JSON")
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        print(f"📤 Datos recibidos: {data}")
        
        id_puesto = data.get('id_puesto')
        punto_inspeccion = data.get('puntoInspeccion', '').strip()
        proceso = data.get('proceso', '').strip()
        carac_inspeccion = data.get('caracInspeccion', '').strip()
        tipo_reg = data.get('tipoReg', '').strip()
        descripcion_control = data.get('descripcionControl', '').strip()
        metodo = data.get('metodo', '').strip()
        ruta_foto = data.get('rutaFoto', '').strip()
        visible_pdf = data.get('visiblePDF', 1)
        
        print(f"📋 Campos procesados:")
        print(f"   - id_puesto: {id_puesto}")
        print(f"   - punto_inspeccion: '{punto_inspeccion}'")
        print(f"   - proceso: '{proceso}'")
        print(f"   - carac_inspeccion: '{carac_inspeccion}'")
        print(f"   - tipo_reg: '{tipo_reg}'")
        print(f"   - descripcion_control: '{descripcion_control}'")
        print(f"   - visiblePDF: {visible_pdf}")
        
        # Validar campos obligatorios
        if not all([id_puesto, punto_inspeccion, proceso, carac_inspeccion, tipo_reg, descripcion_control]):
            print("❌ Faltan campos obligatorios")
            return jsonify({'success': False, 'message': 'Todos los campos obligatorios son requeridos'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ Error de conexión a base de datos")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            print(f"🔍 Verificando existencia del control ID: {control_id}")
            
            # Verificar que el control existe
            cursor.execute("""
                SELECT COUNT(*) 
                FROM [Digitalizacion].[PE].[Controles] 
                WHERE Id_Control = ?
            """, (control_id,))
            
            count = cursor.fetchone()[0]
            print(f"📊 Controles encontrados con ID {control_id}: {count}")
            
            if count == 0:
                print(f"❌ Control con ID {control_id} no encontrado")
                return jsonify({'success': False, 'message': 'Control no encontrado'}), 404
            
            print("🔄 Ejecutando actualización...")
            
            # Actualizar el control usando Id_Control en WHERE y ID_Puesto como FK
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Controles]
                SET ID_Puesto = ?, PuntoInspección = ?, Proceso = ?, CaracInspeccion = ?,
                    TipoReg = ?, DescripcionControl = ?, Metodo = ?, Ruta_foto_mostrar = ?, VisiblePDF = ?
                WHERE Id_Control = ?
            """, (id_puesto, punto_inspeccion, proceso, carac_inspeccion, tipo_reg, 
                  descripcion_control, metodo, ruta_foto, visible_pdf, control_id))
            
            conn.commit()
            
            print(f"✅ Control {control_id} actualizado exitosamente")
            print(f"   ID_Puesto: {id_puesto} - Punto: {punto_inspeccion}")
            
            return jsonify({
                'success': True,
                'message': 'Control actualizado exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error actualizando control {control_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-control-visible-pdf/<int:control_id>', methods=['PUT'])
def update_control_visible_pdf(control_id):
    """Actualiza el campo VisiblePDF de un control"""
    try:
        print(f"🔄 Actualizando VisiblePDF para control ID: {control_id}")
        
        data = request.get_json()
        if not data:
            print("❌ No se recibieron datos JSON")
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        visible_pdf = data.get('VisiblePDF', 1)
        print(f"📤 VisiblePDF: {visible_pdf}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ Error de conexión a base de datos")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el control existe
            cursor.execute("""
                SELECT COUNT(*) 
                FROM [Digitalizacion].[PE].[Controles] 
                WHERE Id_Control = ?
            """, (control_id,))
            
            count = cursor.fetchone()[0]
            if count == 0:
                print(f"❌ Control con ID {control_id} no encontrado")
                return jsonify({'success': False, 'message': 'Control no encontrado'}), 404
            
            print("🔄 Ejecutando actualización de VisiblePDF...")
            
            # Actualizar el campo VisiblePDF
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Controles]
                SET VisiblePDF = ?
                WHERE Id_Control = ?
            """, (visible_pdf, control_id))
            
            conn.commit()
            
            estado = 'visible' if visible_pdf else 'oculto'
            print(f"✅ Control {control_id} actualizado: VisiblePDF={estado}")
            
            return jsonify({
                'success': True,
                'message': f'Control actualizado: ahora {estado} en PDF',
                'VisiblePDF': bool(visible_pdf)
            })
            
    except Exception as e:
        print(f"💥 Error actualizando VisiblePDF para control {control_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/delete-control/<int:control_id>', methods=['DELETE'])
def delete_control(control_id):
    """Elimina un control específico"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el control existe y obtener info para log (ahora con JOIN a PE.Puesto)
            cursor.execute("""
                SELECT p.Nombre_Puesto, c.PuntoInspección 
                FROM [Digitalizacion].[PE].[Controles] c
                INNER JOIN [Digitalizacion].[PE].[Puesto] p ON c.ID_Puesto = p.ID_Puesto
                WHERE c.Id_Control = ?
            """, (control_id,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Control no encontrado'}), 404
            
            nombre_puesto, punto_inspeccion = resultado
            
            # Eliminar el control
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Controles]
                WHERE Id_Control = ?
            """, (control_id,))
            
            conn.commit()
            
            print(f"🗑️ Control eliminado: {nombre_puesto} - {punto_inspeccion}")
            
            return jsonify({
                'success': True,
                'message': f'Control eliminado exitosamente: {punto_inspeccion} de {nombre_puesto}'
            })
            
    except Exception as e:
        print(f"💥 Error eliminando control: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/duplicate-control/<int:control_id>', methods=['POST'])
def duplicate_control(control_id):
    """Duplica un control existente"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos del control original
            cursor.execute("""
                SELECT ID_Puesto, PuntoInspección, Proceso, CaracInspeccion, TipoReg, DescripcionControl, Metodo, Ruta_foto_mostrar, VisiblePDF
                FROM [Digitalizacion].[PE].[Controles]
                WHERE Id_Control = ?
            """, (control_id,))
            
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Control no encontrado'}), 404
            
            # Insertar nuevo control duplicado
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Controles] 
                (ID_Puesto, PuntoInspección, Proceso, CaracInspeccion, TipoReg, DescripcionControl, Metodo, Ruta_foto_mostrar, VisiblePDF)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (row[0], row[1], row[2], row[3], row[4], row[5] + " (Copia)", row[6], row[7], row[8]))
            
            conn.commit()
            
            print(f"✅ Control {control_id} duplicado exitosamente")
            
            return jsonify({'success': True, 'message': 'Control duplicado exitosamente'})
            
    except Exception as e:
        print(f"💥 Error duplicando control: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/get-tipos-reg-controles', methods=['GET'])
def get_tipos_reg_controles():
    """Obtiene los tipos de registro únicos de la tabla PE.Controles para el dropdown"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener tipos de registro únicos
            cursor.execute("""
                SELECT DISTINCT TipoReg
                FROM [Digitalizacion].[PE].[Controles]
                WHERE TipoReg IS NOT NULL AND TipoReg != ''
                ORDER BY TipoReg
            """)
            
            resultados = cursor.fetchall()
            
            # Convertir resultados a lista simple
            tipos_reg = [row[0] for row in resultados if row[0]]  # Filtrar valores nulos si los hay
            
            print(f"📝 Tipos de registro únicos obtenidos: {len(tipos_reg)} registros")
            print(f"📋 Tipos: {tipos_reg}")
            
            return jsonify({
                'success': True,
                'tipos_reg': tipos_reg,
                'message': f'Se obtuvieron {len(tipos_reg)} tipos de registro únicos exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo tipos de registro: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINTS PARA GESTIÓN DE PAUTAS
# ====================================================================================

@app.route('/api/create-pauta', methods=['POST'])
def create_pauta():
    """Crea una nueva pauta con sus controles asociados"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_pauta = data.get('nombre', '').strip()
        controles = data.get('controles', [])
        activo = data.get('activo', 1)  # Por defecto, nueva pauta es activa
        
        print(f"🔄 Creando pauta: {nombre_pauta}")
        print(f"📋 Controles a asociar: {len(controles)}")
        print(f"🔄 Estado (Activo): {activo}")
        
        if not nombre_pauta:
            return jsonify({'success': False, 'message': 'El nombre de la pauta es requerido'}), 400
        
        if not controles or len(controles) == 0:
            return jsonify({'success': False, 'message': 'Debe añadir al menos un control a la pauta'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Generar un nuevo ID_Pauta (obtener el máximo + 1)
            cursor.execute("""
                SELECT ISNULL(MAX(ID_Pauta), 0) + 1 
                FROM [Digitalizacion].[PE].[Pautas]
            """)
            
            nuevo_id_pauta = cursor.fetchone()[0]
            print(f"🆔 Nuevo ID_Pauta: {nuevo_id_pauta}")
            
            # Insertar cada control como un registro separado
            registros_insertados = 0
            for index, control in enumerate(controles):
                print(f"   🔄 Procesando control {index + 1}: {control}")
                
                id_control = control.get('Id_Control')
                orden_pauta = control.get('orden', index + 1)  # Usar orden del frontend o fallback al índice
                print(f"      - Id_Control extraído: {id_control} (tipo: {type(id_control)})")
                print(f"      - Orden_Pauta: {orden_pauta}")
                
                if id_control:
                    try:
                        cursor.execute("""
                            INSERT INTO [Digitalizacion].[PE].[Pautas] 
                            (ID_Pauta, Nombre_Pauta, ID_Control, Fecha_creacion, Orden_Pauta, Activo)
                            VALUES (?, ?, ?, GETDATE(), ?, ?)
                        """, (nuevo_id_pauta, nombre_pauta, id_control, orden_pauta, 1 if activo else 0))
                        
                        registros_insertados += 1
                        print(f"      ✅ Control {id_control} agregado a pauta con orden {orden_pauta}, Activo={activo}")
                        
                    except Exception as e:
                        print(f"      ❌ Error insertando control {id_control}: {e}")
                        continue
                else:
                    print(f"      ⚠️ Control sin Id_Control válido: {control}")
            
            print(f"📊 Resumen: {registros_insertados} de {len(controles)} controles insertados")
            
            if registros_insertados > 0:
                # Confirmar la transacción
                conn.commit()
                estado_text = 'activa' if activo else 'inactiva'
                print(f"✅ Pauta '{nombre_pauta}' creada ({estado_text}) con {registros_insertados} controles")
                
                return jsonify({
                    'success': True,
                    'message': f'Pauta "{nombre_pauta}" creada exitosamente ({estado_text}) con {registros_insertados} controles',
                    'id_pauta': nuevo_id_pauta,
                    'controles_insertados': registros_insertados,
                    'activo': bool(activo)
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'No se pudieron insertar controles en la pauta'
                }), 400
            
    except Exception as e:
        print(f"💥 Error creando pauta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/get-pautas', methods=['GET'])
def get_pautas():
    """Obtiene todas las pautas configuradas con sus controles (con JOIN a PE.Puesto para obtener nombre)"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener todas las pautas con información de controles
            # INNER JOIN con PE.Puesto para obtener el nombre del puesto
            cursor.execute("""
                SELECT p.ID_Pauta, p.Nombre_Pauta, p.ID_Control, p.Fecha_creacion, p.Fecha_modificacion,
                       ISNULL(pu.Nombre_Puesto, ''), ISNULL(c.PuntoInspección, ''), ISNULL(c.TipoReg, ''), 
                       ISNULL(c.DescripcionControl, ''), p.Orden_Pauta, p.Activo, c.ID_Puesto
                FROM [Digitalizacion].[PE].[Pautas] p
                LEFT JOIN [Digitalizacion].[PE].[Controles] c ON p.ID_Control = c.Id_Control
                LEFT JOIN [Digitalizacion].[PE].[Puesto] pu ON c.ID_Puesto = pu.ID_Puesto
                ORDER BY p.ID_Pauta, p.Orden_Pauta, pu.Nombre_Puesto, c.PuntoInspección
            """)
            
            resultados = cursor.fetchall()
            
            # Organizar resultados por pauta
            pautas = {}
            for row in resultados:
                id_pauta = row[0]
                nombre_pauta = row[1]
                id_control = row[2]
                fecha_creacion = row[3]
                fecha_modificacion = row[4]
                nombre_puesto = row[5]
                punto_inspeccion = row[6]
                tipo_reg = row[7]
                descripcion = row[8]
                orden_pauta = row[9]
                activo = row[10]
                id_puesto = row[11]
                
                # Crear pauta si no existe
                if id_pauta not in pautas:
                    pautas[id_pauta] = {
                        'id_pauta': id_pauta,
                        'nombre_pauta': nombre_pauta,
                        'fecha_creacion': fecha_creacion.isoformat() if fecha_creacion else None,
                        'fecha_modificacion': fecha_modificacion.isoformat() if fecha_modificacion else None,
                        'activo': bool(activo),
                        'controles': []
                    }
                
                # Agregar control a la pauta
                if id_control:
                    control_data = {
                        'id_control': id_control,
                        'id_puesto': id_puesto,
                        'puesto': nombre_puesto,  # Ahora obtenido del JOIN con PE.Puesto
                        'punto_inspeccion': punto_inspeccion,
                        'tipo_reg': tipo_reg,
                        'descripcionControl': descripcion,
                        'orden': orden_pauta
                    }
                    print(f"🔍 Agregando control a pauta: {control_data}")
                    pautas[id_pauta]['controles'].append(control_data)
            
            # Convertir a lista
            lista_pautas = list(pautas.values())
            
            print(f"📊 Pautas obtenidas: {len(lista_pautas)}")
            
            return jsonify({
                'success': True,
                'pautas': lista_pautas,
                'message': f'Se obtuvieron {len(lista_pautas)} pautas exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo pautas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-pauta/<int:pauta_id>', methods=['PUT'])
def update_pauta(pauta_id):
    """Actualiza una pauta existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_pauta = data.get('nombre', '').strip()
        controles = data.get('controles', [])
        
        print(f"🔄 Actualizando pauta ID: {pauta_id}")
        print(f"📝 Nuevo nombre: {nombre_pauta}")
        print(f"📋 Controles nuevos: {len(controles)}")
        
        if not nombre_pauta:
            return jsonify({'success': False, 'message': 'El nombre de la pauta es requerido'}), 400
        
        if not controles or len(controles) == 0:
            return jsonify({'success': False, 'message': 'Debe añadir al menos un control a la pauta'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que la pauta existe
            cursor.execute("""
                SELECT COUNT(*) 
                FROM [Digitalizacion].[PE].[Pautas] 
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            if cursor.fetchone()[0] == 0:
                return jsonify({'success': False, 'message': 'Pauta no encontrada'}), 404
            
            # Eliminar todos los registros existentes de esta pauta
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Pautas]
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            print(f"🗑️ Registros anteriores de pauta {pauta_id} eliminados")
            
            # Insertar los nuevos controles
            registros_insertados = 0
            for index, control in enumerate(controles):
                id_control = control.get('Id_Control')
                orden_pauta = control.get('orden', index + 1)  # Usar orden del frontend o fallback al índice
                print(f"   🔄 Procesando control para actualización: {control}")
                print(f"      - Id_Control: {id_control}")
                print(f"      - Orden_Pauta: {orden_pauta}")
                
                if id_control:
                    try:
                        cursor.execute("""
                            INSERT INTO [Digitalizacion].[PE].[Pautas] 
                            (ID_Pauta, Nombre_Pauta, ID_Control, Fecha_creacion, Fecha_modificacion, Orden_Pauta)
                            VALUES (?, ?, ?, GETDATE(), GETDATE(), ?)
                        """, (pauta_id, nombre_pauta, id_control, orden_pauta))
                        
                        registros_insertados += 1
                        print(f"   ✅ Control {id_control} agregado a pauta actualizada con orden {orden_pauta}")
                        
                    except Exception as e:
                        print(f"   ❌ Error insertando control {id_control}: {e}")
                        continue
            
            if registros_insertados > 0:
                # Confirmar la transacción
                conn.commit()
                print(f"✅ Pauta {pauta_id} '{nombre_pauta}' actualizada con {registros_insertados} controles")
                
                return jsonify({
                    'success': True,
                    'message': f'Pauta "{nombre_pauta}" actualizada exitosamente con {registros_insertados} controles',
                    'controles_insertados': registros_insertados
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'No se pudieron insertar controles en la pauta actualizada'
                }), 400
            
    except Exception as e:
        print(f"💥 Error actualizando pauta {pauta_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/delete-pauta/<int:pauta_id>', methods=['DELETE'])
def delete_pauta(pauta_id):
    """Elimina una pauta y todos sus controles asociados"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que la pauta existe y obtener info para log
            cursor.execute("""
                SELECT DISTINCT Nombre_Pauta 
                FROM [Digitalizacion].[PE].[Pautas] 
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Pauta no encontrada'}), 404
            
            nombre_pauta = resultado[0]
            
            # Eliminar todos los registros de la pauta
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Pautas]
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            conn.commit()
            
            print(f"🗑️ Pauta eliminada: {nombre_pauta} (ID: {pauta_id})")
            
            return jsonify({
                'success': True,
                'message': f'Pauta "{nombre_pauta}" eliminada exitosamente'
            })
            
    except Exception as e:
        print(f"💥 Error eliminando pauta {pauta_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/update-pauta-activo/<int:pauta_id>', methods=['PUT'])
def update_pauta_activo(pauta_id):
    """Actualiza el estado Activo de una pauta (todos sus registros)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        activo = data.get('activo', 0)
        
        print(f"🔄 Actualizando estado Activo de pauta ID: {pauta_id} a {activo}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que la pauta existe y obtener su nombre
            cursor.execute("""
                SELECT DISTINCT Nombre_Pauta 
                FROM [Digitalizacion].[PE].[Pautas] 
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Pauta no encontrada'}), 404
            
            nombre_pauta = resultado[0]
            
            # Actualizar el estado Activo para TODOS los registros de esta pauta
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Pautas]
                SET Activo = ?
                WHERE ID_Pauta = ?
            """, (1 if activo else 0, pauta_id))
            
            registros_actualizados = cursor.rowcount
            conn.commit()
            
            estado = 'activada' if activo else 'desactivada'
            print(f"✅ Pauta {nombre_pauta} (ID: {pauta_id}) {estado}. Registros actualizados: {registros_actualizados}")
            
            return jsonify({
                'success': True,
                'message': f'Pauta "{nombre_pauta}" {estado} correctamente',
                'registros_actualizados': registros_actualizados,
                'activo': bool(activo)
            })
            
    except Exception as e:
        print(f"💥 Error actualizando estado de pauta {pauta_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINTS PARA GESTIÓN DE REFERENCIAS DE PAUTAS 🆕
# ====================================================================================

@app.route('/api/save-pauta-referencias/<int:pauta_id>', methods=['POST'])
def save_pauta_referencias(pauta_id):
    """Guarda las referencias de una pauta en la tabla PE.Pauta_Referencias"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

        referencias = data.get('referencias', [])
        
        print(f"💾 Guardando {len(referencias)} referencias para pauta ID: {pauta_id}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()

            # ✅ IMPORTANTE: Primero verificar que la pauta existe
            cursor.execute("""
                SELECT ID_Pauta FROM [Digitalizacion].[PE].[Pautas]
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Pauta no encontrada'}), 404

            # Eliminar referencias antiguas de esta pauta
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Pauta_Referencias]
                WHERE ID_Pauta = ?
            """, (pauta_id,))
            
            print(f"🗑️ Referencias anteriores eliminadas para pauta {pauta_id}")

            # Insertar nuevas referencias
            referencias_insertadas = 0
            for ref_data in referencias:
                referencia = ref_data.get('referencia', '').strip()
                
                if not referencia:
                    continue
                
                cursor.execute("""
                    INSERT INTO [Digitalizacion].[PE].[Pauta_Referencias] 
                    (ID_Pauta, Referencia, Fecha_creacion, Fecha_modificacion)
                    VALUES (?, ?, GETDATE(), GETDATE())
                """, (pauta_id, referencia))
                
                referencias_insertadas += 1
                print(f"  ✅ Referencia insertada: {referencia}")

            conn.commit()

            print(f"📊 {referencias_insertadas} referencias guardadas para pauta {pauta_id}")

            return jsonify({
                'success': True,
                'message': f'{referencias_insertadas} referencias guardadas correctamente',
                'referencias_guardadas': referencias_insertadas
            }), 201

    except Exception as e:
        print(f"💥 Error guardando referencias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/get-pauta-referencias/<int:pauta_id>', methods=['GET'])
def get_pauta_referencias(pauta_id):
    """Obtiene todas las referencias de una pauta específica"""
    try:
        print(f"🔍 Obteniendo referencias para pauta ID: {pauta_id}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()

            # Obtener referencias de la pauta
            cursor.execute("""
                SELECT ID_Referencia, Referencia, Fecha_creacion, Fecha_modificacion
                FROM [Digitalizacion].[PE].[Pauta_Referencias]
                WHERE ID_Pauta = ?
                ORDER BY Fecha_creacion DESC
            """, (pauta_id,))

            resultados = cursor.fetchall()
            
            referencias = []
            for row in resultados:
                referencias.append({
                    'id': row[0],
                    'referencia': row[1],
                    'fecha_creacion': row[2].isoformat() if row[2] else None,
                    'fecha_modificacion': row[3].isoformat() if row[3] else None
                })

            print(f"✅ {len(referencias)} referencias obtenidas para pauta {pauta_id}")

            return jsonify({
                'success': True,
                'pauta_id': pauta_id,
                'referencias': referencias,
                'total': len(referencias)
            }), 200

    except Exception as e:
        print(f"💥 Error obteniendo referencias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PUESTOS Y PAUTAS PARA VISTA ÁRBOL
# ====================================================================================

@app.route('/api/get-puesto-pauta', methods=['GET'])
def get_puesto_pauta():
    """Obtiene la estructura de puestos y pautas desde PE.Puesto_Pauta para vista tipo árbol
    Solo muestra pautas con Activo = 1
    
    🆕 LÓGICA DE PUESTOS COMPARTIDOS:
    Si un puesto tiene puesto_compartido = 1 y 2+ pautas comparten el MISMO control "Listado Armarios":
    - Mostrar UNA sola tarjeta que agrupe todas esas pautas
    - La tarjeta mostrará: "Pauta1, Pauta2" (todas juntas)
    
    Si puesto_compartido = 0 (default):
    - Mostrar una tarjeta por cada pauta (comportamiento actual)
    """
    request_started_at = time.perf_counter()
    print(f"\n🔍 [GET /api/get-puesto-pauta] Iniciando obtención de puestos y pautas...")
    
    try:
        print("📡 Estableciendo conexión a base de datos...")
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ Error: No se pudo establecer conexión a base de datos")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            print("✅ Conexión establecida correctamente")
            cursor = conn.cursor()
            try:
                cursor.timeout = 10
            except Exception:
                pass
            
            print("🗄️ Ejecutando consulta SQL...")
            # Obtener datos de puestos y pautas SOLO donde Activo = 1
            cursor.execute("""
                SELECT DISTINCT pp.[Puesto], pp.[Nombre_Pauta], pp.[Num_Pautas]
                FROM [Digitalizacion].[PE].[Puesto_Pauta] pp
                INNER JOIN [Digitalizacion].[PE].[Pautas] p ON pp.[Nombre_Pauta] = p.[Nombre_Pauta]
                WHERE p.[Activo] = 1
                ORDER BY pp.[Puesto], pp.[Nombre_Pauta]
            """)
            
            resultados = cursor.fetchall()
            print(f"\n🔍 Resultados obtenidos de PE.Puesto_Pauta (solo Activo=1): {len(resultados)} registros en {time.perf_counter() - request_started_at:.2f}s")
            
            # Organizar datos por puesto
            puestos = {}
            for row in resultados:
                puesto = row[0]
                nombre_pauta = row[1]
                num_pautas = row[2]
                
                print(f"   📍 Procesando: {puesto} -> {nombre_pauta} ({num_pautas} controles) [ACTIVO]")
                
                # Crear puesto si no existe
                if puesto not in puestos:
                    puestos[puesto] = {
                        'nombre_puesto': puesto,
                        'pautas': [],
                        'puesto_compartido': 0  # Default
                    }
                
                # Agregar pauta al puesto
                puestos[puesto]['pautas'].append({
                    'nombre_pauta': nombre_pauta,
                    'num_pautas': num_pautas
                })
            
            # 🆕 Ahora, para cada puesto, verificar si es compartido
            for nombre_puesto in puestos:
                # Obtener valor de puesto_compartido desde tabla Puestos
                cursor.execute("""
                    SELECT DISTINCT TOP 1 puesto_compartido
                    FROM [Digitalizacion].[PE].[Puestos]
                    WHERE Nombre_Puesto = ?
                """, (nombre_puesto,))
                
                resultado_compartido = cursor.fetchone()
                if resultado_compartido:
                    puesto_compartido_flag = resultado_compartido[0]
                    puestos[nombre_puesto]['puesto_compartido'] = puesto_compartido_flag
                    
                    print(f"\n   🔎 Puesto '{nombre_puesto}':")
                    print(f"      puesto_compartido = {puesto_compartido_flag}")
                    
                    # Si está marcado como compartido, validar si realmente puede agruparse
                    if puesto_compartido_flag == 1:
                        print(f"      ✅ MARCADO COMO COMPARTIDO - Validando...")
                        
                        # Obtener pautas del puesto
                        pautas_del_puesto = [p['nombre_pauta'] for p in puestos[nombre_puesto]['pautas']]
                        
                        if len(pautas_del_puesto) >= 2:
                            print(f"      📋 Pautas: {pautas_del_puesto}")
                            
                            # Obtener controles "Listado Armarios" por pauta
                            pautas_con_listado = {}
                            for pauta in pautas_del_puesto:
                                # Usar tabla Checklist que vincula Puesto-Pauta-Control
                                cursor.execute("""
                                    SELECT DISTINCT [Id_Control]
                                    FROM [Digitalizacion].[PE].[Checklist]
                                    WHERE [Puesto] = ? 
                                    AND [Nombre_Pauta] = ? 
                                    AND [TipoReg] = 'Listado Armarios'
                                """, (nombre_puesto, pauta))
                                
                                controles = cursor.fetchall()
                                if controles:
                                    pautas_con_listado[pauta] = controles[0][0]
                            
                            if len(pautas_con_listado) >= 2:
                                # Detectar si comparten control
                                controles_por_id = {}
                                for pauta, id_control in pautas_con_listado.items():
                                    if id_control not in controles_por_id:
                                        controles_por_id[id_control] = []
                                    controles_por_id[id_control].append(pauta)
                                
                                controles_compartidos = {
                                    id_ctrl: pautas_list 
                                    for id_ctrl, pautas_list in controles_por_id.items() 
                                    if len(pautas_list) >= 2
                                }
                                
                                if controles_compartidos:
                                    # 🆕 AGRUPAR PAUTAS COMPARTIDAS EN UNA SOLA TARJETA
                                    print(f"      ✅ SE AGRUPAN - Controles compartidos: {controles_compartidos}")
                                    
                                    # Obtener pautas que comparten el control
                                    pautas_agrupadas = list(controles_compartidos.values())[0]
                                    
                                    # Crear una tarjeta única con todas las pautas
                                    num_controles_listado = puestos[nombre_puesto]['pautas'][0]['num_pautas']
                                    
                                    puestos[nombre_puesto]['pautas'] = [{
                                        'nombre_pauta': ', '.join(pautas_agrupadas),  # Mostrar todas juntas
                                        'num_pautas': num_controles_listado,
                                        'es_agrupada': True,
                                        'pautas_reales': pautas_agrupadas  # Guardar las pautas reales para uso interno
                                    }]
                                    
                                    print(f"      📊 Nueva estructura: 1 tarjeta agrupada con {len(pautas_agrupadas)} pautas")
                                else:
                                    print(f"      ❌ No hay controles compartidos - Manteniendo tarjetas separadas")
                            else:
                                print(f"      ❌ Insuficientes pautas con 'Listado Armarios' - Manteniendo tarjetas separadas")
                        else:
                            print(f"      ❌ Insuficientes pautas (<2) - Manteniendo tarjetas separadas")
            
            # Convertir a lista ordenada y eliminar puestos sin pautas activas
            lista_puestos = list(puestos.values())
            lista_puestos.sort(key=lambda x: x['nombre_puesto'])
            
            print(f"\n📊 Estructura final: {len(lista_puestos)} puestos con pautas activas")
            for p in lista_puestos:
                print(f"   📍 {p['nombre_puesto']}: {len(p['pautas'])} tarjeta(s)")
                for pauta in p['pautas']:
                    print(f"      - {pauta['nombre_pauta']}")
            
            response_data = {
                'success': True,
                'puestos': lista_puestos,
                'message': f'Se obtuvieron {len(lista_puestos)} puestos con sus pautas activas exitosamente'
            }
            print(f"✅ [GET /api/get-puesto-pauta] Enviando respuesta exitosa en {time.perf_counter() - request_started_at:.2f}s")
            
            return jsonify(response_data)
            
    except Exception as e:
        print(f"💥 Error obteniendo puesto-pauta tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINT PARA OBTENER CHECKLIST DE CONTROLES POR PAUTA
# ====================================================================================

@app.route('/api/get-checklist/<puesto>/<pauta>', methods=['GET'])
def get_checklist(puesto, pauta):
    """Obtiene el checklist de controles para un puesto y pauta específicos
    Solo si la pauta está activa (Activo = 1)"""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            print(f"🔍 Obteniendo checklist para Puesto: {puesto}, Pauta: {pauta}")
            
            # Verificar que la pauta está activa
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Pautas]
                WHERE [Nombre_Pauta] = ? AND [Activo] = 1
            """, (pauta,))
            
            if cursor.fetchone()[0] == 0:
                return jsonify({
                    'success': False,
                    'message': f'La pauta "{pauta}" no está disponible (inactiva)'
                }), 403
            
            # Obtener las columnas configuradas para este puesto
            cursor.execute("""
                SELECT [Columna], [Nombre_Columna], [Orden_Columna]
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE [Nombre_Puesto] = ?
                ORDER BY [Orden_Columna]
            """, (puesto,))
            
            columnas_config = cursor.fetchall()
            print(f"📋 Columnas configuradas para {puesto}: {len(columnas_config)}")
            
            if not columnas_config:
                return jsonify({
                    'success': False,
                    'message': f'No se encontraron columnas configuradas para el puesto {puesto}'
                }), 404
            
            # Obtener los controles del checklist para esta pauta con ISNULL para evitar valores NULL
            # Incluir VisiblePDF desde la tabla Controles
            cursor.execute("""
                SELECT ch.[Id_Control], 
                       ISNULL(ch.[Puesto], ''), 
                       ISNULL(ch.[PuntoInspección], ''), 
                       ISNULL(ch.[Proceso], ''), 
                       ISNULL(ch.[CaracInspeccion], ''), 
                       ISNULL(ch.[TipoReg], ''), 
                       ISNULL(ch.[DescripcionControl], ''), 
                       ISNULL(ch.[Metodo], ''), 
                       ISNULL(ch.[Ruta_foto_mostrar], ''), 
                       ch.[ID_Pauta], 
                       ISNULL(ch.[Nombre_Pauta], ''), 
                       ch.[Fecha_creacion], 
                       ch.[Fecha_modificacion], 
                       ch.[Orden_Pauta],
                       ISNULL(ctrl.[VisiblePDF], 1)
                FROM [Digitalizacion].[PE].[Checklist] ch
                LEFT JOIN [Digitalizacion].[PE].[Controles] ctrl ON ch.[Id_Control] = ctrl.[Id_Control]
                WHERE ch.[Puesto] = ? AND ch.[Nombre_Pauta] = ?
                ORDER BY ch.[Orden_Pauta]
            """, (puesto, pauta))
            
            controles = cursor.fetchall()
            print(f"🎯 Controles encontrados: {len(controles)}")
            
            # Formatear columnas
            columnas = []
            for row in columnas_config:
                columnas.append({
                    'campo': row[0],
                    'titulo': row[1],
                    'orden': row[2]
                })
            
            # Formatear controles
            controles_lista = []
            for row in controles:
                control = {
                    'Id_Control': row[0],
                    'Puesto': row[1],
                    'PuntoInspección': row[2],
                    'Proceso': row[3],
                    'CaracInspeccion': row[4],
                    'TipoReg': row[5],
                    'DescripcionControl': row[6],
                    'Metodo': row[7],
                    'Ruta_foto_mostrar': row[8],
                    'ID_Pauta': row[9],
                    'Nombre_Pauta': row[10],
                    'Fecha_creacion': row[11].isoformat() if row[11] else None,
                    'Fecha_modificacion': row[12].isoformat() if row[12] else None,
                    'Orden_Pauta': row[13],
                    'VisiblePDF': bool(row[14])  # ← Agregar VisiblePDF
                }
                controles_lista.append(control)
            
            return jsonify({
                'success': True,
                'puesto': puesto,
                'pauta': pauta,
                'columnas': columnas,
                'controles': controles_lista,
                'message': f'Checklist cargado: {len(controles_lista)} controles con {len(columnas)} columnas'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo checklist: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINTS PARA GESTIÓN DE PEDIDOS/ARMARIOS
# ====================================================================================

def _normalizar_cantidad_cache(valor):
    if valor is None:
        return None

    try:
        cantidad_float = float(valor)
        return int(cantidad_float) if cantidad_float.is_integer() else cantidad_float
    except Exception:
        return valor


def _obtener_info_cache_num_pedido(cursor, num_pedido):
    num_pedido_limpio = str(num_pedido or '').strip()

    if not num_pedido_limpio:
        return {
            'exists': False,
            'num_pedido': '',
            'cantidad': None,
            'semana_entrega': ''
        }

    cursor.execute("""
        SELECT TOP 1 [NumPedido], [Cantidad], [SEMANAENTREGA]
        FROM [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache]
        WHERE LTRIM(RTRIM([NumPedido])) = ?
    """, (num_pedido_limpio,))

    row = cursor.fetchone()
    if not row:
        return {
            'exists': False,
            'num_pedido': num_pedido_limpio,
            'cantidad': None,
            'semana_entrega': ''
        }

    return {
        'exists': True,
        'num_pedido': str(row[0]).strip() if row[0] is not None else num_pedido_limpio,
        'cantidad': _normalizar_cantidad_cache(row[1]),
        'semana_entrega': str(row[2]).strip() if row[2] is not None else ''
    }


def _obtener_posicion_num_pedido(cursor, num_pedido, pedido_id):
    num_pedido_limpio = str(num_pedido or '').strip()

    if not num_pedido_limpio or not pedido_id:
        return {
            'NumPedidoConPosicion': num_pedido_limpio,
            'TotalNumPedido': 0 if not num_pedido_limpio else 1,
            'PosicionNumPedido': 0 if not num_pedido_limpio else 1
        }

    cursor.execute("""
        WITH pedidos_ranking AS (
            SELECT 
                [ID_Pedido],
                ROW_NUMBER() OVER (ORDER BY [ID_Pedido] ASC) as posicion,
                COUNT(*) OVER () as total
            FROM [Digitalizacion].[PE].[Pedido]
            WHERE [NumPedido] = ?
        )
        SELECT posicion, total
        FROM pedidos_ranking
        WHERE [ID_Pedido] = ?
    """, (num_pedido_limpio, pedido_id))

    resultado = cursor.fetchone()
    if resultado:
        posicion = resultado[0]
        total = resultado[1]
        return {
            'NumPedidoConPosicion': f"{num_pedido_limpio}/{posicion}",
            'TotalNumPedido': total,
            'PosicionNumPedido': posicion
        }

    return {
        'NumPedidoConPosicion': num_pedido_limpio,
        'TotalNumPedido': 1,
        'PosicionNumPedido': 1
    }


def _enriquecer_pedido_num_pedido(cursor, pedido):
    num_pedido = str(pedido.get('NumPedido') or '').strip()
    pedido_id = pedido.get('ID_Pedido') or pedido.get('id_pedido') or pedido.get('id')

    pedido.update(_obtener_posicion_num_pedido(cursor, num_pedido, pedido_id))

    cache_info = _obtener_info_cache_num_pedido(cursor, num_pedido)
    pedido['CantidadPedido'] = cache_info['cantidad']
    pedido['SemanaEntregaPedido'] = cache_info['semana_entrega']

    return pedido

@app.route('/api/pedido/<armario>', methods=['GET'])
def get_pedido_por_armario(armario):
    """Busca un pedido por número de armario (opcionalmente filtra por pauta)"""
    try:
        # 🆕 Obtener parámetro opcional nombre_pauta
        nombre_pauta = request.args.get('nombre_pauta', None)
        
        # 🆕 Limpiar armario: trim + conversión a string
        armario_limpio = str(armario).strip()
        nombre_pauta_limpio = nombre_pauta.strip() if nombre_pauta else None
        
        print(f"🔍 Buscando pedido por armario: '{armario_limpio}' (original: '{armario}')")
        if nombre_pauta_limpio:
            print(f"   - Filtro por pauta: '{nombre_pauta_limpio}'")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # 🆕 Buscar pedido por armario (y opcionalmente por pauta)
            # Usar CAST/CONVERT para asegurar comparación correcta de tipos
            if nombre_pauta_limpio:
                print(f"   📋 Ejecutando query CON filtro de pauta...")
                cursor.execute("""
                    SELECT [ID_Pedido], [Armario], [Fecha], [Referencia], [Comentarios], [Nombre_Pauta], [Cerrado], [NumPedido]
                    FROM [Digitalizacion].[PE].[Pedido]
                    WHERE CAST(TRIM([Armario]) AS VARCHAR(MAX)) = CAST(? AS VARCHAR(MAX)) AND [Nombre_Pauta] = ?
                """, (armario_limpio, nombre_pauta_limpio))
            else:
                print(f"   📋 Ejecutando query SIN filtro de pauta...")
                cursor.execute("""
                    SELECT [ID_Pedido], [Armario], [Fecha], [Referencia], [Comentarios], [Nombre_Pauta], [Cerrado], [NumPedido]
                    FROM [Digitalizacion].[PE].[Pedido]
                    WHERE CAST(TRIM([Armario]) AS VARCHAR(MAX)) = CAST(? AS VARCHAR(MAX))
                """, (armario_limpio,))
            
            resultado = cursor.fetchone()
            
            if resultado:
                pedido = {
                    'ID_Pedido': resultado[0],
                    'Armario': resultado[1],
                    'Fecha': resultado[2].isoformat() if resultado[2] else None,
                    'Referencia': resultado[3],
                    'Comentarios': resultado[4],
                    'Nombre_Pauta': resultado[5],
                    'Cerrado': resultado[6],
                    'NumPedido': resultado[7] if resultado[7] else ''
                }
                _enriquecer_pedido_num_pedido(cursor, pedido)
                
                print(f"✅ Pedido encontrado: ID={pedido['ID_Pedido']}, Armario={pedido['Armario']}, Pauta={pedido['Nombre_Pauta']}")
                return jsonify(pedido)
            else:
                print(f"❌ Pedido no encontrado para armario: '{armario_limpio}'")
                # 🆕 Debug: mostrar qué armarios existen en la BD
                cursor.execute("""
                    SELECT TOP (10) TRIM([Armario]), [Nombre_Pauta], [ID_Pedido]
                    FROM [Digitalizacion].[PE].[Pedido]
                    ORDER BY [ID_Pedido] DESC
                """)
                debug_armarios = cursor.fetchall()
                if debug_armarios:
                    print(f"   📊 Últimos armarios en BD:")
                    for arm in debug_armarios:
                        print(f"      - Armario: '{arm[0]}', Pauta: '{arm[1]}', ID: {arm[2]}")
                
                return jsonify({'message': 'Pedido no encontrado'}), 404
                
    except Exception as e:
        print(f"💥 Error buscando pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/pedido', methods=['POST'])
def crear_pedido():
    """Crea un nuevo pedido"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        armario = data.get('Armario', '').strip()
        fecha = data.get('Fecha', '').strip()
        referencia = data.get('Referencia', '').strip()
        comentarios = data.get('Comentarios', '').strip()
        nombre_pauta = data.get('Nombre_Pauta', '').strip()  # 🆕 Agregar pauta
        
        print(f"📝 Creando nuevo pedido:")
        print(f"   - Armario: {armario}")
        print(f"   - Fecha: {fecha}")
        print(f"   - Referencia: {referencia}")
        print(f"   - Nombre_Pauta: {nombre_pauta}")
        
        if not armario or not referencia:
            return jsonify({'success': False, 'message': 'Armario y Referencia son obligatorios'}), 400
        
        if not nombre_pauta:
            return jsonify({'success': False, 'message': 'Nombre_Pauta es obligatorio'}), 400
        
        # Si no se proporciona fecha, usar fecha actual
        if not fecha:
            from datetime import datetime
            fecha = datetime.now().strftime('%Y-%m-%d')
        
        # Validar y convertir formato de fecha para SQL Server (datetime)
        try:
            from datetime import datetime
            # Convertir string a objeto datetime con hora 00:00:00
            fecha_datetime = datetime.strptime(fecha, '%Y-%m-%d')
            print(f"📅 Fecha procesada: {fecha} -> {fecha_datetime} (tipo: {type(fecha_datetime)})")
        except ValueError as e:
            print(f"❌ Error en formato de fecha: {e}")
            return jsonify({'success': False, 'message': f'Formato de fecha inválido: {fecha}'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # ❌ VALIDACIÓN CRÍTICA: El número de armario debe ser ÚNICO (sin importar la pauta)
            # 🆕 Usar CAST para asegurar comparación correcta de tipos
            cursor.execute("""
                SELECT COUNT(*), [Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE CAST(TRIM([Armario]) AS VARCHAR(MAX)) = CAST(? AS VARCHAR(MAX))
                GROUP BY [Nombre_Pauta]
            """, (armario,))
            
            resultado_validacion = cursor.fetchone()
            if resultado_validacion:
                pauta_existente = resultado_validacion[1]
                return jsonify({
                    'success': False, 
                    'message': f'El armario "{armario}" ya existe para la pauta "{pauta_existente}". El número de armario no se puede repetir.'
                }), 409
            
            # Insertar nuevo pedido con Nombre_Pauta
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Pedido] 
                ([Armario], [Fecha], [Referencia], [Comentarios], [Nombre_Pauta], [Cerrado])
                OUTPUT INSERTED.[ID_Pedido], INSERTED.[Armario], INSERTED.[Fecha], INSERTED.[Referencia], INSERTED.[Comentarios], INSERTED.[Nombre_Pauta]
                VALUES (?, ?, ?, ?, ?, 0)
            """, (armario, fecha_datetime, referencia, comentarios, nombre_pauta))
            
            resultado = cursor.fetchone()
            conn.commit()
            
            pedido_creado = {
                'ID_Pedido': resultado[0],
                'Armario': resultado[1],
                'Fecha': resultado[2].isoformat() if resultado[2] else None,
                'Referencia': resultado[3],
                'Comentarios': resultado[4],
                'Nombre_Pauta': resultado[5],
                'Cerrado': 0,
                'NumPedido': ''
            }
            _enriquecer_pedido_num_pedido(cursor, pedido_creado)
            
            print(f"✅ Pedido creado exitosamente: {pedido_creado}")
            
            return jsonify(pedido_creado), 201
            
    except Exception as e:
        print(f"💥 Error creando pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

@app.route('/api/pedido/<int:pedido_id>', methods=['PUT'])
def actualizar_pedido(pedido_id):
    """Actualiza la referencia de un pedido existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nueva_referencia = data.get('Referencia', '').strip()
        
        print(f"🔄 Actualizando pedido ID {pedido_id}")
        print(f"   - Nueva referencia: {nueva_referencia}")
        
        if not nueva_referencia:
            return jsonify({'success': False, 'message': 'La referencia es obligatoria'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el pedido existe
            cursor.execute("""
                SELECT [ID_Pedido], [Armario], [Fecha], [Referencia], [Comentarios], [Nombre_Pauta], [NumPedido]
                FROM [Digitalizacion].[PE].[pedido]
                WHERE [ID_Pedido] = ?
            """, (pedido_id,))
            
            pedido_actual = cursor.fetchone()
            if not pedido_actual:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            # Actualizar la referencia
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[pedido]
                SET [Referencia] = ?
                WHERE [ID_Pedido] = ?
            """, (nueva_referencia, pedido_id))
            
            conn.commit()
            
            # Retornar el pedido actualizado
            pedido_actualizado = {
                'ID_Pedido': pedido_actual[0],
                'Armario': pedido_actual[1],
                'Fecha': pedido_actual[2].isoformat() if pedido_actual[2] else None,
                'Referencia': nueva_referencia,  # Nueva referencia
                'Comentarios': pedido_actual[4],
                'Nombre_Pauta': pedido_actual[5],
                'NumPedido': pedido_actual[6] if pedido_actual[6] else ''
            }
            _enriquecer_pedido_num_pedido(cursor, pedido_actualizado)
            
            print(f"✅ Pedido {pedido_id} actualizado exitosamente")
            
            return jsonify(pedido_actualizado)
            
    except Exception as e:
        print(f"💥 Error actualizando pedido {pedido_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINT PARA ACTUALIZAR NOMBRE_PAUTA EN UN PEDIDO
# ====================================================================================
@app.route('/api/pedido/<int:pedido_id>/pauta', methods=['PATCH'])
def actualizar_pauta_pedido(pedido_id):
    """Actualiza el campo Nombre_Pauta de un pedido existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        nombre_pauta = data.get('Nombre_Pauta', '').strip()
        
        print(f"🔄 Actualizando pauta del pedido ID {pedido_id}")
        print(f"   - Nombre_Pauta: {nombre_pauta}")
        
        if not nombre_pauta:
            return jsonify({'success': False, 'message': 'Nombre_Pauta es obligatorio'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el pedido existe
            cursor.execute("""
                SELECT [ID_Pedido], [Armario], [Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE [ID_Pedido] = ?
            """, (pedido_id,))
            
            pedido_actual = cursor.fetchone()
            if not pedido_actual:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            pauta_anterior = pedido_actual[2]
            
            # Si ya tiene una pauta asignada y es diferente, advertir
            if pauta_anterior and pauta_anterior != nombre_pauta:
                print(f"⚠️ Cambiando pauta de '{pauta_anterior}' a '{nombre_pauta}'")
            
            # Actualizar el Nombre_Pauta
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Pedido]
                SET [Nombre_Pauta] = ?
                WHERE [ID_Pedido] = ?
            """, (nombre_pauta, pedido_id))
            
            conn.commit()
            
            print(f"✅ Pedido {pedido_id} actualizado con pauta '{nombre_pauta}'")
            
            return jsonify({
                'success': True,
                'message': f'Pauta actualizada correctamente',
                'ID_Pedido': pedido_id,
                'Nombre_Pauta': nombre_pauta,
                'Pauta_Anterior': pauta_anterior
            })
            
    except Exception as e:
        print(f"💥 Error actualizando pauta del pedido {pedido_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500

# ====================================================================================
# ENDPOINT PARA ACTUALIZAR NUMPEDIDO EN UN PEDIDO
# ====================================================================================
@app.route('/api/validar-num-pedido-cliente', methods=['GET'])
def validar_num_pedido_cliente():
    """Valida si un número de pedido existe en PE.Pedido_Cantidad_SemanaEntrega_Cache."""
    request_started_at = time.perf_counter()

    try:
        num_pedido = (request.args.get('num_pedido') or '').strip()

        if not num_pedido:
            return jsonify({
                'success': False,
                'exists': False,
                'message': 'NumPedido es obligatorio'
            }), 400

        print(f"🔍 Validando NumPedido de cliente: {num_pedido}")

        pedido_id_param = (request.args.get('pedido_id') or '').strip()

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'exists': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            cache_info = _obtener_info_cache_num_pedido(cursor, num_pedido)
            exists = cache_info['exists']

            ya_asignados = 0
            if exists and pedido_id_param and pedido_id_param.isdigit():
                try:
                    cursor.execute("""
                        SELECT COUNT(*) FROM [Digitalizacion].[PE].[Pedido]
                        WHERE LTRIM(RTRIM([NumPedido])) = ? AND [ID_Pedido] <> ?
                    """, (num_pedido, int(pedido_id_param)))
                    ya_asignados = cursor.fetchone()[0]
                except Exception:
                    ya_asignados = 0

            print(
                f"{'✅' if exists else '⚠️'} Validación NumPedido '{num_pedido}': "
                f"{'existe' if exists else 'no existe'} | ya_asignados={ya_asignados} en {time.perf_counter() - request_started_at:.2f}s"
            )

            return jsonify({
                'success': True,
                'exists': exists,
                'cantidad': cache_info['cantidad'],
                'ya_asignados': ya_asignados,
                'semana_entrega': cache_info['semana_entrega'],
                'message': 'NumPedido válido' if exists else 'El número de pedido introducido no es un pedido de cliente'
            })

    except Exception as e:
        print(f"💥 Error validando NumPedido de cliente: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'exists': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/pedido/<int:pedido_id>/numpedido', methods=['PATCH'])
def actualizar_numpedido(pedido_id):
    """Actualiza el campo NumPedido de un pedido existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        num_pedido = data.get('NumPedido', '').strip()
        permitir_no_cliente = bool(data.get('PermitirNoCliente', False))
        
        print(f"🔄 Actualizando NumPedido del pedido ID {pedido_id}")
        print(f"   - NumPedido: {num_pedido}")
        print(f"   - PermitirNoCliente: {permitir_no_cliente}")
        
        if not num_pedido:
            return jsonify({'success': False, 'message': 'NumPedido es obligatorio'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el pedido existe
            cursor.execute("""
                SELECT [ID_Pedido], [Armario], [NumPedido]
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE [ID_Pedido] = ?
            """, (pedido_id,))
            
            pedido_actual = cursor.fetchone()
            if not pedido_actual:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404

            cache_info = _obtener_info_cache_num_pedido(cursor, num_pedido)
            pedido_cliente_existe = cache_info['exists']

            # Siempre bloquear si el NumPedido no existe en la cache ERP
            if not pedido_cliente_existe:
                return jsonify({
                    'success': False,
                    'message': f'El número de pedido "{num_pedido}" no existe en el sistema ERP. Solo se pueden asignar pedidos de cliente.'
                }), 400

            # Validar capacidad: no se pueden asignar más armarios que Cantidad
            cantidad_max = cache_info.get('cantidad') or 0
            if cantidad_max > 0:
                cursor.execute("""
                    SELECT COUNT(*) FROM [Digitalizacion].[PE].[Pedido]
                    WHERE LTRIM(RTRIM([NumPedido])) = ? AND [ID_Pedido] <> ?
                """, (num_pedido, pedido_id))
                ya_asignados = cursor.fetchone()[0]
                if ya_asignados >= cantidad_max:
                    return jsonify({
                        'success': False,
                        'message': f'El pedido "{num_pedido}" ya tiene {ya_asignados}/{int(cantidad_max)} armarios asignados (máximo permitido). No se pueden añadir más.'
                    }), 400
            
            numpedido_anterior = pedido_actual[2]
            
            # Actualizar el NumPedido
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Pedido]
                SET [NumPedido] = ?
                WHERE [ID_Pedido] = ?
            """, (num_pedido, pedido_id))
            
            conn.commit()
            
            print(f"✅ Pedido {pedido_id} actualizado con NumPedido '{num_pedido}'")

            posicion_info = _obtener_posicion_num_pedido(cursor, num_pedido, pedido_id)
            
            return jsonify({
                'success': True,
                'message': f'NumPedido actualizado correctamente',
                'ID_Pedido': pedido_id,
                'NumPedido': num_pedido,
                'NumPedidoConPosicion': posicion_info['NumPedidoConPosicion'],
                'TotalNumPedido': posicion_info['TotalNumPedido'],
                'PosicionNumPedido': posicion_info['PosicionNumPedido'],
                'CantidadPedido': cache_info['cantidad'],
                'SemanaEntregaPedido': cache_info['semana_entrega'],
                'pedido_cliente_validado': True,
                'NumPedido_Anterior': numpedido_anterior
            })
            
    except Exception as e:
        print(f"💥 Error actualizando NumPedido del pedido {pedido_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/datosuser', methods=['POST'])
def guardar_datos_user():
    """Guarda los resultados del checklist en la tabla PE.DatosUser.
    Espera un JSON con { ID_Pedido: int, Puesto: str, registros: [ { ID_Control: int, Resultado: str, Comentario: str } ] }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

        id_pedido = data.get('ID_Pedido')
        puesto = data.get('Puesto', '')
        registros = data.get('registros', [])

        if not id_pedido or not isinstance(registros, list):
            return jsonify({'success': False, 'message': 'Payload inválido'}), 400

        # 🆕 Obtener el usuario autenticado de la sesión
        usuario_logueado = obtener_usuario_sesion()
        print(f"🔐 Usuario autenticado: {usuario_logueado}")

        print(f"📥 Guardando datos del checklist para pedido {id_pedido}, puesto: {puesto}. Filas: {len(registros)}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()

            # Para cada registro, verificar si existe y actualizar o insertar
            for r in registros:
                id_control = r.get('ID_Control')
                # Aceptar tanto 'Datos' como 'Resultado' para compatibilidad
                datos = r.get('Datos') or r.get('Resultado', '')
                comentario = r.get('Comentario', '')
                es_toggle = r.get('EsToggle', False)  # Indica si es OK/NOK o valor numérico/texto
                evaluacion_medicion = r.get('EvaluacionMedicion')  # OK/NOK de mediciones automáticas

                # Determinar qué campos llenar según el tipo de control
                if es_toggle:
                    # Es un toggle OK/NOK -> guardar solo en campo Resultado
                    resultado = datos  # 'OK' o 'NOK'
                    resultado_txt = None
                    tipo_guardado = "Toggle OK/NOK"
                else:
                    # Es medición o texto -> guardar en ambos campos
                    if evaluacion_medicion and (evaluacion_medicion == 'OK' or evaluacion_medicion == 'NOK'):
                        # Medición con evaluación automática -> guardar ambos
                        resultado = evaluacion_medicion  # 'OK' o 'NOK'
                        resultado_txt = datos  # Valor numérico introducido
                        tipo_guardado = f"Medición con evaluación {evaluacion_medicion}"
                    else:
                        # Texto libre sin evaluación -> solo Resultado_txt
                        resultado = None
                        resultado_txt = datos  # Valor de texto
                        tipo_guardado = "Texto libre"

                # 🆕 Limpiar comentario si el resultado es "OK"
                # Los comentarios solo tienen sentido para NOK o estados que necesitan aclaración
                comentario_final = None if resultado == 'OK' else (comentario if comentario else None)

                # Debug: mostrar datos recibidos
                print(f"  📋 Registro: ID_Control={id_control}, Tipo={tipo_guardado}, Resultado='{resultado}', Resultado_txt='{resultado_txt}', Comentario='{comentario_final}'")

                # Asegurar tipos
                try:
                    id_control_val = int(id_control) if id_control is not None else None
                except Exception:
                    id_control_val = None

                if id_control_val is None:
                    print(f"  ⚠️ ID_Control inválido, saltando registro")
                    continue

                # ⚠️ IMPORTANTE: No guardar registros sin resultado (vacíos o None)
                if not datos or datos == '':
                    print(f"  ⚠️ Datos vacíos para ID_Control={id_control_val}, saltando registro")
                    continue

                # Verificar si existe el registro
                cursor.execute("""
                    SELECT COUNT(*) FROM [Digitalizacion].[PE].[DatosUser]
                    WHERE [ID_Pedido] = ? AND [ID_Control] = ?
                """, (id_pedido, id_control_val))
                
                existe = cursor.fetchone()[0] > 0

                if existe:
                    # Actualizar registro existente
                    cursor.execute("""
                        UPDATE [Digitalizacion].[PE].[DatosUser]
                        SET [Resultado] = ?, [Resultado_txt] = ?, [Comentario] = ?, [user] = ?
                        WHERE [ID_Pedido] = ? AND [ID_Control] = ?
                    """, (resultado, resultado_txt, comentario_final, usuario_logueado, id_pedido, id_control_val))
                    print(f"  ✏️ Actualizado ID_Control={id_control_val} -> Resultado={resultado}, Resultado_txt={resultado_txt}, user={usuario_logueado}")
                else:
                    # Insertar nuevo registro
                    cursor.execute("""
                        INSERT INTO [Digitalizacion].[PE].[DatosUser] (ID_Pedido, ID_Control, Resultado, Resultado_txt, Comentario, [user])
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (id_pedido, id_control_val, resultado, resultado_txt, comentario_final, usuario_logueado))
                    print(f"  ➕ Insertado ID_Control={id_control_val} -> Resultado={resultado}, Resultado_txt={resultado_txt}, user={usuario_logueado}")

            # Actualizar el campo Cerrado en PE.Pedido SOLO si el checklist está completo
            if puesto:
                # Obtener información de la pauta para saber cuántos controles DEBE tener
                pauta = data.get('Nombre_Pauta', '')
                
                if pauta:
                    # Contar cuántos controles tiene esta pauta para este puesto
                    cursor.execute("""
                        SELECT COUNT(*) 
                        FROM [Digitalizacion].[PE].[Checklist]
                        WHERE [Puesto] = ? AND [Nombre_Pauta] = ?
                    """, (puesto, pauta))
                    
                    total_controles_pauta = cursor.fetchone()[0]
                    
                    # Contar cuántos controles ha completado el usuario en este guardado
                    # SOLO contar controles con datos (que no fueron saltados)
                    controles_con_datos = sum(1 for r in registros if r.get('Datos') and r.get('Datos') != '')
                    
                    print(f"  📊 Controles: {controles_con_datos}/{total_controles_pauta}")
                    
                    # SOLO actualizar Cerrado si está completo
                    if controles_con_datos >= total_controles_pauta and total_controles_pauta > 0:
                        # Extraer dinámicamente el número del puesto (primeros 2 dígitos antes del guión)
                        # Ej: "10_Soldadura" -> 10, "17_INSPECCION POROS" -> 17, "40_Inspección" -> 40
                        try:
                            nuevo_cerrado = int(puesto.split('_')[0])
                            
                            cursor.execute("""
                                UPDATE [Digitalizacion].[PE].[Pedido]
                                SET [Cerrado] = ?
                                WHERE [ID_Pedido] = ?
                            """, (nuevo_cerrado, id_pedido))
                            
                            print(f"  ✅ Checklist COMPLETO - Actualizado Cerrado = {nuevo_cerrado} (puesto: {puesto})")
                        except (ValueError, IndexError):
                            print(f"  ⚠️ No se pudo extraer número del puesto: {puesto}")
                    else:
                        # Checklist incompleto - extraer número del puesto anterior dinámicamente
                        try:
                            puesto_numero = int(puesto.split('_')[0])
                            # El valor anterior es el puesto anterior (restamos 1 nivel)
                            # Obtenemos la lista ordenada de puestos para este pedido y pauta
                            cursor.execute("""
                                SELECT DISTINCT [Puesto]
                                FROM [Digitalizacion].[PE].[Checklist]
                                WHERE [Nombre_Pauta] = ?
                                ORDER BY [Puesto]
                            """, (pauta,))
                            
                            puestos_lista = [row[0] for row in cursor.fetchall()]
                            
                            if puesto in puestos_lista:
                                indice_actual = puestos_lista.index(puesto)
                                if indice_actual > 0:
                                    # Hay un puesto anterior - usar su número
                                    puesto_anterior = puestos_lista[indice_actual - 1]
                                    valor_anterior = int(puesto_anterior.split('_')[0])
                                else:
                                    # Es el primer puesto - valor anterior es 0
                                    valor_anterior = 0
                            else:
                                valor_anterior = 0
                            
                            cursor.execute("""
                                UPDATE [Digitalizacion].[PE].[Pedido]
                                SET [Cerrado] = ?
                                WHERE [ID_Pedido] = ?
                            """, (valor_anterior, id_pedido))
                            
                            print(f"  ⚠️ Checklist INCOMPLETO ({controles_con_datos}/{total_controles_pauta}) - Cerrado = {valor_anterior}")
                        except (ValueError, IndexError):
                            print(f"  ⚠️ No se pudo procesar puesto incompleto: {puesto}")

            # Obtener el valor actual de Cerrado para devolverlo en la respuesta
            cursor.execute("""
                SELECT [Cerrado] FROM [Digitalizacion].[PE].[Pedido]
                WHERE [ID_Pedido] = ?
            """, (id_pedido,))
            
            cerrado_actual_row = cursor.fetchone()
            cerrado_actual = cerrado_actual_row[0] if cerrado_actual_row else None

            conn.commit()

        return jsonify({
            'success': True, 
            'message': 'Datos guardados correctamente',
            'Cerrado': cerrado_actual
        }), 200

    except Exception as e:
        print(f"💥 Error guardando datos user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/datosuser/<int:id_pedido>', methods=['GET'])
def obtener_datos_user(id_pedido):
    """Recupera los datos guardados del checklist para un pedido específico desde PE.DatosUser"""
    try:
        print(f"📥 Recuperando datos del checklist para pedido {id_pedido}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()

            # Obtener todos los registros para este pedido
            cursor.execute("""
                SELECT [ID_Control], [Resultado], [Resultado_txt], [Comentario]
                FROM [Digitalizacion].[PE].[DatosUser]
                WHERE [ID_Pedido] = ?
                ORDER BY [ID_Control]
            """, (id_pedido,))

            resultados = cursor.fetchall()
            
            # Convertir a diccionario para fácil acceso en frontend
            datos = {}
            for row in resultados:
                id_control = row[0]
                resultado = row[1] if row[1] is not None else ''
                resultado_txt = row[2] if row[2] is not None else ''
                comentario = row[3] if row[3] is not None else ''
                
                datos[str(id_control)] = {
                    'resultado': resultado,
                    'resultado_txt': resultado_txt,
                    'comentario': comentario
                }

            print(f"✅ Encontrados {len(datos)} registros guardados para pedido {id_pedido}")
            
            return jsonify({
                'success': True, 
                'data': datos,
                'count': len(datos)
            }), 200

    except Exception as e:
        print(f"💥 Error recuperando datos user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# 🆕 ENDPOINT PARA FINALIZAR/CERRAR EL CHECKLIST
# ====================================================================================

@app.route('/api/finalizar-checklist', methods=['POST'])
def finalizar_checklist():
    """Finaliza el checklist actualizando el campo Cerrado en PE.Pedido
    Espera un JSON con { ID_Pedido: int, Puesto: str, Nombre_Pauta: str }
    El Cerrado se establece al número del puesto (ej: "30_INSPECCIONES" -> 30)
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

        id_pedido = data.get('ID_Pedido')
        puesto = data.get('Puesto', '')
        pauta = data.get('Nombre_Pauta', '')

        if not id_pedido or not puesto or not pauta:
            return jsonify({'success': False, 'message': 'Payload inválido: falta ID_Pedido, Puesto o Nombre_Pauta'}), 400

        print(f"🏁 Finalizando checklist para pedido {id_pedido}, puesto: {puesto}, pauta: {pauta}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()

            # Extraer el número del puesto de la cadena "XX_NombrePuesto"
            # Ej: "30_INSPECCIONES PREMONTAJE" -> 30
            try:
                nuevo_cerrado = int(puesto.split('_')[0])
                print(f"  📍 Número de puesto extraído: {nuevo_cerrado}")
            except (ValueError, IndexError):
                return jsonify({'success': False, 'message': f'No se pudo extraer número del puesto: {puesto}'}), 400

            # Actualizar el campo Cerrado en PE.Pedido
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Pedido]
                SET [Cerrado] = ?
                WHERE [ID_Pedido] = ?
            """, (nuevo_cerrado, id_pedido))

            conn.commit()

            print(f"  ✅ Checklist finalizado - Actualizado Cerrado = {nuevo_cerrado} para pedido {id_pedido}")

            return jsonify({
                'success': True,
                'message': f'Checklist finalizado correctamente',
                'Cerrado': nuevo_cerrado
            }), 200

    except Exception as e:
        print(f"💥 Error finalizando checklist: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA SERVIR IMÁGENES
# ====================================================================================

@app.route('/api/imagen/<path:filename>')
def servir_imagen(filename):
    """
    Endpoint para servir imágenes de forma segura
    Convierte rutas locales en URLs web accesibles
    """
    try:
        import os
        
        print(f"🖼️ Solicitando imagen: {filename}")
        
        # Normalizamos la ruta: convertir slashes a backslashes (Windows)
        ruta_normalizada = filename.replace('/', '\\')
        print(f"📁 Ruta normalizada: {ruta_normalizada}")
        
        # INTENTO 1: Probar la ruta tal como viene (si es una ruta absoluta válida)
        if os.path.exists(ruta_normalizada) and os.path.isfile(ruta_normalizada):
            directory = os.path.dirname(ruta_normalizada)
            file_name = os.path.basename(ruta_normalizada)
            print(f"✅ Imagen encontrada directamente: {ruta_normalizada}")
            return send_from_directory(directory, file_name)
        
        # INTENTO 2: Si la ruta no comienza con \\ o letra de unidad, intentar reconstruirla como UNC
        if not ruta_normalizada.startswith('\\\\') and ':' not in ruta_normalizada:
            ruta_unc_reconstruida = '\\\\' + ruta_normalizada
            print(f"🔍 Intento UNC reconstruida: {ruta_unc_reconstruida}")
            if os.path.exists(ruta_unc_reconstruida) and os.path.isfile(ruta_unc_reconstruida):
                directory = os.path.dirname(ruta_unc_reconstruida)
                file_name = os.path.basename(ruta_unc_reconstruida)
                print(f"✅ Imagen encontrada (UNC reconstruida): {ruta_unc_reconstruida}")
                return send_from_directory(directory, file_name)
        
        # INTENTO 3: Si no es una ruta absoluta, buscar solo el nombre del archivo en carpetas predefinidas
        filename_clean = os.path.basename(filename)  # Extraer solo el nombre del archivo
        
        print(f"🔍 Buscando por nombre: {filename_clean}")
        
        # Buscar la imagen en los directorios configurados
        rutas_busqueda = [
            RUTA_IMAGENES,  # \\EMEBIDWH\DIgitalizacion\Checklist Power\IMAGENES
            os.path.join(BASE_DIR, 'fotos'),
            'C:/Foto',  # Ruta mencionada en el error
            'C:/Fotos',
            r'\\servidor\unidad compartida\Checklist Calidad\Imagenes',  # Red compartida
            r'\\servidor\unidad compartida\Fotos'  # Alternativa en red
        ]
        
        for ruta_base in rutas_busqueda:
            if os.path.exists(ruta_base):
                ruta_completa = os.path.join(ruta_base, filename_clean)
                if os.path.exists(ruta_completa) and os.path.isfile(ruta_completa):
                    print(f"✅ Imagen encontrada en: {ruta_completa}")
                    return send_from_directory(ruta_base, filename_clean)
        
        print(f"❌ Imagen no encontrada: {filename_clean}")
        print(f"   Rutas buscadas: {rutas_busqueda}")
        
        # Devolver imagen placeholder o error 404
        return jsonify({'error': 'Imagen no encontrada'}), 404
        
    except Exception as e:
        print(f"💥 Error sirviendo imagen {filename}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/archivo/<path:filename>')
def servir_archivo(filename):
    """
    Endpoint para servir archivos (PDFs, imágenes, documentos, etc.) de forma segura
    Soporta rutas locales y rutas compartidas en red
    """
    try:
        import os
        
        print(f"📄 Solicitando archivo: {filename}")
        
        # Normalizamos la ruta: convertir slashes a backslashes (Windows)
        ruta_normalizada = filename.replace('/', '\\')
        print(f"📁 Ruta normalizada: {ruta_normalizada}")
        
        # INTENTO 1: Probar la ruta tal como viene (si es una ruta absoluta válida)
        if os.path.exists(ruta_normalizada) and os.path.isfile(ruta_normalizada):
            directory = os.path.dirname(ruta_normalizada)
            file_name = os.path.basename(ruta_normalizada)
            print(f"✅ Archivo encontrado directamente: {ruta_normalizada}")
            return send_from_directory(directory, file_name)
        
        # INTENTO 2: Si la ruta no comienza con \\ o letra de unidad, intentar reconstruirla como UNC
        if not ruta_normalizada.startswith('\\\\') and ':' not in ruta_normalizada:
            ruta_unc_reconstruida = '\\\\' + ruta_normalizada
            print(f"🔍 Intento UNC reconstruida: {ruta_unc_reconstruida}")
            if os.path.exists(ruta_unc_reconstruida) and os.path.isfile(ruta_unc_reconstruida):
                directory = os.path.dirname(ruta_unc_reconstruida)
                file_name = os.path.basename(ruta_unc_reconstruida)
                print(f"✅ Archivo encontrado (UNC reconstruida): {ruta_unc_reconstruida}")
                return send_from_directory(directory, file_name)
        
        # INTENTO 3: Si no es una ruta absoluta, buscar solo el nombre del archivo en carpetas predefinidas
        filename_clean = os.path.basename(filename)  # Extraer solo el nombre del archivo
        
        print(f"🔍 Buscando por nombre: {filename_clean}")
        
        # Buscar el archivo en los directorios configurados
        rutas_busqueda = [
            RUTA_IMAGENES,  # \\EMEBIDWH\DIgitalizacion\Checklist Power\IMAGENES
            RUTA_PDFS,      # PDFs generados localmente
            os.path.join(BASE_DIR, 'fotos'),
            os.path.join(BASE_DIR, 'pdfs'),
            'C:/Foto',
            'C:/Fotos',
            'C:/PDF',
            'C:/PDFs',
            r'\\servidor\unidad compartida\Checklist Calidad\Imagenes',
            r'\\servidor\unidad compartida\Checklist Calidad\Fotos',
            r'\\servidor\unidad compartida\Checklist Calidad\PDFs'
        ]
        
        for ruta_base in rutas_busqueda:
            if os.path.exists(ruta_base):
                ruta_completa = os.path.join(ruta_base, filename_clean)
                if os.path.exists(ruta_completa) and os.path.isfile(ruta_completa):
                    print(f"✅ Archivo encontrado en: {ruta_completa}")
                    return send_from_directory(ruta_base, filename_clean)
        
        print(f"❌ Archivo no encontrado: {filename_clean}")
        print(f"   Rutas buscadas: {rutas_busqueda}")
        
        # Devolver error 404
        return jsonify({'error': 'Archivo no encontrado'}), 404
        
    except Exception as e:
        print(f"💥 Error sirviendo archivo {filename}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PUESTOS DE UNA PAUTA (WORKFLOW DINÁMICO)
# ====================================================================================
@app.route('/api/pauta-puestos/<pauta>', methods=['GET'])
def obtener_puestos_pauta(pauta):
    """
    Obtiene los puestos que tienen controles para una pauta específica.
    Devuelve lista ordenada, primer puesto y último puesto.
    """
    try:
        print(f"🔍 Consultando puestos con controles para pauta: {pauta}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener puestos distintos que tienen controles para esta pauta
            cursor.execute("""
                SELECT DISTINCT [Puesto]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Nombre_Pauta] = ?
                ORDER BY [Puesto]
            """, (pauta,))
            
            puestos_rows = cursor.fetchall()
            
            if not puestos_rows:
                return jsonify({
                    'success': False,
                    'message': f'No se encontraron puestos con controles para la pauta {pauta}'
                }), 404
            
            # Extraer nombres de puestos
            puestos = [row[0] for row in puestos_rows]
            
            # Extraer valores numéricos para ordenar (10, 20, 30, 40)
            def extraer_valor_numerico(puesto):
                try:
                    return int(puesto.split('_')[0])
                except:
                    return 0
            
            # Ordenar por valor numérico
            puestos_ordenados = sorted(puestos, key=extraer_valor_numerico)
            valores_numericos = [extraer_valor_numerico(p) for p in puestos_ordenados]
            
            primer_puesto = puestos_ordenados[0]
            ultimo_puesto = puestos_ordenados[-1]
            
            print(f"✅ Puestos encontrados para '{pauta}': {puestos_ordenados}")
            print(f"   📍 Primer puesto: {primer_puesto}, Último puesto: {ultimo_puesto}")
            
            return jsonify({
                'success': True,
                'pauta': pauta,
                'puestos': puestos_ordenados,
                'valores_numericos': valores_numericos,
                'primer_puesto': primer_puesto,
                'ultimo_puesto': ultimo_puesto,
                'total_puestos': len(puestos_ordenados)
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo puestos de pauta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINTS PARA MONITORIZACIÓN - ESTADO DE ARMARIOS POR PAUTA Y PUESTO
# ====================================================================================
def _extraer_valor_puesto(nombre_puesto):
    try:
        return int(''.join(filter(str.isdigit, nombre_puesto)))
    except Exception:
        return 999


def _obtener_resumen_monitoreo(cursor):
    print("📋 PASO 1: Obteniendo pautas...")
    cursor.execute("""
        SELECT DISTINCT [Nombre_Pauta]
        FROM [Digitalizacion].[PE].[Checklist]
        WHERE [Nombre_Pauta] IS NOT NULL AND [Nombre_Pauta] != ''
        ORDER BY [Nombre_Pauta]
    """)

    pautas = [row[0] for row in cursor.fetchall()]
    print(f"✅ PASO 1 COMPLETADO: Pautas encontradas: {pautas}")

    print("📍 PASO 2: Obteniendo puestos por pauta...")
    cursor.execute("""
        SELECT DISTINCT [Nombre_Pauta], [Puesto]
        FROM [Digitalizacion].[PE].[Checklist]
        WHERE [Nombre_Pauta] IS NOT NULL AND [Nombre_Pauta] != ''
          AND [Puesto] IS NOT NULL AND [Puesto] != ''
        ORDER BY [Nombre_Pauta], [Puesto]
    """)

    pauta_puestos_raw = {}
    for row in cursor.fetchall():
        pauta_puestos_raw.setdefault(row[0], []).append(row[1])

    pauta_puestos = {}
    pauta_primer_ultimo = {}
    for pauta in pautas:
        puestos_ordenados = sorted(pauta_puestos_raw.get(pauta, []), key=_extraer_valor_puesto)
        pauta_puestos[pauta] = puestos_ordenados
        pauta_primer_ultimo[pauta] = {
            'primer': puestos_ordenados[0] if puestos_ordenados else None,
            'ultimo': puestos_ordenados[-1] if puestos_ordenados else None,
            'todos': puestos_ordenados
        }

    cursor.execute("""
        SELECT [Nombre_Pauta], COUNT(DISTINCT [ID_Pedido]) AS total_armarios
        FROM [Digitalizacion].[PE].[Pedido]
        WHERE [Nombre_Pauta] IS NOT NULL AND [Nombre_Pauta] != ''
        GROUP BY [Nombre_Pauta]
    """)

    armarios_por_pauta = {row[0]: row[1] for row in cursor.fetchall()}

    print(f"✅ PASO 2 COMPLETADO: {len(pauta_puestos)} pautas con puestos")
    return pautas, pauta_puestos, pauta_primer_ultimo, armarios_por_pauta


def _obtener_estado_armarios_por_pauta(cursor, pauta, puestos_pauta, primer_ultimo_pauta):
    print(f"🛒 PASO 3: Obteniendo armarios/pedidos para pauta '{pauta}'...")
    cursor.execute("""
        SELECT DISTINCT
            [ID_Pedido],
            [Armario],
            [Referencia],
            [Nombre_Pauta],
            [Cerrado],
            [Fecha]
        FROM [Digitalizacion].[PE].[Pedido]
        WHERE [Nombre_Pauta] = ?
        ORDER BY [Armario]
    """, (pauta,))

    pedidos_rows = cursor.fetchall()
    print(f"✅ PASO 3 COMPLETADO: {len(pedidos_rows)} armarios encontrados para pauta '{pauta}'")

    print(f"⚡ PASO 3B: Precargando datos de controles y estados para pauta '{pauta}'...")
    cursor.execute("""
        WITH datos_agregados AS (
            SELECT
                du.[ID_Pedido],
                c.[Puesto],
                c.[Nombre_Pauta],
                COUNT(DISTINCT du.[ID_Control]) as total_controles,
                SUM(CASE WHEN du.[Resultado] = 'NOK' THEN 1 ELSE 0 END) as noks_count,
                MAX(CASE WHEN c.[TipoReg] = 'Listado Armarios' AND du.[Resultado] = 'OK' THEN du.[FechaRegistro] ELSE NULL END) as fecha_listado
            FROM [Digitalizacion].[PE].[DatosUser] du
            INNER JOIN [Digitalizacion].[PE].[Checklist] c ON du.[ID_Control] = c.[Id_Control]
            WHERE c.[Nombre_Pauta] = ?
            GROUP BY du.[ID_Pedido], c.[Puesto], c.[Nombre_Pauta]
        )
        SELECT [ID_Pedido], [Puesto], [Nombre_Pauta], [total_controles], [noks_count], [fecha_listado]
        FROM datos_agregados
    """, (pauta,))

    datos_pedido_puesto = {}
    for row in cursor.fetchall():
        fecha_str = row[5].strftime('%d/%m/%Y %H:%M') if row[5] else None
        datos_pedido_puesto[(row[0], row[1], row[2])] = {
            'completados': row[3],
            'noks': row[4] or 0,
            'fecha': fecha_str
        }
    print(f"✅ PASO 3B COMPLETADO: {len(datos_pedido_puesto)} registros precargados")

    cursor.execute("""
        SELECT [Puesto], COUNT(*) AS total_controles, MAX([TipoReg]) AS tipo_reg
        FROM [Digitalizacion].[PE].[Checklist]
        WHERE [Nombre_Pauta] = ?
        GROUP BY [Puesto]
    """, (pauta,))
    controles_por_puesto = {
        row[0]: {'total': row[1], 'tipo': row[2]}
        for row in cursor.fetchall()
    }

    print(f"📦 PASO 4: Procesando {len(pedidos_rows)} armarios para pauta '{pauta}'...")
    estado_armarios_pauta = {}
    contador = 0
    primer_puesto = primer_ultimo_pauta.get('primer')
    ultimo_puesto = primer_ultimo_pauta.get('ultimo')

    for pedido_row in pedidos_rows:
        contador += 1
        if contador % 50 == 0:
            print(f"   ⏳ Procesados {contador}/{len(pedidos_rows)} armarios de '{pauta}'...")

        id_pedido = pedido_row[0]
        armario = pedido_row[1]
        referencia = pedido_row[2] or ''
        cerrado_valor = pedido_row[4]
        armario_ref = f"{armario} - {referencia}" if referencia else armario

        if armario_ref not in estado_armarios_pauta:
            estado_armarios_pauta[armario_ref] = {
                'id_pedido': id_pedido,
                'armario': armario,
                'referencia': referencia,
                'puestos_estado': {}
            }

        for indice_puesto, puesto in enumerate(puestos_pauta):
            datos = datos_pedido_puesto.get((id_pedido, puesto, pauta), {'completados': 0, 'noks': 0, 'fecha': None})
            controles_config = controles_por_puesto.get(puesto, {'total': 0, 'tipo': None})
            total_controles = controles_config['total']
            tipo_reg = controles_config['tipo']
            tiene_listado = tipo_reg == 'Listado Armarios'
            puesto_numero = _extraer_valor_puesto(puesto)

            checklist_cerrado_ok = (
                bool(cerrado_valor)
                and cerrado_valor >= puesto_numero
                and (datos['noks'] or 0) == 0
                and (datos['completados'] or 0) > 0
            )

            if checklist_cerrado_ok and total_controles > datos['completados']:
                total_controles = datos['completados']

            es_visible = puesto == primer_puesto
            if not es_visible and indice_puesto > 0:
                puesto_anterior = puestos_pauta[indice_puesto - 1]
                valor_anterior = _extraer_valor_puesto(puesto_anterior)
                if cerrado_valor and cerrado_valor >= valor_anterior:
                    datos_anterior = datos_pedido_puesto.get((id_pedido, puesto_anterior, pauta), {'noks': 0})
                    es_visible = (datos_anterior.get('noks') or 0) == 0

            if es_visible:
                if datos['noks'] > 0:
                    color = 'ROJO'
                elif datos['completados'] < total_controles:
                    color = 'GRIS'
                elif puesto == ultimo_puesto:
                    color = 'VERDE'
                else:
                    color = 'GRIS'

                estado = {
                    'color': color,
                    'controles_completados': datos['completados'],
                    'total_controles': total_controles,
                    'noks': datos['noks'],
                    'es_listado_armarios': tiene_listado,
                    'fecha_registro': datos['fecha']
                }
            else:
                estado = 'NO_VISIBLE'

            estado_armarios_pauta[armario_ref]['puestos_estado'][puesto] = estado

    print(f"✅ PASO 4 COMPLETADO: {contador} armarios procesados para pauta '{pauta}'")
    return estado_armarios_pauta


@app.route('/api/monitoreo/pautas', methods=['GET'])
def monitoreo_resumen_pautas():
    try:
        print("🔍 INICIO: monitoreo_resumen_pautas")
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ No se pudo conectar a BD")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            try:
                cursor.timeout = 15
            except Exception:
                pass

            pautas, pauta_puestos, pauta_primer_ultimo, armarios_por_pauta = _obtener_resumen_monitoreo(cursor)

            return jsonify({
                'success': True,
                'pautas': pautas,
                'pauta_puestos': pauta_puestos,
                'pauta_primer_ultimo': pauta_primer_ultimo,
                'armarios_por_pauta': armarios_por_pauta
            })
    except Exception as e:
        print(f"💥 Error en monitoreo_resumen_pautas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/monitoreo/estado-armarios', methods=['GET'])
def monitoreo_estado_armarios():
    """
    Obtiene el estado de armarios para una pauta concreta.
    Si no se informa pauta, devuelve un resumen ligero con pautas y puestos.
    """
    request_started_at = time.perf_counter()
    pauta_seleccionada = (request.args.get('pauta') or '').strip()

    try:
        print(f"🔍 INICIO: monitoreo_estado_armarios pauta='{pauta_seleccionada or 'TODAS'}'")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ No se pudo conectar a BD")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            try:
                cursor.timeout = 30
            except Exception:
                pass

            pautas, pauta_puestos, pauta_primer_ultimo, armarios_por_pauta = _obtener_resumen_monitoreo(cursor)

            if not pauta_seleccionada:
                print(f"✅ MONITOREO RESUMEN COMPLETADO en {time.perf_counter() - request_started_at:.2f}s")
                return jsonify({
                    'success': True,
                    'pautas': pautas,
                    'pauta_puestos': pauta_puestos,
                    'pauta_primer_ultimo': pauta_primer_ultimo,
                    'armarios_por_pauta': armarios_por_pauta,
                    'estado_armarios': {}
                })

            if pauta_seleccionada not in pauta_puestos:
                return jsonify({'success': False, 'message': f'La pauta {pauta_seleccionada} no existe'}), 404

            estado_armarios_pauta = _obtener_estado_armarios_por_pauta(
                cursor,
                pauta_seleccionada,
                pauta_puestos.get(pauta_seleccionada, []),
                pauta_primer_ultimo.get(pauta_seleccionada, {})
            )

            total_armarios = len(estado_armarios_pauta)
            print(f"📋 PASO 5: Preparando respuesta JSON para pauta '{pauta_seleccionada}'...")
            print(f"✅ MONITOREO COMPLETADO en {time.perf_counter() - request_started_at:.2f}s. Pauta: {pauta_seleccionada}, Armarios: {total_armarios}")

            return jsonify({
                'success': True,
                'pautas': [pauta_seleccionada],
                'pauta_puestos': {pauta_seleccionada: pauta_puestos.get(pauta_seleccionada, [])},
                'pauta_primer_ultimo': {pauta_seleccionada: pauta_primer_ultimo.get(pauta_seleccionada, {})},
                'armarios_por_pauta': {pauta_seleccionada: total_armarios},
                'estado_armarios': {pauta_seleccionada: estado_armarios_pauta}
            })

    except Exception as e:
        print(f"💥 Error en monitoreo_estado_armarios tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PEDIDOS CON CHECKLIST INCOMPLETO (PRIMER PUESTO)
# ====================================================================================
@app.route('/api/pedidos-borradores/<puesto>/<pauta>', methods=['GET'])
def obtener_pedidos_borradores(puesto, pauta):
    """
    Obtiene pedidos INCOMPLETOS para RETRABAJAR:
    
    Un pedido es incompleto si:
    1. Nunca se completó este puesto (Cerrado < número_del_puesto_actual)
    2. O tiene datos parciales/NOK en este puesto
    
    Esto permite al usuario completar o mejorar el checklist.
    """
    try:
        print(f"🔍 Buscando pedidos INCOMPLETOS para Puesto: {puesto}, Pauta: {pauta}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # 🆕 Extraer número del puesto (ej: "10_Soldadura" → 10)
            puesto_numero = int(puesto.split('_')[0]) if '_' in puesto else 10
            print(f"   📍 Número del puesto extraído: {puesto_numero}")
            
            # 🆕 Obtener total de controles esperados para este puesto+pauta
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Puesto] = ? AND [Nombre_Pauta] = ?
            """, (puesto, pauta))
            
            total_controles_esperados = cursor.fetchone()[0]
            print(f"   📊 Total de controles esperados para {puesto}/{pauta}: {total_controles_esperados}")
            
            if total_controles_esperados == 0:
                return jsonify({
                    'success': True,
                    'pedidos': [],
                    'message': 'No hay controles configurados para este puesto/pauta'
                })
            
            # 🆕 LÓGICA MEJORADA: Obtener pedidos INCOMPLETOS o con NOKs (para retrabajo)
            cursor.execute(f"""
                SELECT 
                    p.[ID_Pedido], 
                    p.[Armario], 
                    p.[Referencia], 
                    p.[Fecha], 
                    p.[Nombre_Pauta], 
                    p.[Cerrado],
                    p.[NumPedido],
                    COALESCE((
                        SELECT COUNT(*) 
                        FROM [Digitalizacion].[PE].[DatosUser] du
                        JOIN [Digitalizacion].[PE].[Checklist] c ON du.[ID_Control] = c.[Id_Control]
                        WHERE du.[ID_Pedido] = p.[ID_Pedido] AND c.[Puesto] = ?
                    ), 0) as controles_completados,
                    {total_controles_esperados} as total_esperado,
                    CASE 
                        WHEN EXISTS (
                            SELECT 1 
                            FROM [Digitalizacion].[PE].[DatosUser] du_nok
                            JOIN [Digitalizacion].[PE].[Checklist] c_nok ON du_nok.[ID_Control] = c_nok.[Id_Control]
                            WHERE du_nok.[ID_Pedido] = p.[ID_Pedido] 
                              AND c_nok.[Puesto] = ?
                              AND du_nok.[Resultado] = 'NOK'
                        ) THEN 'CON_NOK'
                        ELSE 'INCOMPLETO'
                    END as tipo_retrabajo
                FROM [Digitalizacion].[PE].[Pedido] p
                WHERE p.[Nombre_Pauta] = ? 
                  AND (
                      ISNULL(p.[Cerrado], 0) < {puesto_numero}
                      OR
                      (ISNULL(p.[Cerrado], 0) = {puesto_numero} AND EXISTS (
                          SELECT 1 
                          FROM [Digitalizacion].[PE].[DatosUser] du_err
                          JOIN [Digitalizacion].[PE].[Checklist] c_err ON du_err.[ID_Control] = c_err.[Id_Control]
                          WHERE du_err.[ID_Pedido] = p.[ID_Pedido]
                            AND c_err.[Puesto] = ?
                            AND du_err.[Resultado] = 'NOK'
                      ))
                  )
                ORDER BY 
                   CASE WHEN EXISTS (
                        SELECT 1 
                        FROM [Digitalizacion].[PE].[DatosUser] du_ord
                        JOIN [Digitalizacion].[PE].[Checklist] c_ord ON du_ord.[ID_Control] = c_ord.[Id_Control]
                        WHERE du_ord.[ID_Pedido] = p.[ID_Pedido] 
                          AND c_ord.[Puesto] = ?
                          AND du_ord.[Resultado] = 'NOK'
                   ) THEN 1 ELSE 2 END,
                   p.[Cerrado] DESC, p.[Armario] ASC
            """, (puesto, puesto, pauta, puesto, puesto))
            
            resultados = cursor.fetchall()
            
            if not resultados:
                print(f"ℹ️ No hay pedidos INCOMPLETOS para pauta '{pauta}' en puesto {puesto}")
                return jsonify({
                    'success': True,
                    'pedidos': [],
                    'message': 'No hay checklists incompletos para este puesto'
                })
            
            # Convertir a lista de diccionarios
            pedidos_borradores = []
            for row in resultados:
                pedido = {
                    'ID_Pedido': row[0],
                    'Armario': row[1],
                    'Referencia': row[2],
                    'Fecha': row[3].isoformat() if row[3] else None,
                    'Nombre_Pauta': row[4],
                    'Cerrado': row[5],
                    'NumPedido': row[6] if row[6] else '',
                    'ControlesCompletados': row[7],
                    'TotalEsperado': row[8],
                    'TipoRetrabajo': row[9]  # INCOMPLETO o CON_NOK
                }
                pedidos_borradores.append(pedido)
                print(f"  📋 INCOMPLETO encontrado: Armario={row[1]}, Cerrado={row[5]}, Controles={row[7]}/{row[8]}, Tipo={row[9]}")
            
            for pedido in pedidos_borradores:
                _enriquecer_pedido_num_pedido(cursor, pedido)
            
            print(f"✅ {len(pedidos_borradores)} pedidos INCOMPLETOS encontrados")
            
            return jsonify({
                'success': True,
                'pedidos': pedidos_borradores,
                'total': len(pedidos_borradores),
                'message': f'Se encontraron {len(pedidos_borradores)} checklists incompletos'
            })
            
    except Exception as e:
        print(f"💥 Error obteniendo pedidos borradores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PEDIDOS NO CERRADOS (PROGRESIÓN DE PUESTOS) - DINÁMICO
# ====================================================================================

@app.route('/api/validar-configuracion-puesto/<puesto>/<pauta>', methods=['GET'])
def validar_configuracion_puesto(puesto, pauta):
    """
    Valida si un puesto está correctamente configurado para la pauta.
    
    RESTRICCIÓN: Si un puesto tiene >1 control Y al menos uno es "Listado Armarios", 
    es una configuración inválida (no se puede mezclar "Listado Armarios" con otros controles).
    
    Respuesta:
    {
        "success": true,
        "es_valido": true,
        "num_controles": 2,
        "tipos_controles": ["Atributo OK/NOK", "Listado Armarios"],
        "tiene_listado_armarios": true,
        "mensaje": "Configuración válida - Solo 1 control de tipo Listado Armarios"
    }
    
    O si es inválido:
    {
        "success": true,
        "es_valido": false,
        "num_controles": 3,
        "tipos_controles": ["Atributo OK/NOK", "Listado Armarios", "Otro tipo"],
        "tiene_listado_armarios": true,
        "mensaje": "❌ Pauta mal configurada: No puedes mezclar 'Listado Armarios' con otros tipos de controles. Contacta con tu responsable."
    }
    """
    try:
        print(f"🔍 Validando configuración de puesto: {puesto} / {pauta}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener todos los controles de este puesto y pauta
            cursor.execute("""
                SELECT [TipoReg]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Puesto] = ? AND [Nombre_Pauta] = ?
            """, (puesto, pauta))
            
            controles = cursor.fetchall()
            
            if not controles:
                return jsonify({
                    'success': True,
                    'es_valido': False,
                    'num_controles': 0,
                    'tipos_controles': [],
                    'tiene_listado_armarios': False,
                    'mensaje': 'No se encontraron controles para este puesto y pauta'
                }), 200
            
            tipos_controles = [row[0] for row in controles]
            tiene_listado_armarios = 'Listado Armarios' in tipos_controles
            num_controles = len(tipos_controles)
            
            # Validación: Si hay "Listado Armarios", no puede haber más de 1 control
            if tiene_listado_armarios and num_controles > 1:
                print(f"❌ CONFIGURACIÓN INVÁLIDA: {puesto}/{pauta} tiene {num_controles} controles incluyendo 'Listado Armarios'")
                return jsonify({
                    'success': True,
                    'es_valido': False,
                    'num_controles': num_controles,
                    'tipos_controles': tipos_controles,
                    'tiene_listado_armarios': True,
                    'mensaje': "❌ Pauta mal configurada: No puedes mezclar 'Listado Armarios' con otros tipos de controles. Contacta con tu responsable."
                }), 200
            
            # Si es válido
            es_valido = True
            if tiene_listado_armarios and num_controles == 1:
                mensaje = "✅ Configuración válida - Solo 1 control de tipo Listado Armarios"
            elif num_controles == 1:
                mensaje = f"✅ Configuración válida - 1 control de tipo {tipos_controles[0]}"
            else:
                mensaje = f"✅ Configuración válida - {num_controles} controles (sin Listado Armarios)"
            
            print(f"✅ Configuración VÁLIDA: {puesto}/{pauta}")
            
            return jsonify({
                'success': True,
                'es_valido': es_valido,
                'num_controles': num_controles,
                'tipos_controles': tipos_controles,
                'tiene_listado_armarios': tiene_listado_armarios,
                'mensaje': mensaje
            }), 200
            
    except Exception as e:
        print(f"💥 Error validando configuración: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/pedidos-disponibles/<puesto>/<pauta>', methods=['GET'])
def obtener_pedidos_disponibles(puesto, pauta):
    """
    Obtiene pedidos que han completado el puesto anterior de esta pauta y están listos para el puesto actual.
    RESTRICCIÓN: Solo si NO hay controles NOK en el puesto anterior.
    
    Lógica DINÁMICA:
    1. Consulta qué puestos tiene la pauta
    2. Encuentra el puesto anterior al actual en esa secuencia
    3. Busca pedidos con Cerrado = valor_puesto_anterior
    4. Valida que NO hay NOK en los controles del puesto anterior
    """
    try:
        print(f"🔍 Buscando pedidos disponibles para Puesto: {puesto}, Pauta: {pauta}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener puestos distintos que tienen controles para esta pauta
            cursor.execute("""
                SELECT DISTINCT [Puesto]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Nombre_Pauta] = ?
                ORDER BY [Puesto]
            """, (pauta,))
            
            puestos_rows = cursor.fetchall()
            
            if not puestos_rows:
                return jsonify({
                    'success': False,
                    'message': f'No se encontraron puestos con controles para la pauta {pauta}'
                }), 404
            
            # Extraer nombres de puestos y valores numéricos
            puestos = [row[0] for row in puestos_rows]
            
            def extraer_valor_numerico(puesto_nombre):
                try:
                    return int(puesto_nombre.split('_')[0])
                except:
                    return 0
            
            # Ordenar por valor numérico
            puestos_ordenados = sorted(puestos, key=extraer_valor_numerico)
            valores_numericos = [extraer_valor_numerico(p) for p in puestos_ordenados]
            
            print(f"📋 Puestos de la pauta '{pauta}': {puestos_ordenados}")
            
            # Verificar que el puesto actual existe en la pauta
            if puesto not in puestos_ordenados:
                return jsonify({
                    'success': False,
                    'message': f'El puesto {puesto} no tiene controles para la pauta {pauta}'
                }), 400
            
            # Encontrar índice del puesto actual
            indice_actual = puestos_ordenados.index(puesto)
            
            # Si es el primer puesto, no hay pedidos previos (debe crear nuevo)
            if indice_actual == 0:
                return jsonify({
                    'success': True,
                    'pedidos': [],
                    'count': 0,
                    'es_primer_puesto': True,
                    'message': 'Este es el primer puesto de la pauta. Debe crear un nuevo pedido.'
                }), 200
            
            # Obtener el puesto anterior
            puesto_anterior = puestos_ordenados[indice_actual - 1]
            valor_cerrado_anterior = extraer_valor_numerico(puesto_anterior)
            valor_cerrado_actual = extraer_valor_numerico(puesto)
            
            print(f"📍 Puesto actual: {puesto} (índice {indice_actual}, Cerrado = {valor_cerrado_actual})")
            print(f"📍 Puesto anterior: {puesto_anterior} (Cerrado = {valor_cerrado_anterior})")
            
            # Buscar pedidos:
            # 1. Que vengan del puesto anterior (Cerrado = valor_anterior)
            # 2. O que estén cerrados en el puesto actual PERO tengan NOKs (para retrabajar)
            cursor.execute("""
                SELECT [ID_Pedido], [Armario], [Referencia], [Fecha], [Cerrado], [Nombre_Pauta], [NumPedido]
                FROM [Digitalizacion].[PE].[Pedido] p
                WHERE ([Cerrado] = ? AND [Nombre_Pauta] = ?)
                   OR (
                       [Cerrado] = ? AND [Nombre_Pauta] = ?
                       AND EXISTS (
                           SELECT 1
                           FROM [Digitalizacion].[PE].[DatosUser] du
                           JOIN [Digitalizacion].[PE].[Checklist] c ON du.[ID_Control] = c.[Id_Control]
                           WHERE du.[ID_Pedido] = p.[ID_Pedido]
                             AND c.[Puesto] = ?
                             AND du.[Resultado] = 'NOK'
                       )
                   )
                ORDER BY [Fecha] DESC, [ID_Pedido] DESC
            """, (valor_cerrado_anterior, pauta, valor_cerrado_actual, pauta, puesto))
            
            resultados = cursor.fetchall()
            
            pedidos_validos = []
            for row in resultados:
                id_pedido = row[0]
                cerrado_val = row[4]
                
                # Si viene del puesto anterior, validar que no tenga NOKs pendientes ALLÍ
                # (Si ya está en el puesto actual, asumimos que pasó el anterior)
                if cerrado_val == valor_cerrado_anterior:
                    cursor.execute("""
                        SELECT COUNT(*)
                        FROM [Digitalizacion].[PE].[DatosUser] du
                        INNER JOIN [Digitalizacion].[PE].[Checklist] ch ON du.[ID_Control] = ch.[Id_Control]
                        WHERE du.[ID_Pedido] = ? 
                          AND ch.[Puesto] = ? 
                          AND du.[Resultado] = 'NOK'
                    """, (id_pedido, puesto_anterior))
                    
                    hay_nok_anterior = cursor.fetchone()[0] > 0
                    
                    if hay_nok_anterior:
                        print(f"⚠️ Pedido {id_pedido} tiene NOK en {puesto_anterior} - NO permitido avanzar")
                        continue  # Saltar este pedido
                    
                    print(f"✅ Pedido {id_pedido} SIN NOK en {puesto_anterior} - Permitido avanzar")
                else:
                    print(f"🔄 Pedido {id_pedido} ya está en puesto actual (Cerrado={cerrado_val}) - Verificando NOKs")
                
                # 🆕 Verificar si ya tiene NOKs en el puesto ACTUAL (para marcarlo en rojo)
                cursor.execute("""
                    SELECT COUNT(*)
                    FROM [Digitalizacion].[PE].[DatosUser] du
                    INNER JOIN [Digitalizacion].[PE].[Checklist] ch ON du.[ID_Control] = ch.[Id_Control]
                    WHERE du.[ID_Pedido] = ? 
                      AND ch.[Puesto] = ? 
                      AND du.[Resultado] = 'NOK'
                """, (id_pedido, puesto))
                
                hay_nok_actual = cursor.fetchone()[0] > 0
                tipo_retrabajo = 'CON_NOK' if hay_nok_actual else 'NORMAL'
                
                pedidos_validos.append({
                    'id': id_pedido,  # 🆕 Normalizado a 'id' para consistencia con frontend
                    'ID_Pedido': id_pedido,  # Mantener por compatibilidad
                    'Armario': row[1],
                    'Referencia': row[2] if row[2] else '',
                    'Fecha': row[3].strftime('%d/%m/%Y') if row[3] else '',
                    'Cerrado': row[4],
                    'Nombre_Pauta': row[5],
                    'NumPedido': row[6] if row[6] else '',
                    'TipoRetrabajo': tipo_retrabajo
                })
            
            for pedido in pedidos_validos:
                _enriquecer_pedido_num_pedido(cursor, pedido)
            
            print(f"✅ Encontrados {len(pedidos_validos)} pedidos permitidos (sin NOK en {puesto_anterior})")
            
            # 🆕 Obtener controles de esta pauta y puesto para detectar "Listado Armarios"
            cursor.execute("""
                SELECT [TipoReg]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Nombre_Pauta] = ? AND [Puesto] = ?
            """, (pauta, puesto))
            
            controles = cursor.fetchall()
            controles_pauta = [{'TipoReg': row[0]} for row in controles]
            
            return jsonify({
                'success': True,
                'pedidos': pedidos_validos,
                'pedidos_disponibles': pedidos_validos,  # Alias para compatibilidad
                'controles_pauta': controles_pauta,  # 🆕 Para detectar Listado Armarios
                'count': len(pedidos_validos),
                'puesto_actual': puesto,
                'puesto_anterior': puesto_anterior,
                'cerrado_requerido': valor_cerrado_anterior,
                'es_primer_puesto': False,
                'puestos_pauta': puestos_ordenados,
                'nombre_pauta': pauta
            }), 200
            
    except Exception as e:
        print(f"💥 Error obteniendo pedidos disponibles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# 🆕 ENDPOINT PARA VALIDAR PAUTAS COMPARTIBLES (PASO 2)
# ====================================================================================
@app.route('/api/validar-pautas-compartibles/<puesto>', methods=['GET'])
def validar_pautas_compartibles(puesto):
    """
    Detecta automáticamente si 2+ pautas comparten el MISMO control ID de tipo "Listado Armarios".
    
    Lógica:
    1. Obtener todas las pautas del puesto
    2. Para cada pauta: listar controles de tipo "Listado Armarios"
    3. Si 2+ pautas comparten MISMO ID de control → agrupable
    4. Retornar estructura para activar toggle "Compartir Puesto"
    
    Respuesta si ES compartible:
    {
        "success": true,
        "compartible": true,
        "pautas": ["Pauta1", "Pauta2"],
        "control_id_comun": 42,
        "nombre_control": "Listado Armarios",
        "mensaje": "Este puesto puede ser compartido: 2 pautas comparten el control 'Listado Armarios' (ID: 42)"
    }
    
    Respuesta si NO es compartible:
    {
        "success": true,
        "compartible": false,
        "pautas": ["Pauta1", "Pauta2"],
        "mensaje": "Este puesto NO puede ser compartido: Las pautas no comparten un control 'Listado Armarios' común"
    }
    """
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a BD'}), 500
            
            cursor = conn.cursor()
            
            print(f"\n🔍 VALIDANDO PAUTAS COMPARTIBLES")
            print(f"   📍 Puesto: {puesto}")
            
            # PASO 1: Obtener todas las pautas del puesto
            # Primero obtener el ID_Puesto del nombre
            cursor.execute("""
                SELECT ID_Puesto FROM [Digitalizacion].[PE].[Puesto]
                WHERE Nombre_Puesto = ?
            """, (puesto,))
            
            id_puesto_row = cursor.fetchone()
            if not id_puesto_row:
                print(f"   ❌ Puesto '{puesto}' no encontrado")
                return jsonify({
                    'success': False,
                    'message': f'Puesto "{puesto}" no encontrado en base de datos'
                }), 404
            
            id_puesto = id_puesto_row[0]
            
            cursor.execute("""
                SELECT DISTINCT p.[Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pautas] p
                INNER JOIN [Digitalizacion].[PE].[Controles] c ON p.[ID_Control] = c.[Id_Control]
                WHERE c.[ID_Puesto] = ? AND p.[Activo] = 1
                ORDER BY p.[Nombre_Pauta]
            """, (id_puesto,))
            
            pautas = [row[0] for row in cursor.fetchall()]
            print(f"   📋 Pautas encontradas: {pautas}")
            
            if len(pautas) < 2:
                print(f"   ❌ Insuficientes pautas para agrupar (necesita 2+)")
                return jsonify({
                    'success': True,
                    'compartible': False,
                    'pautas': pautas,
                    'mensaje': f'Se necesitan 2+ pautas para compartir puesto. Encontradas: {len(pautas)}'
                }), 200
            
            # PASO 2: Para cada pauta, obtener controles de tipo "Listado Armarios"
            pautas_con_listado = {}
            
            for pauta in pautas:
                # Usar tabla Checklist que vincula Puesto-Pauta-Control
                cursor.execute("""
                    SELECT DISTINCT [Id_Control]
                    FROM [Digitalizacion].[PE].[Checklist]
                    WHERE [Puesto] = ? 
                    AND [Nombre_Pauta] = ? 
                    AND [TipoReg] = 'Listado Armarios'
                """, (puesto, pauta))
                
                controles = cursor.fetchall()
                print(f"   📋 Pauta '{pauta}': {len(controles)} controles 'Listado Armarios'")
                
                if controles:
                    # Tomar el primer (y normalmente único) "Listado Armarios" de esta pauta
                    pautas_con_listado[pauta] = {
                        'id_control': controles[0][0],
                        'nombre_control': 'Listado Armarios', # Nombre genérico ya que no consultamos tabla Controles
                        'tipo': 'Listado Armarios'
                    }
            
            print(f"   ✅ Pautas con 'Listado Armarios': {list(pautas_con_listado.keys())}")
            
            # PASO 3: Detectar si 2+ pautas comparten el MISMO control ID
            if len(pautas_con_listado) < 2:
                print(f"   ❌ No hay suficientes pautas con 'Listado Armarios' común")
                return jsonify({
                    'success': True,
                    'compartible': False,
                    'pautas': pautas,
                    'pautas_con_listado': list(pautas_con_listado.keys()),
                    'mensaje': 'Las pautas no comparten un control "Listado Armarios" común'
                }), 200
            
            # Agrupar por ID de control para detectar duplicados
            controles_por_id = {}
            for pauta, info_control in pautas_con_listado.items():
                id_control = info_control['id_control']
                if id_control not in controles_por_id:
                    controles_por_id[id_control] = []
                controles_por_id[id_control].append(pauta)
            
            # Buscar controles compartidos por 2+ pautas
            controles_compartidos = {
                id_ctrl: pautas_list 
                for id_ctrl, pautas_list in controles_por_id.items() 
                if len(pautas_list) >= 2
            }
            
            print(f"   📊 Controles compartidos: {controles_compartidos}")
            
            if not controles_compartidos:
                print(f"   ❌ Ningún control es compartido por 2+ pautas")
                return jsonify({
                    'success': True,
                    'compartible': False,
                    'pautas': pautas,
                    'pautas_con_listado': list(pautas_con_listado.keys()),
                    'mensaje': 'Aunque hay "Listado Armarios", cada pauta usa un control diferente'
                }), 200
            
            # ÉXITO: Hay al menos un control compartido
            control_id_comun = list(controles_compartidos.keys())[0]
            pautas_agrupadas = controles_compartidos[control_id_comun]
            nombre_control = pautas_con_listado[pautas_agrupadas[0]]['nombre_control']
            
            print(f"   ✅ COMPARTIBLE: {len(pautas_agrupadas)} pautas comparten control ID {control_id_comun}")
            
            return jsonify({
                'success': True,
                'compartible': True,
                'pautas': pautas_agrupadas,
                'control_id_comun': control_id_comun,
                'nombre_control': nombre_control,
                'tipo_control': 'Listado Armarios',
                'mensaje': f'✅ Este puesto PUEDE compartirse: {len(pautas_agrupadas)} pautas comparten el control "{nombre_control}" (ID: {control_id_comun})'
            }), 200
            
    except Exception as e:
        print(f"💥 Error validando pautas compartibles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA REGISTRAR LISTADO ARMARIOS (puestos con solo "Listado Armarios")
# ====================================================================================
@app.route('/api/registrar-listado-armarios', methods=['POST'])
def registrar_listado_armarios():
    """
    Registra múltiples armarios seleccionados para un puesto con control "Listado Armarios".
    
    Body esperado:
    {
        "puesto": "nombre_puesto",
        "pauta": "nombre_pauta",
        "armarios": [
            {"id": "ID_PEDIDO_REAL", "nombre": "73_24"},
            {"id": "ID_PEDIDO_REAL", "nombre": "73_25"}
        ]
    }
    
    IMPORTANTE: El campo "id" DEBE ser el ID_Pedido de la tabla Pedido, NO el número de armario
    """
    try:
        data = request.get_json()
        puesto = data.get('puesto')
        pauta = data.get('pauta')
        armarios = data.get('armarios', [])
        usuario_logueado = obtener_usuario_sesion()
        
        print(f"\n💾 INICIANDO REGISTRO DE LISTADO ARMARIOS")
        print(f"   📍 Puesto: {puesto}")
        print(f"   📋 Pauta: {pauta}")
        print(f"   👤 Usuario: {usuario_logueado}")
        print(f"   🔢 Armarios a registrar: {len(armarios)}")
        
        # Log detallado de armarios recibidos
        for i, arm in enumerate(armarios, 1):
            print(f"      {i}. ID={arm.get('id')} | Nombre={arm.get('nombre')}")
        
        if not puesto or not pauta or not armarios:
            print(f"❌ ERROR: Faltan datos requeridos (puesto={puesto}, pauta={pauta}, armarios={len(armarios)})")
            return jsonify({'success': False, 'message': 'Faltan datos requeridos'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ ERROR: No se pudo conectar a la base de datos")
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # 🆕 PASO 1: Validar que el puesto tiene SOLO control "Listado Armarios"
            cursor.execute("""
                SELECT [TipoReg]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Nombre_Pauta] = ? AND [Puesto] = ?
            """, (pauta, puesto))
            
            controles = cursor.fetchall()
            tipos_controles = [row[0] for row in controles]
            
            print(f"\n🔍 VALIDACIÓN DE CONFIGURACIÓN")
            print(f"   Tipos de controles encontrados: {tipos_controles}")
            
            tiene_listado_armarios = 'Listado Armarios' in tipos_controles
            num_controles = len(tipos_controles)
            
            if not tiene_listado_armarios:
                print(f"❌ ERROR: Pauta {pauta} NO tiene control 'Listado Armarios' en puesto {puesto}")
                return jsonify({'success': False, 'message': f'Pauta no tiene control "Listado Armarios" en {puesto}'}), 400
            
            if num_controles > 1:
                print(f"❌ ERROR: Pauta mal configurada. Puesto {puesto} tiene {num_controles} controles: {tipos_controles}")
                return jsonify({'success': False, 'message': 'Pauta mal configurada. Contacta con tu responsable'}), 400
            
            print(f"   ✅ Configuración válida: 1 control 'Listado Armarios'")
            
            # Obtener el ID del control "Listado Armarios"
            cursor.execute("""
                SELECT [Id_Control]
                FROM [Digitalizacion].[PE].[Checklist]
                WHERE [Nombre_Pauta] = ? 
                  AND [Puesto] = ? 
                  AND [TipoReg] = 'Listado Armarios'
            """, (pauta, puesto))
            
            result = cursor.fetchone()
            if not result:
                print(f"❌ ERROR: No se encontró control 'Listado Armarios' para {puesto} / {pauta}")
                return jsonify({'success': False, 'message': f'No se encontró control "Listado Armarios"'}), 404
            
            id_control = result[0]
            print(f"   ✅ Control ID_Control={id_control}")
            
            # 🆕 PASO 2: Validar que los IDs de armarios son ID_Pedido válidos
            print(f"\n🔍 VALIDACIÓN DE ARMARIOS")
            
            armarios_validos = []
            for armario in armarios:
                armario_id = str(armario.get('id', '0')).strip()
                armario_nombre = armario.get('nombre', '')
                
                # Intentar convertir a entero
                try:
                    id_pedido = int(armario_id)
                except (ValueError, TypeError):
                    print(f"   ⚠️ ADVERTENCIA: ID armario inválido '{armario_id}' para {armario_nombre} - Saltando")
                    continue
                
                # Verificar que el ID_Pedido existe en la tabla Pedido
                # ⚠️ MODIFICADO: Se elimina el filtro por Nombre_Pauta para soportar puestos compartidos
                # donde los armarios pueden pertenecer a diferentes pautas del grupo
                cursor.execute("""
                    SELECT [ID_Pedido], [Armario], [Referencia], [Nombre_Pauta]
                    FROM [Digitalizacion].[PE].[Pedido]
                    WHERE [ID_Pedido] = ?
                """, (id_pedido,))
                
                pedido_row = cursor.fetchone()
                
                if pedido_row:
                    print(f"   ✅ Armario {armario_nombre} (ID_Pedido={id_pedido}) - Pauta Real: {pedido_row[3]} - Válido")
                    armarios_validos.append({
                        'id': id_pedido,
                        'nombre': armario_nombre,
                        'armario_numero': pedido_row[1]
                    })
                else:
                    print(f"   ❌ Armario {armario_nombre} (ID_Pedido={id_pedido}) - NO ENCONTRADO en tabla Pedido")
            
            if not armarios_validos:
                print(f"❌ ERROR: Ninguno de los armarios proporcionados es válido")
                return jsonify({'success': False, 'message': 'No se encontraron armarios válidos'}), 400
            
            print(f"   Total válidos: {len(armarios_validos)} / {len(armarios)}")
            
            # 🆕 PASO 3: Insertar o actualizar registros en DatosUser
            print(f"\n📝 CREANDO/ACTUALIZANDO REGISTROS EN DATOSUSER")
            
            registros_creados = 0
            registros_actualizados = 0
            
            for armario in armarios_validos:
                try:
                    id_pedido = armario['id']
                    nombre = armario['nombre']
                    
                    # Usar MERGE para insertar o actualizar (INSERT si no existe, UPDATE si existe)
                    cursor.execute("""
                        MERGE INTO [Digitalizacion].[PE].[DatosUser] AS target
                        USING (SELECT ? AS ID_Pedido, ? AS ID_Control) AS source
                        ON target.[ID_Pedido] = source.[ID_Pedido] 
                           AND target.[ID_Control] = source.[ID_Control]
                        WHEN MATCHED THEN
                            UPDATE SET 
                                target.[Resultado] = 'OK',
                                target.[Comentario] = '',
                                target.[Resultado_txt] = NULL,
                                target.[User] = ?
                        WHEN NOT MATCHED THEN
                            INSERT ([ID_Pedido], [ID_Control], [Resultado], [Comentario], [Resultado_txt], [User])
                            VALUES (?, ?, 'OK', '', NULL, ?);
                    """, (
                        id_pedido,
                        id_control,
                        usuario_logueado,
                        id_pedido,
                        id_control,
                        usuario_logueado
                    ))
                    
                    # Verificar si fue INSERT o UPDATE
                    if cursor.rowcount == 1:
                        registros_creados += 1
                        print(f"   ✅ Insertado: ID_Pedido={id_pedido} ({nombre}) con Resultado=OK")
                    else:
                        registros_actualizados += 1
                        print(f"   🔄 Actualizado: ID_Pedido={id_pedido} ({nombre}) con Resultado=OK")
                    
                except Exception as e:
                    print(f"   ❌ Error procesando armario {nombre}: {e}")
                    import traceback
                    traceback.print_exc()
            
            # 🆕 PASO 4: Actualizar campo Cerrado en tabla Pedido
            print(f"\n🔄 ACTUALIZANDO CAMPO CERRADO")
            
            try:
                valor_puesto = int(puesto.split('_')[0])
                print(f"   📍 Valor numérico del puesto: {valor_puesto}")
            except (ValueError, IndexError):
                print(f"   ⚠️ No se pudo extraer valor numérico de '{puesto}'")
                valor_puesto = 99
            
            ids_pedidos = [arm['id'] for arm in armarios_validos]
            
            if ids_pedidos:
                placeholders = ','.join('?' * len(ids_pedidos))
                # ⚠️ MODIFICADO: Actualizar solo por ID_Pedido, sin filtrar por pauta
                query = f"""
                    UPDATE [Digitalizacion].[PE].[Pedido]
                    SET [Cerrado] = ?
                    WHERE [ID_Pedido] IN ({placeholders})
                """
                
                params = [valor_puesto] + ids_pedidos
                cursor.execute(query, params)
                rows_affected = cursor.rowcount
                
                print(f"   ✅ Actualizados {rows_affected} pedidos: Cerrado = {valor_puesto}")
                for id_p in ids_pedidos:
                    print(f"      - ID_Pedido {id_p}")
            
            # 🆕 PASO 5: Detectar si puesto es compartido y retornar info de filtros ANTES del commit
            print(f"\n🔍 DETECTANDO SI PUESTO ES COMPARTIDO")
            
            cursor.execute("""
                SELECT TOP 1 [puesto_compartido]
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE [Nombre_Puesto] = ?
            """, (puesto,))
            
            puesto_result = cursor.fetchone()
            puesto_compartido = puesto_result[0] if puesto_result else 0
            
            # Obtener pautas si el puesto es compartido
            pautas_compartidas = []
            if puesto_compartido == 1:
                print(f"   ✅ Puesto ES compartido - Obteniendo pautas para filtros")
                
                # Obtener ID_Puesto del nombre
                cursor.execute("""
                    SELECT ID_Puesto FROM [Digitalizacion].[PE].[Puesto]
                    WHERE Nombre_Puesto = ?
                """, (puesto,))
                id_puesto = cursor.fetchone()[0]
                
                cursor.execute("""
                    SELECT DISTINCT p.[Nombre_Pauta]
                    FROM [Digitalizacion].[PE].[Pautas] p
                    INNER JOIN [Digitalizacion].[PE].[Controles] c ON p.[ID_Control] = c.[Id_Control]
                    WHERE c.[ID_Puesto] = ? AND p.[Activo] = 1
                    ORDER BY p.[Nombre_Pauta]
                """, (id_puesto,))
                
                pautas_compartidas = [row[0] for row in cursor.fetchall()]
                print(f"   📋 Pautas disponibles para filtros: {pautas_compartidas}")
            else:
                print(f"   ❌ Puesto NO es compartido")
                pautas_compartidas = [pauta]
            
            # Ahora hacer el commit
            conn.commit()
            
            print(f"\n✅ REGISTRO COMPLETADO EXITOSAMENTE")
            print(f"   Total registros creados: {registros_creados}")
            print(f"   Total registros actualizados: {registros_actualizados}")
            print(f"   Pedidos actualizados: {len(ids_pedidos)}")
            
            respuesta = {
                'success': True,
                'message': f'Se procesaron {registros_creados + registros_actualizados} armarios correctamente',
                'registros_creados': registros_creados,
                'registros_actualizados': registros_actualizados,
                'pedidos_actualizados': len(ids_pedidos),
                'armarios_procesados': armarios_validos,
                'puesto_compartido': puesto_compartido,
                'pautas_compartidas': pautas_compartidas,
                'mostrar_filtros': puesto_compartido == 1
            }
            
            return jsonify(respuesta), 200
            
    except Exception as e:
        print(f"💥 Error registrando listado armarios: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# 🆕 ENDPOINT PARA OBTENER ARMARIOS CON FILTRO POR PAUTA (PASO 5)
# ====================================================================================
@app.route('/api/obtener-armarios-filtrados/<puesto>/<pauta_filtro>', methods=['GET'])
def obtener_armarios_filtrados(puesto, pauta_filtro):
    """
    Obtiene armarios para un puesto compartido, aplicando lógica dinámica de secuencia.
    
    Lógica:
    1. Determina la secuencia de puestos (ej: 10->20->30)
    2. Identifica el puesto anterior al actual
    3. Filtra pedidos donde Cerrado = Valor_Puesto_Anterior
    4. Excluye pedidos con NOKs en el puesto anterior
    """
    try:
        filtro = request.args.get('filtro', 'todos')
        
        print(f"\n🔍 OBTENIENDO ARMARIOS FILTRADOS (DINÁMICO)")
        print(f"   📍 Puesto: {puesto}")
        print(f"   🔽 Filtro: {filtro}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a BD'}), 500
            
            cursor = conn.cursor()
            
            # Verificar si el puesto es compartido
            cursor.execute("""
                SELECT TOP 1 [puesto_compartido]
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE [Nombre_Puesto] = ?
            """, (puesto,))
            
            puesto_result = cursor.fetchone()
            puesto_compartido = puesto_result[0] if puesto_result else 0
            
            print(f"   🔄 Puesto compartido: {puesto_compartido}")
            
            # Obtener pautas disponibles para este puesto
            cursor.execute("""
                SELECT DISTINCT p.[Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pautas] p
                INNER JOIN [Digitalizacion].[PE].[Controles] c ON p.[ID_Control] = c.[Id_Control]
                INNER JOIN [Digitalizacion].[PE].[Puesto] pu ON c.[ID_Puesto] = pu.[ID_Puesto]
                WHERE pu.[Nombre_Puesto] = ? AND p.[Activo] = 1
                ORDER BY p.[Nombre_Pauta]
            """, (puesto,))
            
            pautas_disponibles = [row[0] for row in cursor.fetchall()]
            print(f"   📋 Pautas disponibles: {pautas_disponibles}")
            
            if not pautas_disponibles:
                print(f"   ❌ NO SE ENCONTRARON PAUTAS - Retornando lista vacía")
                return jsonify({
                    'success': True,
                    'puesto': puesto,
                    'puesto_compartido': puesto_compartido == 1,
                    'pautas_disponibles': [],
                    'filtro_actual': filtro,
                    'armarios': [],
                    'total': 0
                }), 200
            
            # ---------------------------------------------------------
            # LÓGICA DINÁMICA DE SECUENCIA (Cerrado = Puesto Anterior)
            # ---------------------------------------------------------
            
            condiciones_or = []
            params_or = []
            
            # Iterar sobre cada pauta disponible para determinar su puesto anterior específico
            for pauta_nombre in pautas_disponibles:
                # Obtener secuencia de puestos para esta pauta
                cursor.execute("""
                    SELECT DISTINCT [Puesto]
                    FROM [Digitalizacion].[PE].[Checklist]
                    WHERE [Nombre_Pauta] = ?
                    ORDER BY [Puesto]
                """, (pauta_nombre,))
                
                puestos_rows = cursor.fetchall()
                puestos_lista = [row[0] for row in puestos_rows]
                
                def extraer_valor(p):
                    try: return int(p.split('_')[0])
                    except: return 0
                
                puestos_ordenados = sorted(puestos_lista, key=extraer_valor)
                
                if puesto in puestos_ordenados:
                    idx = puestos_ordenados.index(puesto)
                    if idx > 0:
                        puesto_anterior = puestos_ordenados[idx - 1]
                        valor_cerrado_requerido = extraer_valor(puesto_anterior)
                        
                        print(f"   👉 Pauta '{pauta_nombre}': Puesto anterior es '{puesto_anterior}' (Cerrado={valor_cerrado_requerido})")
                        
                        # Construir condición para esta pauta:
                        # (Nombre_Pauta = 'X' AND Cerrado = Y AND NOT EXISTS (NOK en puesto anterior))
                        # ADEMÁS: Incluir armarios CERRADOS pero con controles incompletos (para reabrirlos)
                        condicion = f"""
                        (
                            p.[Nombre_Pauta] = ? 
                            AND (
                                /* Armarios nuevos: Cerrado = puesto anterior */
                                (p.[Cerrado] = ? AND NOT EXISTS (
                                    SELECT 1
                                    FROM [Digitalizacion].[PE].[DatosUser] du
                                    JOIN [Digitalizacion].[PE].[Checklist] c ON du.[ID_Control] = c.[Id_Control]
                                    WHERE du.[ID_Pedido] = p.[ID_Pedido]
                                      AND c.[Puesto] = ?
                                      AND du.[Resultado] = 'NOK'
                                ))
                                OR
                                /* Armarios CERRADOS pero con controles incompletos (para reabrirlos) */
                                (p.[Cerrado] = ? AND EXISTS (
                                    SELECT 1
                                    FROM [Digitalizacion].[PE].[DatosUser] du
                                    JOIN [Digitalizacion].[PE].[Checklist] c ON du.[ID_Control] = c.[Id_Control]
                                    WHERE du.[ID_Pedido] = p.[ID_Pedido]
                                      AND c.[Puesto] = ?
                                      AND (du.[Resultado] IS NULL OR du.[Resultado] = 'NOK')
                                ))
                            )
                        )
                        """
                        condiciones_or.append(condicion)
                        params_or.extend([pauta_nombre, valor_cerrado_requerido, puesto_anterior, valor_cerrado_requerido, puesto])

                    else:
                        print(f"   ℹ️ Pauta '{pauta_nombre}': Es el primer puesto. No requiere cerrado anterior.")
                else:
                    print(f"   ⚠️ El puesto {puesto} no está en la secuencia de la pauta {pauta_nombre}")

            if not condiciones_or:
                print(f"   ❌ Ninguna pauta tiene puesto anterior válido (o no están en secuencia).")
                return jsonify({
                    'success': True,
                    'puesto': puesto,
                    'puesto_compartido': puesto_compartido == 1,
                    'pautas_disponibles': pautas_disponibles,
                    'filtro_actual': filtro,
                    'armarios': [],
                    'total': 0
                }), 200
                
            print(f"   🎯 Filtro Dinámico: Generadas {len(condiciones_or)} condiciones OR para pautas")
            
            # Validar filtro de pauta
            if filtro != 'todos' and filtro not in pautas_disponibles:
                print(f"   ⚠️ Filtro '{filtro}' no es válido - Usando 'todos'")
                filtro = 'todos'
            
            # Construir Query Final
            sql_base = """
                SELECT DISTINCT
                    p.[ID_Pedido], 
                    p.[Armario], 
                    p.[Referencia], 
                    p.[Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pedido] p
                WHERE 
            """
            
            # Combinar condiciones OR
            sql_where = " OR ".join(condiciones_or)
            
            # Si hay filtro específico de pauta, lo aplicamos envolviendo todo
            if filtro != 'todos':
                sql_final = f"{sql_base} (({sql_where}) AND p.[Nombre_Pauta] = ?)"
                params_or.append(filtro)
            else:
                sql_final = f"{sql_base} ({sql_where})"
            
            sql_final += " ORDER BY p.[Nombre_Pauta], p.[Armario]"
            
            print(f"   📝 Ejecutando Query...")
            cursor.execute(sql_final, params_or)
            armarios_rows = cursor.fetchall()
            armarios = [
                {
                    'id': row[0],
                    'nombre': row[1],
                    'referencia': row[2],
                    'pauta': row[3]
                }
                for row in armarios_rows
            ]
            
            print(f"   📦 Total armarios encontrados: {len(armarios)}")
            
            return jsonify({
                'success': True,
                'puesto': puesto,
                'puesto_compartido': puesto_compartido == 1,
                'pautas_disponibles': pautas_disponibles,
                'filtro_actual': filtro,
                'armarios': armarios,
                'total': len(armarios)
            }), 200
            
    except Exception as e:
        print(f"💥 Error obteniendo armarios filtrados: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# 🆕 ENDPOINTS PARA PDF EN LISTADO ARMARIOS
# ====================================================================================
@app.route('/api/obtener-pdf-status/<pauta>/<puesto>', methods=['GET'])
def obtener_pdf_status(pauta, puesto):
    """
    Obtiene el estado PDF de una pauta para un puesto.
    Retorna: { "pdf": 0 o 1 }
    """
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Primero obtener ID_Pauta del nombre de pauta
            cursor.execute("""
                SELECT [ID_Pauta]
                FROM [Digitalizacion].[PE].[Pautas]
                WHERE [Nombre_Pauta] = ?
            """, (pauta,))
            
            pauta_result = cursor.fetchone()
            if not pauta_result:
                print(f"⚠️ Pauta no encontrada: {pauta}")
                return jsonify({'pdf': 0}), 200
            
            id_pauta = pauta_result[0]
            
            # Obtener ID_Puesto del nombre de puesto
            cursor.execute("""
                SELECT [ID_Puesto]
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE [Nombre_Puesto] = ?
            """, (puesto,))
            
            puesto_result = cursor.fetchone()
            if not puesto_result:
                print(f"⚠️ Puesto no encontrado: {puesto}")
                return jsonify({'pdf': 0}), 200
            
            id_puesto = puesto_result[0]
            
            # Obtener el estado PDF de Proveedores_Pautas
            cursor.execute("""
                SELECT [pdf]
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE [ID_Pauta] = ? AND [ID_Puesto] = ?
            """, (id_pauta, id_puesto))
            
            result = cursor.fetchone()
            
            if result:
                # Convertir a entero (True -> 1, False -> 0, None -> 0)
                pdf_status = int(result[0]) if result[0] is not None else 0
                print(f"📄 PDF status para {pauta}/{puesto}: {pdf_status}")
                return jsonify({'pdf': pdf_status}), 200
            else:
                print(f"⚠️ No se encontró registro en Proveedores_Pautas para {pauta}/{puesto}")
                return jsonify({'pdf': 0}), 200
    
    except Exception as e:
        print(f"❌ Error obteniendo estado PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/obtener-proveedores-pdf/<pauta>/<puesto>', methods=['GET'])
def obtener_proveedores_pdf(pauta, puesto):
    """
    Obtiene los proveedores asociados a una pauta y puesto con PDF=1.
    """
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Primero obtener ID_Pauta del nombre de pauta
            cursor.execute("""
                SELECT [ID_Pauta]
                FROM [Digitalizacion].[PE].[Pautas]
                WHERE [Nombre_Pauta] = ?
            """, (pauta,))
            
            pauta_result = cursor.fetchone()
            if not pauta_result:
                print(f"⚠️ Pauta no encontrada: {pauta}")
                return jsonify({'proveedores': []}), 200
            
            id_pauta = pauta_result[0]
            
            # Obtener ID_Puesto del nombre de puesto
            cursor.execute("""
                SELECT [ID_Puesto]
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE [Nombre_Puesto] = ?
            """, (puesto,))
            
            puesto_result = cursor.fetchone()
            if not puesto_result:
                print(f"⚠️ Puesto no encontrado: {puesto}")
                return jsonify({'proveedores': []}), 200
            
            id_puesto = puesto_result[0]
            
            # Obtener todos los registros de Proveedores_Pautas para esta pauta y puesto con PDF=1
            cursor.execute("""
                SELECT DISTINCT pp.[ID_Proveedor], p.[Nombre_Proveedor]
                FROM [Digitalizacion].[PE].[Proveedores_Pautas] pp
                JOIN [Digitalizacion].[PE].[Proveedores] p ON pp.[ID_Proveedor] = p.[ID_Proveedor]
                WHERE pp.[ID_Pauta] = ? 
                  AND pp.[ID_Puesto] = ? 
                  AND pp.[pdf] = 1
            """, (id_pauta, id_puesto))
            
            proveedores = []
            for row in cursor.fetchall():
                proveedor = {
                    'ID_Proveedor': row[0],
                    'Nombre_Proveedor': row[1]
                }
                proveedores.append(proveedor)
                print(f"   ✅ Proveedor: {row[1]} (ID={row[0]})")
            
            print(f"📦 Total de proveedores con PDF=1 para {pauta}/{puesto}: {len(proveedores)}")
            return jsonify({'proveedores': proveedores}), 200
    
    except Exception as e:
        print(f"❌ Error obteniendo proveedores para PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# ====================================================================================
# ====================================================================================
# ENDPOINT PARA VISUALIZAR PDF
# ====================================================================================
@app.route('/api/visualizar-pdf/<filename>', methods=['GET'])
def visualizar_pdf(filename):
    """
    Sirve el PDF generado desde la carpeta compartida para visualizar en navegador.
    Soporta visualización en modal o descarga.
    """
    try:
        # Validar nombre de archivo (evitar path traversal)
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'success': False, 'message': 'Nombre de archivo inválido'}), 400
        
        filepath = os.path.join(RUTA_PDFS_COMPARTIDA, filename)
        
        if not os.path.exists(filepath):
            print(f"❌ PDF no encontrado: {filepath}")
            return jsonify({'success': False, 'message': f'Archivo no encontrado: {filename}'}), 404
        
        print(f"✅ Sirviendo PDF: {filepath}")
        
        # Servir el PDF para visualización en navegador con headers CORS
        response = send_file(
            filepath,
            mimetype='application/pdf',
            as_attachment=False,  # Mostrar en navegador, no descargar
            download_name=filename
        )
        
        # Agregar headers CORS y caché
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
    
    except Exception as e:
        print(f"❌ Error visualizando PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA GENERAR CARTA DE PORTE (con Origen, Destino, Transportista)
# ====================================================================================
@app.route('/api/generar-pdf-carta-porte', methods=['POST'])
def generar_pdf_carta_porte():
    """
    Genera la Carta de Porte con Origen (EMESA), Destino y Transportista.
    También guarda los registros en PE.Envíos y PE.Cargas.
    
    Body esperado:
    {
        "puesto": "15_Expedición",
        "pauta": "Armario 73_24",
        "origen": "EMESA",
        "id_destino": 1,
        "id_transportista": 2,
        "matricula_vehiculo": "AB-1234-CD",
        "matricula_remolque": "AB-1234-EF",
        "armarios": [62, 63, 64]  # IDs de los pedidos registrados
    }
    """
    if not PDF_DISPONIBLE:
        return jsonify({'success': False, 'message': 'PDF no disponible'}), 500
    
    try:
        data = request.get_json()
        puesto = data.get('puesto')
        pauta = data.get('pauta')
        origen = data.get('origen', 'EMESA')
        id_destino = data.get('id_destino')
        id_transportista = data.get('id_transportista')
        matricula_vehiculo = data.get('matricula_vehiculo', '').strip().upper()
        matricula_remolque = data.get('matricula_remolque', '').strip().upper()
        armarios_ids = data.get('armarios', [])
        
        print(f"\n📄 GENERANDO CARTA DE PORTE")
        print(f"   Puesto: {puesto}")
        print(f"   Pauta: {pauta}")
        print(f"   Origen: {origen}")
        print(f"   Destino ID: {id_destino}")
        print(f"   Transportista ID: {id_transportista}")
        print(f"   Matrícula Vehículo: {matricula_vehiculo}")
        print(f"   Matrícula Remolque: {matricula_remolque}")
        print(f"   Armarios: {armarios_ids}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos del Destino
            cursor.execute("""
                SELECT [Nombre_Proveedor], [Direccion], [CP], [Ciudad], [CIF]
                FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (id_destino,))
            
            destino_row = cursor.fetchone()
            if not destino_row:
                return jsonify({'success': False, 'message': 'Destino no encontrado'}), 404
            
            destino_nombre = destino_row[0]
            destino_direccion = destino_row[1] or ''
            destino_cp = destino_row[2] or ''
            destino_ciudad = destino_row[3] or ''
            destino_cif = destino_row[4] or ''
            
            print(f"   ✅ Destino: {destino_nombre}")
            
            # Obtener datos del Transportista
            cursor.execute("""
                SELECT [Nombre_Proveedor], [Direccion], [CP], [Ciudad], [CIF]
                FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (id_transportista,))
            
            transportista_row = cursor.fetchone()
            if not transportista_row:
                return jsonify({'success': False, 'message': 'Transportista no encontrado'}), 404
            
            transportista_nombre = transportista_row[0]
            transportista_direccion = transportista_row[1] or ''
            transportista_cp = transportista_row[2] or ''
            transportista_ciudad = transportista_row[3] or ''
            transportista_cif = transportista_row[4] or ''
            
            print(f"   ✅ Transportista: {transportista_nombre}")
            
            # 🆕 Crear o reutilizar registro en PE.Envíos
            print(f"\n   🚚 Procesando Envío...")
            id_envio = None
            
            try:
                # Verificar si ya existe un envío con esta matrícula de vehículo
                cursor.execute("""
                    SELECT [ID_Envío] FROM [PE].[Envíos]
                    WHERE [Matrícula_Vehículo] = ? AND [Activo] = 1
                    ORDER BY [Fecha_Creacion] DESC
                """, (matricula_vehiculo,))
                
                envio_row = cursor.fetchone()
                
                if envio_row:
                    id_envio = envio_row[0]
                    print(f"      ✅ Envío existente reutilizado: ID={id_envio}")
                else:
                    # Crear nuevo envío
                    cursor.execute("""
                        INSERT INTO [PE].[Envíos] 
                            ([Matrícula_Vehículo], [Matrícula_Remolque], [Fecha_Creacion], [Activo])
                        VALUES (?, ?, GETDATE(), 1)
                    """, (matricula_vehiculo, matricula_remolque if matricula_remolque else None))
                    
                    conn.commit()
                    
                    # Obtener el ID del envío recién creado
                    cursor.execute("""
                        SELECT [ID_Envío] FROM [PE].[Envíos]
                        WHERE [Matrícula_Vehículo] = ?
                        ORDER BY [Fecha_Creacion] DESC
                    """, (matricula_vehiculo,))
                    
                    envio_row = cursor.fetchone()
                    id_envio = envio_row[0] if envio_row else None
                    print(f"      ✅ Nuevo Envío creado: ID={id_envio}")
            
            except pyodbc.IntegrityError as e:
                # Si hay conflicto UNIQUE, reutilizar el existente
                print(f"      ⚠️ Matrícula duplicada, reutilizando envío existente")
                cursor.execute("""
                    SELECT [ID_Envío] FROM [PE].[Envíos]
                    WHERE [Matrícula_Vehículo] = ?
                    ORDER BY [Fecha_Creacion] DESC
                """, (matricula_vehiculo,))
                
                envio_row = cursor.fetchone()
                id_envio = envio_row[0] if envio_row else None
            
            # Generar fecha de registro
            fecha_registro = datetime.now().strftime('%d/%m/%Y')
            
            # Obtener datos de los armarios
            datos_armarios = []
            
            for armario_id in armarios_ids:
                # MODIFICADO: Obtener también la Referencia
                cursor.execute("""
                    SELECT [Armario], [Nombre_Pauta], [ID_Pedido], [Referencia]
                    FROM [Digitalizacion].[PE].[Pedido]
                    WHERE [ID_Pedido] = ?
                """, (armario_id,))
                
                pedido_row = cursor.fetchone()
                if not pedido_row:
                    print(f"   ⚠️ Pedido {armario_id} no encontrado")
                    continue
                
                codigo_armario = pedido_row[0]
                nombre_pauta_pedido = pedido_row[1]
                id_pedido = pedido_row[2]
                referencia_pedido = pedido_row[3] or '' # Nueva variable
                
                # Obtener peso y precio basado en la pauta ESPECÍFICA del armario
                cursor.execute("""
                    SELECT TOP (1) [ID_Pauta]
                    FROM [Digitalizacion].[PE].[Pautas]
                    WHERE [Nombre_Pauta] = ?
                """, (nombre_pauta_pedido,))
                
                pauta_row = cursor.fetchone()
                
                peso = 0
                precio = 0 # Nueva variable
                
                if not pauta_row:
                    print(f"   ⚠️ Pauta {nombre_pauta_pedido} no encontrada")
                else:
                    id_pauta_armario = pauta_row[0]
                    
                    cursor.execute("""
                        SELECT [ID_Puesto]
                        FROM [Digitalizacion].[PE].[Puestos]
                        WHERE [Nombre_Puesto] = ?
                    """, (puesto,))
                    
                    puesto_row = cursor.fetchone()
                    id_puesto = puesto_row[0] if puesto_row else None
                    
                    if not id_puesto:
                        print(f"   ⚠️ Puesto {puesto} no encontrado")
                    else:
                        # MODIFICADO: Obtener peso y precio
                        cursor.execute("""
                            SELECT TOP (1) [Peso], [Precio]
                            FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                            WHERE [ID_Proveedor] = ?
                              AND [ID_Pauta] = ?
                              AND [ID_Puesto] = ?
                        """, (id_destino, id_pauta_armario, id_puesto))
                        
                        datos_prov_pauta = cursor.fetchone()
                        if datos_prov_pauta:
                            peso = datos_prov_pauta[0] or 0
                            precio = datos_prov_pauta[1] or 0
                        
                        print(f"   📊 Datos obtenidos para {codigo_armario} (Pauta: {nombre_pauta_pedido}): Peso={peso}kg, Precio={precio}")
                
                datos_armarios.append({
                    'codigo': codigo_armario,
                    'descripcion': nombre_pauta_pedido,
                    'peso': peso,
                    'uds': 1,
                    'fecha': fecha_registro,
                    'id_pedido': id_pedido,
                    'referencia': referencia_pedido, # Nuevo campo
                    'precio': precio # Nuevo campo
                })
                
                print(f"   ✅ Armario: {codigo_armario} - Pauta: {nombre_pauta_pedido} - Peso: {peso} - Precio: {precio}")
                
                # 🆕 Insertar registro en PE.Cargas
                if id_envio:
                    try:
                        cursor.execute("""
                            INSERT INTO [PE].[Cargas]
                                ([ID_Proveedor_Destino], [ID_Proveedor_Transportista], [ID_Envío], [Fecha],
                                 [Origen], [ID_Mercancia], [Código_Mercancia], [Descripción_Mercancia],
                                 [Peso_KG], [Matrícula_Vehículo], [Matrícula_Remolque], [Fecha_Creacion], [Activo])
                            VALUES
                                (?, ?, ?, GETDATE(), ?, ?, ?, ?, ?, ?, ?, GETDATE(), 1)
                        """, (
                            id_destino, id_transportista, id_envio, origen, id_pedido,
                            codigo_armario, nombre_pauta_pedido, peso,
                            matricula_vehiculo, matricula_remolque if matricula_remolque else None
                        ))
                        
                        conn.commit()
                        print(f"      ✅ Carga registrada en PE.Cargas para armario {codigo_armario}")
                    
                    except Exception as e:
                        print(f"      ⚠️ Error al registrar carga: {e}")
                        conn.rollback()
            
            # Generar PDF
            fecha_registro_completa = datetime.now().strftime('%d/%m/%Y %H:%M')
            nombre_archivo = f"CartaPorte_{puesto}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            ruta_completa = os.path.join(RUTA_PDFS_COMPARTIDA, nombre_archivo)
            
            # Crear carpeta compartida si no existe
            if not os.path.exists(RUTA_PDFS_COMPARTIDA):
                try:
                    os.makedirs(RUTA_PDFS_COMPARTIDA)
                    print(f"   📁 Carpeta compartida creada: {RUTA_PDFS_COMPARTIDA}")
                except Exception as e:
                    print(f"   ⚠️ No se pudo crear carpeta compartida: {e}")
                    return jsonify({'success': False, 'message': f'No se puede acceder a carpeta compartida'}), 500
            
            print(f"\n📝 GENERANDO CARTA DE PORTE: {nombre_archivo}")
            
            # Crear documento PDF
            doc = SimpleDocTemplate(ruta_completa, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            
            # Logo y título
            logo_path = os.path.join(BASE_DIR, "IMAGENES", "Logo_EMESA.png")
            
            if os.path.exists(logo_path):
                try:
                    logo = Image(logo_path, width=3.49*cm, height=1.09*cm)
                    titulo_para = Paragraph("<b style='font-size:14px; color:#003366'>CARTA DE PORTE</b>", styles['Normal'])
                    
                    header_table_data = [[logo], [Spacer(1, 0.1*cm)], [titulo_para]]
                    header_table = Table(header_table_data, colWidths=[17.5*cm])
                    header_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (0, 2), (0, 2), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ]))
                    story.append(header_table)
                    print(f"   ✅ Logo cargado correctamente")
                except Exception as e:
                    print(f"   ⚠️ Error cargando logo: {e}")
                    story.append(Paragraph("<b style='font-size:14px; color:#003366'>CARTA DE PORTE</b>", styles['Normal']))
            else:
                story.append(Paragraph("<b style='font-size:14px; color:#003366'>CARTA DE PORTE</b>", styles['Normal']))
            
            story.append(Spacer(1, 0.5*cm))
            
            # Sección ORIGEN (EMESA - siempre fijo)
            origen_text = """<b>ORIGEN :</b>  EMESA Estampaciones Metálicas Épila S.A.<br/>
NIF: A-50796861<br/>
C/ Camino del Sabinar, s/n<br/>
Pol. Ind. Valdemuel - 50290<br/>
EPILA (Zaragoza)"""
            origen_para = Paragraph(origen_text, styles['Normal'])
            story.append(origen_para)
            story.append(Spacer(1, 0.6*cm))
            
            # Sección DESTINO (del dropdown)
            destino_text = f"""<b>DESTINO :</b>  {destino_nombre}<br/>
NIF: {destino_cif}<br/>
{destino_direccion}<br/>
{destino_cp} {destino_ciudad}"""
            destino_para = Paragraph(destino_text, styles['Normal'])
            story.append(destino_para)
            story.append(Spacer(1, 0.6*cm))
            
            # Sección TRANSPORTISTA (del dropdown)
            transportista_text = f"""<b>TRANSPORTISTA :</b>  {transportista_nombre}<br/>
NIF: {transportista_cif}<br/>
{transportista_direccion}<br/>
{transportista_cp} {transportista_ciudad}"""
            transportista_para = Paragraph(transportista_text, styles['Normal'])
            story.append(transportista_para)
            story.append(Spacer(1, 0.4*cm))
            
            # 🆕 Sección MATRÍCULAS (debajo del TRANSPORTISTA)
            matriculas_text = f"""<b>MATRÍCULA VEHÍCULO:</b> {matricula_vehiculo or 'No especificada'}<br/>
<b>MATRÍCULA REMOLQUE:</b> {matricula_remolque or 'No especificada'}"""
            matriculas_para = Paragraph(matriculas_text, styles['Normal'])
            story.append(matriculas_para)
            story.append(Spacer(1, 0.8*cm))
            
            # Tabla de armarios con BULTOS y FECHA
            fecha_hoy = datetime.now().strftime('%d/%m/%Y')
            data_table = [['BULTOS', 'CÓDIGO', 'DESCRIPCIÓN', 'PESO (KG)', 'UDS', 'FECHA']]
            
            for armario in datos_armarios:
                data_table.append([
                    '1',  # BULTOS siempre = 1
                    str(armario['codigo']),
                    str(armario['descripcion']),
                    str(armario['peso']),
                    str(armario['uds']),
                    fecha_hoy  # FECHA de hoy
                ])
            
            table = Table(data_table, colWidths=[1.5*cm, 2*cm, 4*cm, 2*cm, 1.2*cm, 2*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            
            # Pie de página con fecha
            story.append(Spacer(1, 1*cm))
            fecha_pie = f"Generado el: {fecha_registro_completa}"
            story.append(Paragraph(f"<i style='font-size:9px; color:#666'>{fecha_pie}</i>", styles['Normal']))
            
            # Generar PDF
            doc.build(story)
            
            # Verificar que el archivo fue creado
            if os.path.exists(ruta_completa):
                tamaño_kb = os.path.getsize(ruta_completa) / 1024
                print(f"   ✅ Carta de Porte generada correctamente: {ruta_completa}")
                print(f"   📊 Tamaño: {tamaño_kb:.2f} KB")
            else:
                print(f"   ❌ Error: Archivo PDF no se creó en {ruta_completa}")
                return jsonify({'success': False, 'message': 'Error al guardar el PDF'}), 500
            
            # 🆕 GENERAR EXCEL EN PARALELO
            if EXCEL_DISPONIBLE:
                print(f"\n📊 GENERANDO EXCEL: {nombre_archivo.replace('.pdf', '.xlsx')}")
                
                try:
                    # Crear libro de Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Cargas"
                    
                    # Definir cabeceras con estilos
                    # MODIFICADO: Nuevas cabeceras solicitadas
                    cabeceras = [
                        'Proveedor', 
                        'Cantidad', 
                        'Código', 
                        'Descripción material', 
                        'Referencia EMESA', 
                        'Precio', 
                        'GFH entrega', 
                        'Observaciones', 
                        'Plazo', 
                        'Nº Inmovilizado'
                    ]
                    
                    # Aplicar estilo a las cabeceras (fondo azul, fuente blanca, negrita)
                    blue_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
                    white_font = Font(bold=True, color='FFFFFF')
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for col_num, cabecera in enumerate(cabeceras, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.value = cabecera
                        cell.fill = blue_fill
                        cell.font = white_font
                        cell.border = thin_border
                    
                    # Llenar datos en Excel
                    fila = 2
                    for armario in datos_armarios:
                        # Columna 1: Proveedor (Nombre del destino)
                        ws.cell(row=fila, column=1).value = destino_nombre
                        
                        # Columna 2: Cantidad (Fijo 1)
                        ws.cell(row=fila, column=2).value = 1
                        
                        # Columna 3: Código (Armario)
                        ws.cell(row=fila, column=3).value = armario['codigo']
                        
                        # Columna 4: Descripción material (Referencia)
                        ws.cell(row=fila, column=4).value = armario['referencia']
                        
                        # Columna 5: Referencia EMESA (Referencia)
                        ws.cell(row=fila, column=5).value = armario['referencia']
                        
                        # Columna 6: Precio
                        ws.cell(row=fila, column=6).value = armario['precio']
                        
                        # Columna 7: GFH entrega (Vacío)
                        ws.cell(row=fila, column=7).value = ""
                        
                        # Columna 8: Observaciones (Vacío)
                        ws.cell(row=fila, column=8).value = ""
                        
                        # Columna 9: Plazo (Vacío)
                        ws.cell(row=fila, column=9).value = ""
                        
                        # Columna 10: Nº Inmovilizado (Vacío)
                        ws.cell(row=fila, column=10).value = ""
                        
                        fila += 1
                    
                    # Ajustar ancho de columnas
                    for col in range(1, 11):
                        ws.column_dimensions[chr(64 + col)].width = 18
                    
                    # Guardar Excel con el mismo nombre que el PDF
                    nombre_excel = nombre_archivo.replace('.pdf', '.xlsx')
                    ruta_excel = os.path.join(RUTA_PDFS_COMPARTIDA, nombre_excel)
                    
                    wb.save(ruta_excel)
                    
                    tamaño_excel_kb = os.path.getsize(ruta_excel) / 1024
                    print(f"   ✅ Excel generado correctamente: {ruta_excel}")
                    print(f"   📊 Tamaño: {tamaño_excel_kb:.2f} KB")
                    
                except Exception as e:
                    print(f"   ⚠️ Error generando Excel: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"   ⚠️ openpyxl no disponible - Excel no se generará")
            
            print(f"✅ Carta de Porte generada exitosamente")
            return jsonify({
                'success': True,
                'message': 'Carta de Porte y Excel generados exitosamente',
                'archivo': nombre_archivo,
                'origen': origen,
                'destino': destino_nombre,
                'transportista': transportista_nombre
            }), 200
    
    except Exception as e:
        print(f"❌ Error generando Carta de Porte: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ====================================================================================
# ENDPOINT PARA OBTENER PDF COMO BASE64
# ====================================================================================
@app.route('/api/obtener-pdf-base64/<filename>', methods=['GET'])
def obtener_pdf_base64(filename):
    """
    Retorna el PDF en formato base64 para visualización en canvas/embed
    """
    try:
        # Validar nombre de archivo
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'success': False, 'message': 'Nombre de archivo inválido'}), 400
        
        filepath = os.path.join(RUTA_PDFS_COMPARTIDA, filename)
        
        if not os.path.exists(filepath):
            print(f"❌ PDF no encontrado: {filepath}")
            return jsonify({'success': False, 'message': f'Archivo no encontrado'}), 404
        
        # Leer archivo y convertir a base64
        with open(filepath, 'rb') as f:
            pdf_data = f.read()
            pdf_base64 = __import__('base64').b64encode(pdf_data).decode('utf-8')
        
        print(f"✅ PDF convertido a base64: {filename}")
        
        return jsonify({
            'success': True,
            'data': pdf_base64,
            'filename': filename
        }), 200
    
    except Exception as e:
        print(f"❌ Error en obtener-pdf-base64: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA DESCARGAR PDF
# ====================================================================================
@app.route('/api/descargar-pdf/<filename>', methods=['GET'])
def descargar_pdf(filename):
    """
    Descarga el PDF generado desde la carpeta compartida.
    """
    try:
        # Validar nombre de archivo (evitar path traversal)
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'success': False, 'message': 'Nombre de archivo inválido'}), 400
        
        filepath = os.path.join(RUTA_PDFS_COMPARTIDA, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': f'Archivo no encontrado: {filename}'}), 404
        
        # Servir el PDF para descarga
        return send_file(
            filepath,
            mimetype='application/pdf',
            as_attachment=True,  # Descargar
            download_name=filename
        )
    
    except Exception as e:
        print(f"❌ Error descargando PDF: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA GUARDAR PDF AUTOMÁTICAMENTE EN CARPETA COMPARTIDA
# ====================================================================================
@app.route('/api/guardar-pdf-compartida/<int:id_pedido>', methods=['GET'])
def guardar_pdf_compartida(id_pedido):
    """
    Genera y guarda automáticamente el PDF en la carpeta compartida.
    Nombre del archivo:
    - Si todos los controles son OK: {NumPedido}.pdf
    - Si hay algún NOK: {NumPedido}_NOK.pdf
    """
    if not PDF_DISPONIBLE:
        return jsonify({'success': False, 'message': 'Librería reportlab no disponible'}), 500
    
    try:
        print(f"💾 Guardando PDF en carpeta compartida para pedido {id_pedido}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos del pedido
            cursor.execute("""
                SELECT [ID_Pedido], [Armario], [Referencia], [Fecha], [NumPedido], [Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE [ID_Pedido] = ?
            """, (id_pedido,))
            
            pedido_row = cursor.fetchone()
            if not pedido_row:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            armario = pedido_row[1]
            referencia = pedido_row[2] or 'Sin referencia'
            fecha_pedido = pedido_row[3].strftime('%d/%m/%Y') if pedido_row[3] else 'Sin fecha'
            num_pedido = pedido_row[4] or 'SIN_NUM_PEDIDO'
            nombre_pauta = pedido_row[5] or 'Sin pauta'
            
            print(f"  📋 Datos: NumPedido={num_pedido}, Armario={armario}")
            
            # Obtener todos los controles del pedido para detectar NOKs
            cursor.execute("""
                SELECT COUNT(*) as total, 
                       SUM(CASE WHEN [Resultado] = 'NOK' THEN 1 ELSE 0 END) as noks
                FROM [Digitalizacion].[PE].[DatosUser]
                WHERE [ID_Pedido] = ?
            """, (id_pedido,))
            
            resultado_row = cursor.fetchone()
            total_controles = resultado_row[0] or 0
            noks = resultado_row[1] or 0
            
            print(f"  📊 Análisis: Total={total_controles}, NOKs={noks}")
            
            # Determinar si hay NOKs
            hay_noks = noks > 0
            
            # Determinar nombre del archivo
            # Formato solicitado: "Pedido " + NumPedido + "_ Armario " + Nº armario
            nombre_base = f"Pedido {num_pedido}_ Armario {armario}"
            
            if hay_noks:
                nombre_archivo = f"{nombre_base}_NOK.pdf"
                print(f"  🚨 Hay {noks} controles NOK - Nombre: {nombre_archivo}")
            else:
                nombre_archivo = f"{nombre_base}.pdf"
                print(f"  ✅ Todos los controles OK - Nombre: {nombre_archivo}")
            
            # Generar el PDF
            buffer = BytesIO()
            
            # Obtener controles para el PDF
            # 🆕 USAMOS DISTINCT y FILTRO DE PAUTA para evitar duplicados
            cursor.execute("""
                SELECT DISTINCT
                    c.[Puesto],
                    c.[Nombre_Pauta],
                    c.[Id_Control],
                    c.[DescripcionControl],
                    c.[Metodo],
                    c.[PuntoInspección],
                    c.[Proceso],
                    c.[TipoReg],
                    c.[CaracInspeccion],
                    du.[Resultado],
                    du.[Resultado_txt],
                    du.[Comentario],
                    c.[Orden_Pauta],
                    ctrl.[VisiblePDF]
                FROM [Digitalizacion].[PE].[DatosUser] du
                INNER JOIN [Digitalizacion].[PE].[Checklist] c 
                    ON c.[Id_Control] = du.[ID_Control]
                INNER JOIN [Digitalizacion].[PE].[Controles] ctrl
                    ON ctrl.[Id_Control] = c.[Id_Control]
                WHERE du.[ID_Pedido] = ?
                AND c.[Nombre_Pauta] = ?
                AND ISNULL(ctrl.[VisiblePDF], 1) = 1
                ORDER BY c.[Puesto], c.[Nombre_Pauta], c.[Orden_Pauta]
            """, (id_pedido, nombre_pauta))
            
            controles = cursor.fetchall()
            
            if not controles:
                return jsonify({'success': False, 'message': 'No hay datos de checklist para generar PDF'}), 404
            
            # Generar PDF en buffer
            generar_pdf_resumen(buffer, armario, referencia, fecha_pedido, num_pedido, controles, cursor)
            buffer.seek(0)
            
            # Guardar en carpeta compartida
            try:
                ruta_completa = os.path.join(RUTA_PDFS_COMPARTIDA, nombre_archivo)
                
                # Crear carpeta si no existe
                if not os.path.exists(RUTA_PDFS_COMPARTIDA):
                    os.makedirs(RUTA_PDFS_COMPARTIDA)
                    print(f"  📁 Carpeta creada: {RUTA_PDFS_COMPARTIDA}")
                
                # Guardar archivo
                with open(ruta_completa, 'wb') as f:
                    f.write(buffer.getvalue())
                
                print(f"  ✅ PDF guardado en: {ruta_completa}")
                
                return jsonify({
                    'success': True,
                    'message': f'PDF guardado exitosamente: {nombre_archivo}',
                    'nombre_archivo': nombre_archivo,
                    'ruta': ruta_completa,
                    'hay_noks': hay_noks,
                    'total_noks': noks
                }), 200
                
            except Exception as e:
                print(f"  ❌ Error guardando PDF en carpeta compartida: {e}")
                return jsonify({
                    'success': False,
                    'message': f'Error guardando PDF: {str(e)}'
                }), 500
            
    except Exception as e:
        print(f"💥 Error en guardar_pdf_compartida: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

# ====================================================================================
# ENDPOINT PARA GENERAR PDF RESUMEN DEL CHECKLIST COMPLETO
# ====================================================================================
@app.route('/api/generar-pdf/<int:id_pedido>', methods=['GET'])
def generar_pdf_checklist(id_pedido):
    """
    Genera un PDF resumen con todos los controles completados de todos los puestos.
    Organizado por puestos: 10_Soldadura, 20_Subcontratación, 30_Montaje, 40_Inspección
    """
    if not PDF_DISPONIBLE:
        return jsonify({'success': False, 'message': 'Librería reportlab no disponible'}), 500
    
    try:
        print(f"📄 Generando PDF para pedido {id_pedido}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos del pedido (incluyendo Nombre_Pauta)
            cursor.execute("""
                SELECT [Armario], [Referencia], [Fecha], [NumPedido], [Nombre_Pauta]
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE [ID_Pedido] = ?
            """, (id_pedido,))
            
            pedido_row = cursor.fetchone()
            if not pedido_row:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            armario = pedido_row[0]
            referencia = pedido_row[1] or 'Sin referencia'
            fecha_pedido = pedido_row[2].strftime('%d/%m/%Y') if pedido_row[2] else 'Sin fecha'
            num_pedido = pedido_row[3] or 'Sin número de pedido'
            nombre_pauta = pedido_row[4] or 'Sin pauta'
            
            print(f"📋 Datos del pedido: Armario={armario}, Pauta={nombre_pauta}, NumPedido={num_pedido}")
            
            # Obtener todos los datos guardados del checklist con información de controles
            # FILTRANDO: Solo controles donde VisiblePDF = 1 Y Nombre_Pauta coincide con la del pedido
            # 🆕 USAMOS DISTINCT para evitar duplicados si DatosUser tiene registros sucios o históricos
            cursor.execute("""
                SELECT DISTINCT
                    c.[Puesto],
                    c.[Nombre_Pauta],
                    c.[Id_Control],
                    c.[DescripcionControl],
                    c.[Metodo],
                    c.[PuntoInspección],
                    c.[Proceso],
                    c.[TipoReg],
                    c.[CaracInspeccion],
                    du.[Resultado],
                    du.[Resultado_txt],
                    du.[Comentario],
                    c.[Orden_Pauta],
                    ctrl.[VisiblePDF]
                FROM [Digitalizacion].[PE].[DatosUser] du
                INNER JOIN [Digitalizacion].[PE].[Checklist] c 
                    ON c.[Id_Control] = du.[ID_Control]
                INNER JOIN [Digitalizacion].[PE].[Controles] ctrl
                    ON ctrl.[Id_Control] = c.[Id_Control]
                WHERE du.[ID_Pedido] = ?
                AND c.[Nombre_Pauta] = ?
                AND ISNULL(ctrl.[VisiblePDF], 1) = 1
                ORDER BY c.[Puesto], c.[Nombre_Pauta], c.[Orden_Pauta]
            """, (id_pedido, nombre_pauta))
            
            controles = cursor.fetchall()
            
            if not controles:
                return jsonify({'success': False, 'message': 'No hay datos de checklist para este pedido'}), 404
            
            # Generar PDF (pasar cursor para consultas adicionales)
            buffer = BytesIO()
            pdf = generar_pdf_resumen(buffer, armario, referencia, fecha_pedido, num_pedido, controles, cursor)
            buffer.seek(0)
            
            # 🆕 GUARDAR AUTOMÁTICAMENTE EL PDF EN LA CARPETA DEDICADA
            try:
                # Generar nombre de archivo con timestamp para evitar sobrescrituras
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                nombre_archivo_pdf = f'{armario}_{timestamp}.pdf'
                ruta_completa_pdf = os.path.join(RUTA_PDFS, nombre_archivo_pdf)
                
                # Guardar PDF en disco
                with open(ruta_completa_pdf, 'wb') as f:
                    f.write(buffer.getvalue())
                
                print(f"💾 PDF guardado automáticamente en: {ruta_completa_pdf}")
                
                # Resetear el buffer para enviarlo al navegador
                buffer.seek(0)
                
            except Exception as e:
                print(f"⚠️ Error guardando PDF automáticamente: {e}")
                # Continuar aunque falle el guardado - no bloquear la visualización
                buffer.seek(0)
            
            print(f"✅ PDF generado exitosamente para armario {armario}")
            
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=f'{armario}.pdf'
            )
            
    except Exception as e:
        print(f"💥 Error generando PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


def generar_pdf_resumen(buffer, armario, referencia, fecha_pedido, num_pedido, controles, cursor):
    """
    Función que genera el PDF con reportlab usando columnas configuradas por puesto
    """
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=2*cm, bottomMargin=1*cm)
    
    elementos = []
    styles_sheet = getSampleStyleSheet()
    
    # Estilo personalizado para títulos
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles_sheet['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c6bb3'),
        spaceAfter=20,
        alignment=1  # Centrado
    )
    
    subtitulo_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles_sheet['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#555'),
        spaceAfter=10
    )
    
    # Logo EMESA en cabecera
    logo_path = os.path.join(RUTA_IMAGENES, 'Logo_EMESA.png')
    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=3.49*cm, height=1.09*cm)
            elementos.append(logo)
            elementos.append(Spacer(1, 1*cm))
        except Exception as e:
            print(f"⚠️ Error cargando logo: {e}")
    
    # Título principal
    elementos.append(Paragraph(f"RESUMEN DE CHECKLIST", titulo_style))
    elementos.append(Spacer(1, 0.3*cm))
    
    # Información del pedido - 2 filas: Fila 1: Armario | Referencia, Fila 2: Nº Pedido | Fecha Pedido
    info_data = [
        ['Armario:', armario, 'Referencia:', referencia],
        ['Nº Pedido:', num_pedido, 'Fecha Pedido:', fecha_pedido]
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 6*cm, 4*cm, 6*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f7ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        # Hacer negrita las columnas de etiquetas (0 y 2)
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        # Fuente normal para valores (1 y 3)
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (3, 0), (3, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    
    elementos.append(info_table)
    elementos.append(Spacer(1, 0.5*cm))
    
    # Mapeo de campos de base de datos a índices en la tupla de controles
    campo_a_indice = {
        'Id_Control': 2,
        'DescripcionControl': 3,
        'Metodo': 4,
        'PuntoInspección': 5,
        'Proceso': 6,
        'TipoReg': 7,
        'CaracInspeccion': 8
    }
    
    # Organizar controles por puesto
    controles_por_puesto = {}
    for control in controles:
        puesto = control[0]
        if puesto not in controles_por_puesto:
            controles_por_puesto[puesto] = []
        controles_por_puesto[puesto].append(control)
    
    # Usar los puestos dinámicamente del checklist (no asuma un orden fijo)
    orden_puestos = sorted(controles_por_puesto.keys())
    
    for puesto in orden_puestos:
        if puesto not in controles_por_puesto:
            continue
        
        # Título del puesto
        elementos.append(Paragraph(f"<b>{puesto}</b>", subtitulo_style))
        
        # Obtener columnas configuradas para este puesto
        cursor.execute("""
            SELECT [Columna], [Nombre_Columna], [Orden_Columna]
            FROM [Digitalizacion].[PE].[Puestos]
            WHERE [Nombre_Puesto] = ?
            ORDER BY [Orden_Columna]
        """, (puesto,))
        
        columnas_config = cursor.fetchall()
        
        if not columnas_config:
            elementos.append(Paragraph(f"⚠️ No se encontraron columnas configuradas para {puesto}", styles_sheet['Normal']))
            continue
        
        print(f"📋 Columnas configuradas para {puesto}: {[col[1] for col in columnas_config]}")
        
        # Obtener las pautas únicas de este puesto
        pautas_unicas = list(set([c[1] for c in controles_por_puesto[puesto]]))
        
        for pauta in pautas_unicas:
            # Subtítulo de la pauta
            elementos.append(Paragraph(f"Pauta: <i>{pauta}</i>", styles_sheet['Normal']))
            elementos.append(Spacer(1, 0.2*cm))
            
            # Filtrar controles de esta pauta
            controles_pauta = [c for c in controles_por_puesto[puesto] if c[1] == pauta]
            
            # Crear cabeceras dinámicas según columnas configuradas
            cabeceras = [col[1] for col in columnas_config]  # Nombre_Columna
            cabeceras.append('Resultado')  # Agregar columna de resultado
            
            tabla_data = [cabeceras]
            
            # Calcular anchos de columna dinámicamente
            num_columnas = len(cabeceras)
            ancho_disponible = 27 * cm  # Ancho disponible en landscape A4
            ancho_resultado = 3 * cm
            ancho_restante = ancho_disponible - ancho_resultado
            ancho_por_columna = ancho_restante / (num_columnas - 1)
            
            col_widths = [ancho_por_columna] * (num_columnas - 1) + [ancho_resultado]
            
            # Estilo para párrafos en celdas (permite word wrap automático)
            celda_style = ParagraphStyle(
                'CeldaStyle',
                parent=styles_sheet['Normal'],
                fontSize=7,
                leading=9,
                wordWrap='CJK',
                alignment=0  # Left align
            )
            
            # Generar filas de datos
            for control in controles_pauta:
                fila = []
                
                # Extraer valores según columnas configuradas
                for col_config in columnas_config:
                    campo = col_config[0]  # Columna (ej: 'Id_Control', 'DescripcionControl')
                    
                    if campo in campo_a_indice:
                        indice = campo_a_indice[campo]
                        valor = control[indice] if control[indice] is not None else ''
                        
                        # Usar Paragraph para permitir saltos de línea automáticos
                        # NO truncar, dejar que reportlab lo maneje
                        valor_str = str(valor) if valor else ''
                        fila.append(Paragraph(valor_str, celda_style))
                    else:
                        fila.append(Paragraph('-', celda_style))
                
                # Agregar resultado (combinar Resultado y Resultado_txt)
                resultado = control[9] or ''  # Índice 9: Resultado
                resultado_txt = control[10] or ''  # Índice 10: Resultado_txt
                
                if resultado and resultado_txt:
                    resultado_final = f"{resultado_txt} ({resultado})"
                elif resultado_txt:
                    resultado_final = resultado_txt
                elif resultado:
                    resultado_final = resultado
                else:
                    resultado_final = '-'
                
                fila.append(Paragraph(str(resultado_final), celda_style))
                
                tabla_data.append(fila)
            
            # Crear tabla
            tabla = Table(tabla_data, colWidths=col_widths)
            tabla.setStyle(TableStyle([
                # Cabecera
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3880c7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alineación superior para texto con múltiples líneas
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('LEFTPADDING', (0, 1), (-1, -1), 4),
                ('RIGHTPADDING', (0, 1), (-1, -1), 4),
            ]))
            
            elementos.append(tabla)
            elementos.append(Spacer(1, 0.5*cm))
        
        # PageBreak entre puestos (excepto el último)
        if puesto != orden_puestos[-1] and (orden_puestos.index(puesto) + 1 < len(orden_puestos)):
            # Solo agregar PageBreak si hay más puestos con datos
            hay_mas_puestos = any(p in controles_por_puesto for p in orden_puestos[orden_puestos.index(puesto) + 1:])
            if hay_mas_puestos:
                elementos.append(PageBreak())
    
    # Construir PDF
    doc.build(elementos)
    return buffer

# ====================================================================================
# ENDPOINTS PARA GESTIÓN DE PROVEEDORES
# ====================================================================================

@app.route('/api/proveedores', methods=['GET'])
def get_proveedores():
    """Obtener lista de todos los proveedores"""
    try:
        print("🔍 Obteniendo lista de proveedores...")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                print("❌ Error de conexión a la base de datos")
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [ID_Proveedor], [Nombre_Proveedor], [Direccion], [CP], [Ciudad], [CIF], [Tipo]
                FROM [Digitalizacion].[PE].[Proveedores]
                ORDER BY [Nombre_Proveedor]
            """)
            
            proveedores = []
            for row in cursor.fetchall():
                proveedores.append({
                    'id': row[0],
                    'nombre': row[1] if row[1] else '',
                    'direccion': row[2] if row[2] else '',
                    'cp': row[3] if row[3] else '',
                    'ciudad': row[4] if row[4] else '',
                    'cif': row[5] if row[5] else '',
                    'tipo': row[6] if row[6] else ''
                })
            
            print(f"✅ Se obtuvieron {len(proveedores)} proveedores")
            return jsonify(proveedores)
    
    except Exception as e:
        print(f"❌ Error obteniendo proveedores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/proveedor/<int:proveedor_id>', methods=['GET'])
def get_proveedor(proveedor_id):
    """Obtener datos de un proveedor específico"""
    try:
        print(f"🔍 Obteniendo proveedor ID: {proveedor_id}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [ID_Proveedor], [Nombre_Proveedor], [Direccion], [CP], [Ciudad], [CIF], [Tipo]
                FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (proveedor_id,))
            
            row = cursor.fetchone()
            if not row:
                print(f"❌ Proveedor {proveedor_id} no encontrado")
                return jsonify({'success': False, 'message': 'Proveedor no encontrado'}), 404
            
            proveedor = {
                'id': row[0],
                'nombre': row[1] if row[1] else '',
                'direccion': row[2] if row[2] else '',
                'cp': row[3] if row[3] else '',
                'ciudad': row[4] if row[4] else '',
                'cif': row[5] if row[5] else '',
                'tipo': row[6] if row[6] else ''
            }
            
            print(f"✅ Proveedor encontrado: {proveedor['nombre']}")
            return jsonify(proveedor)
    
    except Exception as e:
        print(f"❌ Error obteniendo proveedor {proveedor_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/crear-proveedor', methods=['POST'])
def crear_proveedor():
    """Crear un nuevo proveedor"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Datos inválidos'}), 400
            
        nombre = data.get('nombre', '').strip()
        direccion = data.get('direccion', '').strip()
        cp = data.get('cp', '').strip()
        ciudad = data.get('ciudad', '').strip()
        cif = data.get('cif', '').strip()
        tipo = data.get('tipo', '').strip()
        
        print(f"📝 Creando proveedor: {nombre}")
        
        if not nombre:
            return jsonify({'success': False, 'message': 'El nombre es obligatorio'}), 400
        
        # Validar CP si se proporciona
        if cp and not cp.isdigit():
            return jsonify({'success': False, 'message': 'El CP debe ser numérico'}), 400
            
        if cp and len(cp) != 5:
            return jsonify({'success': False, 'message': 'El CP debe tener 5 dígitos'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Proveedores] 
                ([Nombre_Proveedor], [Direccion], [CP], [Ciudad], [CIF], [Tipo])
                VALUES (?, ?, ?, ?, ?, ?)
            """, (nombre, direccion if direccion else None, cp if cp else None, ciudad if ciudad else None, cif if cif else None, tipo if tipo else None))
            
            conn.commit()
            print(f"✅ Proveedor creado exitosamente: {nombre}")
            return jsonify({'success': True, 'message': 'Proveedor creado exitosamente'})
    
    except Exception as e:
        print(f"❌ Error creando proveedor: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/proveedor/<int:proveedor_id>', methods=['PUT'])
def actualizar_proveedor(proveedor_id):
    """Actualizar datos de un proveedor"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Datos inválidos'}), 400
            
        nombre = data.get('nombre', '').strip()
        direccion = data.get('direccion', '').strip()
        cp = data.get('cp', '').strip()
        ciudad = data.get('ciudad', '').strip()
        cif = data.get('cif', '').strip()
        tipo = data.get('tipo', '').strip()
        
        print(f"🔄 Actualizando proveedor ID: {proveedor_id}")
        
        if not nombre:
            return jsonify({'success': False, 'message': 'El nombre es obligatorio'}), 400
            
        # Validar CP si se proporciona
        if cp and not cp.isdigit():
            return jsonify({'success': False, 'message': 'El CP debe ser numérico'}), 400
            
        if cp and len(cp) != 5:
            return jsonify({'success': False, 'message': 'El CP debe tener 5 dígitos'}), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el proveedor existe
            cursor.execute("""
                SELECT [ID_Proveedor] FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (proveedor_id,))
            
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Proveedor no encontrado'}), 404
            
            # Actualizar el proveedor
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Proveedores]
                SET [Nombre_Proveedor] = ?, [Direccion] = ?, [CP] = ?, [Ciudad] = ?, [CIF] = ?, [Tipo] = ?
                WHERE [ID_Proveedor] = ?
            """, (nombre, direccion if direccion else None, cp if cp else None, ciudad if ciudad else None, cif if cif else None, tipo if tipo else None, proveedor_id))
            
            conn.commit()
            print(f"✅ Proveedor {proveedor_id} actualizado exitosamente")
            return jsonify({'success': True, 'message': 'Proveedor actualizado exitosamente'})
    
    except Exception as e:
        print(f"❌ Error actualizando proveedor {proveedor_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/proveedor/<int:proveedor_id>', methods=['DELETE'])
def eliminar_proveedor(proveedor_id):
    """Eliminar un proveedor"""
    try:
        print(f"🗑️ Eliminando proveedor ID: {proveedor_id}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el proveedor existe
            cursor.execute("""
                SELECT [Nombre_Proveedor] FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (proveedor_id,))
            
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Proveedor no encontrado'}), 404
            
            nombre_proveedor = row[0]
            
            # Eliminar el proveedor
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Proveedores]
                WHERE [ID_Proveedor] = ?
            """, (proveedor_id,))
            
            conn.commit()
            print(f"✅ Proveedor '{nombre_proveedor}' eliminado exitosamente")
            return jsonify({'success': True, 'message': 'Proveedor eliminado exitosamente'})
    
    except Exception as e:
        print(f"❌ Error eliminando proveedor {proveedor_id}: {e}")
        import traceback
        traceback.print_exc()
# ====================================================================================
# ENDPOINTS PARA GESTIÓN DE PROVEEDORES-PAUTAS (TABLA COMBINADA)
# ====================================================================================

@app.route('/api/get-puestos-dropdown', methods=['GET'])
def get_puestos_dropdown():
    """Obtener puestos únicos con ID y nombre para dropdown"""
    request_started_at = time.perf_counter()
    try:
        print("🔍 Obteniendo puestos para dropdown...")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            try:
                cursor.timeout = 8
            except Exception:
                pass
            
            # Obtener puestos con nombre válido para dropdown
            cursor.execute("""
                SELECT DISTINCT ID_Puesto, Nombre_Puesto
                FROM [Digitalizacion].[PE].[Puestos]
                WHERE Nombre_Puesto IS NOT NULL AND LTRIM(RTRIM(Nombre_Puesto)) <> ''
                ORDER BY Nombre_Puesto
            """)
            
            resultados = cursor.fetchall()
            
            puestos = []
            for row in resultados:
                puestos.append({
                    'id_puesto': row[0],
                    'nombre_puesto': row[1]
                })
            
            print(f"📊 Puestos obtenidos para dropdown: {len(puestos)} en {time.perf_counter() - request_started_at:.2f}s")
            
            return jsonify({
                'success': True,
                'puestos': puestos,
                'message': f'Se obtuvieron {len(puestos)} puestos exitosamente'
            })
    
    except pyodbc.Error as e:
        print(f"❌ Error SQL obteniendo puestos dropdown tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'Error consultando puestos en base de datos'
        }), 500
    except Exception as e:
        print(f"❌ Error obteniendo puestos dropdown tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas', methods=['GET'])
def get_proveedores_pautas():
    """Obtener lista completa de relaciones proveedores-pautas con nombres legibles"""
    try:
        print("🔍 Obteniendo lista de proveedores-pautas...")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # JOIN para obtener nombres legibles en lugar de IDs
            cursor.execute("""
                SELECT 
                    pp.ID_Prov_Pauta,
                    pp.ID_Proveedor,
                    prov.Nombre_Proveedor,
                    pp.ID_Pauta,
                    pauta.Nombre_Pauta,
                    pp.ID_Puesto,
                    puesto.Nombre_Puesto,
                    pp.Codigo,
                    pp.Precio,
                    pp.Peso,
                    pp.pdf
                FROM [Digitalizacion].[PE].[Proveedores_Pautas] pp
                LEFT JOIN [Digitalizacion].[PE].[Proveedores] prov ON pp.ID_Proveedor = prov.ID_Proveedor
                LEFT JOIN (
                    SELECT DISTINCT ID_Pauta, Nombre_Pauta 
                    FROM [Digitalizacion].[PE].[Pautas]
                ) pauta ON pp.ID_Pauta = pauta.ID_Pauta
                LEFT JOIN (
                    SELECT DISTINCT ID_Puesto, Nombre_Puesto 
                    FROM [Digitalizacion].[PE].[Puestos]
                ) puesto ON pp.ID_Puesto = puesto.ID_Puesto
                ORDER BY prov.Nombre_Proveedor, pauta.Nombre_Pauta, puesto.Nombre_Puesto
            """)
            
            resultados = cursor.fetchall()
            
            registros = []
            for row in resultados:
                registros.append({
                    'id': row[0],
                    'id_proveedor': row[1],
                    'nombre_proveedor': row[2] or 'Proveedor no encontrado',
                    'id_pauta': row[3],
                    'nombre_pauta': row[4] or 'Pauta no encontrada',
                    'id_puesto': row[5],
                    'nombre_puesto': row[6] or 'Puesto no encontrado',
                    'codigo': row[7],
                    'precio': float(row[8]) if row[8] is not None else None,
                    'peso': float(row[9]) if row[9] is not None else None,
                    'pdf': bool(row[10]) if row[10] is not None else False
                })
            
            print(f"📊 Registros obtenidos: {len(registros)}")
            
            return jsonify({
                'success': True,
                'registros': registros,
                'total': len(registros)
            })
    
    except Exception as e:
        print(f"❌ Error obteniendo proveedores-pautas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas', methods=['POST'])
def crear_proveedor_pauta_completo():
    """Crear nueva relación proveedor-pauta-puesto con validación de unicidad"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
            
        id_proveedor = data.get('id_proveedor')
        id_pauta = data.get('id_pauta')
        id_puesto = data.get('id_puesto', '').strip() if data.get('id_puesto') else ''
        codigo = data.get('codigo', '').strip() if data.get('codigo') else None
        # Usar 0.00 como valor por defecto si no se proporciona (la tabla no admite NULL)
        precio = data.get('precio') if data.get('precio') is not None else 0.00
        peso = data.get('peso') if data.get('peso') is not None else 0.00
        pdf = 1 if data.get('pdf') else 0
        
        print(f"📝 Creando registro proveedor-pauta-puesto:")
        print(f"   - ID_Proveedor: {id_proveedor}")
        print(f"   - ID_Pauta: {id_pauta}")
        print(f"   - ID_Puesto: {id_puesto}")
        print(f"   - Codigo: {codigo}")
        print(f"   - Precio: {precio}")
        print(f"   - Peso: {peso}")
        print(f"   - PDF: {pdf}")
        
        # Validar campos obligatorios
        if not id_proveedor or not id_pauta or not id_puesto:
            return jsonify({
                'success': False, 
                'message': 'Proveedor, Pauta y Puesto son obligatorios'
            }), 400
        
        # Validar que precio y peso sean positivos
        if precio < 0:
            return jsonify({
                'success': False, 
                'message': 'El precio debe ser positivo'
            }), 400
            
        if peso < 0:
            return jsonify({
                'success': False, 
                'message': 'El peso debe ser positivo'
            }), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que no existe la combinación (unicidad)
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Proveedor = ? AND ID_Pauta = ? AND ID_Puesto = ?
            """, (id_proveedor, id_pauta, id_puesto))
            
            if cursor.fetchone()[0] > 0:
                return jsonify({
                    'success': False,
                    'message': 'Ya existe un registro con esta combinación de Proveedor-Pauta-Puesto'
                }), 400
            
            # Insertar nuevo registro
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Proveedores_Pautas]
                (ID_Proveedor, ID_Pauta, ID_Puesto, Codigo, Precio, Peso, pdf)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (id_proveedor, id_pauta, id_puesto, codigo, precio, peso, pdf))
            
            conn.commit()
            
            print(f"✅ Registro creado exitosamente")
            
            return jsonify({
                'success': True,
                'message': 'Registro creado exitosamente'
            })
    
    except Exception as e:
        print(f"❌ Error creando registro proveedor-pauta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/<int:registro_id>', methods=['GET'])
def get_proveedor_pauta_individual(registro_id):
    """Obtener datos de un registro específico"""
    try:
        print(f"🔍 Obteniendo registro ID: {registro_id}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    pp.ID_Prov_Pauta,
                    pp.ID_Proveedor,
                    prov.Nombre_Proveedor,
                    pp.ID_Pauta,
                    pauta.Nombre_Pauta,
                    pp.ID_Puesto,
                    puesto.Nombre_Puesto,
                    pp.Codigo,
                    pp.Precio,
                    pp.Peso,
                    pp.pdf
                FROM [Digitalizacion].[PE].[Proveedores_Pautas] pp
                LEFT JOIN [Digitalizacion].[PE].[Proveedores] prov ON pp.ID_Proveedor = prov.ID_Proveedor
                LEFT JOIN (
                    SELECT DISTINCT ID_Pauta, Nombre_Pauta 
                    FROM [Digitalizacion].[PE].[Pautas]
                ) pauta ON pp.ID_Pauta = pauta.ID_Pauta
                LEFT JOIN (
                    SELECT DISTINCT ID_Puesto, Nombre_Puesto 
                    FROM [Digitalizacion].[PE].[Puestos]
                ) puesto ON pp.ID_Puesto = puesto.ID_Puesto
                WHERE pp.ID_Prov_Pauta = ?
            """, (registro_id,))
            
            resultado = cursor.fetchone()
            
            if not resultado:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            
            registro = {
                'id': resultado[0],
                'id_proveedor': resultado[1],
                'nombre_proveedor': resultado[2],
                'id_pauta': resultado[3],
                'nombre_pauta': resultado[4],
                'id_puesto': resultado[5],
                'nombre_puesto': resultado[6],
                'codigo': resultado[7],
                'precio': float(resultado[8]) if resultado[8] is not None else None,
                'peso': float(resultado[9]) if resultado[9] is not None else None,
                'pdf': bool(resultado[10]) if resultado[10] is not None else False
            }
            
            return jsonify({
                'success': True,
                'registro': registro
            })
    
    except Exception as e:
        print(f"❌ Error obteniendo registro {registro_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/<int:registro_id>', methods=['PUT'])
def actualizar_proveedor_pauta(registro_id):
    """Actualizar registro proveedor-pauta existente"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
            
        id_proveedor = data.get('id_proveedor')
        id_pauta = data.get('id_pauta')
        id_puesto = data.get('id_puesto', '').strip() if data.get('id_puesto') else ''
        codigo = data.get('codigo', '').strip() if data.get('codigo') else None
        # Usar 0.00 como valor por defecto si no se proporciona (la tabla no admite NULL)
        precio = data.get('precio') if data.get('precio') is not None else 0.00
        peso = data.get('peso') if data.get('peso') is not None else 0.00
        pdf = 1 if data.get('pdf') else 0
        
        print(f"🔄 Actualizando registro ID: {registro_id}")
        
        # Validar campos obligatorios
        if not id_proveedor or not id_pauta or not id_puesto:
            return jsonify({
                'success': False, 
                'message': 'Proveedor, Pauta y Puesto son obligatorios'
            }), 400
        
        # Validar que precio y peso sean positivos
        if precio < 0:
            return jsonify({
                'success': False, 
                'message': 'El precio debe ser positivo'
            }), 400
            
        if peso < 0:
            return jsonify({
                'success': False, 
                'message': 'El peso debe ser positivo'
            }), 400
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el registro existe
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Prov_Pauta = ?
            """, (registro_id,))
            
            if cursor.fetchone()[0] == 0:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            
            # Verificar que no existe otra combinación igual (excepto el registro actual)
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Proveedor = ? AND ID_Pauta = ? AND ID_Puesto = ? AND ID_Prov_Pauta != ?
            """, (id_proveedor, id_pauta, id_puesto, registro_id))
            
            if cursor.fetchone()[0] > 0:
                return jsonify({
                    'success': False,
                    'message': 'Ya existe otro registro con esta combinación de Proveedor-Pauta-Puesto'
                }), 400
            
            # Actualizar registro
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Proveedores_Pautas]
                SET ID_Proveedor = ?, ID_Pauta = ?, ID_Puesto = ?, Codigo = ?, Precio = ?, Peso = ?, pdf = ?
                WHERE ID_Prov_Pauta = ?
            """, (id_proveedor, id_pauta, id_puesto, codigo, precio, peso, pdf, registro_id))
            
            conn.commit()
            
            print(f"✅ Registro {registro_id} actualizado")
            
            return jsonify({
                'success': True,
                'message': 'Registro actualizado exitosamente'
            })
    
    except Exception as e:
        print(f"❌ Error actualizando registro {registro_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/<int:registro_id>', methods=['DELETE'])
def eliminar_proveedor_pauta(registro_id):
    """Eliminar registro proveedor-pauta"""
    try:
        print(f"🗑️ Eliminando registro ID: {registro_id}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el registro existe y obtener info para log
            cursor.execute("""
                SELECT 
                    prov.Nombre_Proveedor,
                    pauta.Nombre_Pauta,
                    puesto.Nombre_Puesto
                FROM [Digitalizacion].[PE].[Proveedores_Pautas] pp
                LEFT JOIN [Digitalizacion].[PE].[Proveedores] prov ON pp.ID_Proveedor = prov.ID_Proveedor
                LEFT JOIN (
                    SELECT DISTINCT ID_Pauta, Nombre_Pauta 
                    FROM [Digitalizacion].[PE].[Pautas]
                ) pauta ON pp.ID_Pauta = pauta.ID_Pauta
                LEFT JOIN (
                    SELECT DISTINCT ID_Puesto, Nombre_Puesto 
                    FROM [Digitalizacion].[PE].[Puestos]
                ) puesto ON pp.ID_Puesto = puesto.ID_Puesto
                WHERE pp.ID_Prov_Pauta = ?
            """, (registro_id,))
            
            resultado = cursor.fetchone()
            if not resultado:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            
            nombre_info = f"{resultado[0]} - {resultado[1]} - {resultado[2]}"
            
            # Eliminar registro
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Prov_Pauta = ?
            """, (registro_id,))
            
            conn.commit()
            
            print(f"✅ Registro eliminado: {nombre_info} (ID: {registro_id})")
            
            return jsonify({
                'success': True,
                'message': f'Registro "{nombre_info}" eliminado exitosamente'
            })
    
    except Exception as e:
        print(f"❌ Error eliminando registro {registro_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/<int:registro_id>/pdf', methods=['PATCH'])
def toggle_pdf_proveedor_pauta(registro_id):
    """Actualizar solo el campo pdf de un registro"""
    try:
        data = request.get_json()
        if data is None:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        pdf_value = 1 if data.get('pdf') else 0
        
        print(f"🔄 Actualizando PDF del registro ID: {registro_id} -> {pdf_value}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el registro existe
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Prov_Pauta = ?
            """, (registro_id,))
            
            if cursor.fetchone()[0] == 0:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            
            # Actualizar solo el campo pdf
            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Proveedores_Pautas]
                SET pdf = ?
                WHERE ID_Prov_Pauta = ?
            """, (pdf_value, registro_id))
            
            conn.commit()
            
            print(f"✅ Campo PDF actualizado para registro {registro_id}")
            
            return jsonify({
                'success': True,
                'message': 'Estado PDF actualizado correctamente',
                'pdf': bool(pdf_value)
            })
    
    except Exception as e:
        print(f"❌ Error actualizando PDF del registro {registro_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/<int:registro_id>/duplicate', methods=['POST'])
def duplicar_proveedor_pauta(registro_id):
    """Duplicar registro proveedor-pauta con nuevos valores"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        print(f"📋 Duplicando registro ID: {registro_id}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos del registro original
            cursor.execute("""
                SELECT ID_Proveedor, ID_Pauta, ID_Puesto, Precio, Peso
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Prov_Pauta = ?
            """, (registro_id,))
            
            original = cursor.fetchone()
            if not original:
                return jsonify({'success': False, 'message': 'Registro original no encontrado'}), 404
            
            # Usar datos del formulario o mantener originales
            id_proveedor = data.get('id_proveedor', original[0])
            id_pauta = data.get('id_pauta', original[1])
            id_puesto = data.get('id_puesto', original[2])
            precio = data.get('precio', original[3])
            peso = data.get('peso', original[4])
            
            # Validar combinación única
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Proveedores_Pautas]
                WHERE ID_Proveedor = ? AND ID_Pauta = ? AND ID_Puesto = ?
            """, (id_proveedor, id_pauta, id_puesto))
            
            if cursor.fetchone()[0] > 0:
                return jsonify({
                    'success': False,
                    'message': 'Ya existe un registro con esta combinación de Proveedor-Pauta-Puesto'
                }), 400
            
            # Crear duplicado
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Proveedores_Pautas]
                (ID_Proveedor, ID_Pauta, ID_Puesto, Precio, Peso)
                VALUES (?, ?, ?, ?, ?)
            """, (id_proveedor, id_pauta, id_puesto, precio, peso))
            
            conn.commit()
            
            print(f"✅ Registro duplicado exitosamente")
            
            return jsonify({
                'success': True,
                'message': 'Registro duplicado exitosamente'
            })
    
    except Exception as e:
        print(f"❌ Error duplicando registro {registro_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/proveedores-pautas/export', methods=['GET'])
def exportar_proveedores_pautas():
    """Exportar datos a Excel (.xlsx)"""
    try:
        print("📊 Iniciando exportación a Excel...")
        
        # Importar openpyxl para Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({
                'success': False,
                'message': 'Error: openpyxl no está instalado. Ejecutar: pip install openpyxl'
            }), 500
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener datos para exportar
            cursor.execute("""
                SELECT 
                    pp.ID_Prov_Pauta,
                    ISNULL(prov.Nombre_Proveedor, 'Sin proveedor') as Proveedor,
                    ISNULL(pauta.Nombre_Pauta, 'Sin pauta') as Pauta,
                    ISNULL(puesto.Nombre_Puesto, 'Sin puesto') as Puesto,
                    pp.Precio,
                    pp.Peso
                FROM [Digitalizacion].[PE].[Proveedores_Pautas] pp
                LEFT JOIN [Digitalizacion].[PE].[Proveedores] prov ON pp.ID_Proveedor = prov.ID_Proveedor
                LEFT JOIN (
                    SELECT DISTINCT ID_Pauta, Nombre_Pauta 
                    FROM [Digitalizacion].[PE].[Pautas]
                ) pauta ON pp.ID_Pauta = pauta.ID_Pauta
                LEFT JOIN (
                    SELECT DISTINCT ID_Puesto, Nombre_Puesto 
                    FROM [Digitalizacion].[PE].[Puestos]
                ) puesto ON pp.ID_Puesto = puesto.ID_Puesto
                ORDER BY prov.Nombre_Proveedor, pauta.Nombre_Pauta, puesto.Nombre_Puesto
            """)
            
            resultados = cursor.fetchall()
            
            # Crear libro Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Proveedores-Pautas"
            
            # Encabezados
            encabezados = ['ID', 'Proveedor', 'Pauta', 'Puesto', 'Precio', 'Peso']
            ws.append(encabezados)
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for row in resultados:
                precio_str = f"{row[4]:.2f}" if row[4] is not None else ""
                peso_str = f"{row[5]:.2f}" if row[5] is not None else ""
                
                ws.append([
                    row[0],  # ID
                    row[1],  # Proveedor
                    row[2],  # Pauta
                    row[3],  # Puesto
                    precio_str,  # Precio
                    peso_str   # Peso
                ])
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"proveedores_pautas_{timestamp}.xlsx"
            filepath = os.path.join(RUTA_PDFS, filename)
            
            wb.save(filepath)
            
            print(f"✅ Excel generado: {filepath}")
            
            # Enviar archivo
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    except Exception as e:
        print(f"❌ Error exportando a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }), 500


@app.route('/api/crear-proveedor-pauta', methods=['POST'])
def crear_proveedor_pauta():
    """Crear relación entre proveedor y pauta (LEGACY - mantenido por compatibilidad)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Datos inválidos'}), 400
            
        proveedor_id = data.get('proveedor_id')
        pauta_id = data.get('pauta_id')
        
        print(f"🔗 Endpoint LEGACY detectado: crear-proveedor-pauta")
        
        # Este endpoint está deprecado - redirigir al usuario al nuevo sistema
        return jsonify({
            'success': False,
            'message': 'Este endpoint está obsoleto. Usa /api/proveedores-pautas con campos completos (proveedor, pauta, puesto, precio, peso)'
        }), 400
    
    except Exception as e:
        print(f"❌ Error en endpoint legacy: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ====================================================================================
# 🆕 ENDPOINTS PARA GESTIÓN DE CORRECCIONES (MODIFICAR/ELIMINAR RESULTADOS)
# ====================================================================================

@app.route('/api/correcciones/armarios', methods=['GET'])
def get_armarios_correcciones():
    """Obtiene lista de armarios para el dropdown de correcciones"""
    try:
        print("🔍 Obteniendo lista de armarios para correcciones...")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500
            
            cursor = conn.cursor()
            
            # Obtener armarios ordenados por ID_Pedido descendente
            cursor.execute("""
                SELECT TOP 1000 ID_Pedido, Armario, NumPedido
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE Armario IS NOT NULL AND Armario != ''
                ORDER BY ID_Pedido DESC
            """)
            
            armarios = []
            for row in cursor.fetchall():
                armarios.append({
                    'id_pedido': row[0],
                    'armario': row[1],
                    'num_pedido': row[2] if row[2] else ''
                })
            
            print(f"📊 {len(armarios)} armarios obtenidos")
            
            return jsonify({
                'success': True,
                'armarios': armarios
            })
    
    except Exception as e:
        print(f"❌ Error obteniendo armarios: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/correcciones/armario/<int:id_pedido>', methods=['GET'])
def get_controles_armario(id_pedido):
    """Obtiene TODOS los registros del HISTÓRICO de cambios en los controles de un armario"""
    try:
        print(f"🔍 Obteniendo histórico completo del armario ID_Pedido: {id_pedido}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Paso 1: Obtener información del armario
            cursor.execute("""
                SELECT Armario, Nombre_Pauta
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))
            
            armario_row = cursor.fetchone()
            if not armario_row:
                return jsonify({'success': False, 'message': 'Armario no encontrado'}), 404
            
            armario = armario_row[0]
            nombre_pauta = armario_row[1]
            
            # Paso 2: Obtener CADA FILA del histórico de DatosUser_Historico
            # Ordenado por FechaHistorico DESC (más recientes primero)
            cursor.execute("""
                SELECT 
                    h.[ID_Control],
                    h.[Resultado],
                    h.[Comentario],
                    h.[Resultado_txt],
                    h.[User],
                    h.[FechaHistorico],
                    h.[Accion],
                    h.[FechaOriginal],
                    ISNULL(pu.[Nombre_Puesto], ''),
                    ISNULL(c.[DescripcionControl], '')
                FROM [Digitalizacion].[PE].[DatosUser_Historico] h
                LEFT JOIN [Digitalizacion].[PE].[Controles] c ON h.[ID_Control] = c.[Id_Control]
                LEFT JOIN [Digitalizacion].[PE].[Puesto] pu ON c.[ID_Puesto] = pu.[ID_Puesto]
                WHERE h.[ID_Pedido] = ?
                ORDER BY h.[FechaHistorico] DESC
            """, (id_pedido,))
            
            rows_hist = cursor.fetchall()
            
            # Procesar: CADA fila del histórico es un registro separado
            controles = []
            
            for row in rows_hist:
                # Convertir fecha
                fecha_str = ''
                if row[5]:
                    if isinstance(row[5], datetime):
                        fecha_str = row[5].strftime('%a, %d %b %Y %H:%M:%S GMT')
                    else:
                        fecha_str = str(row[5])
                
                controles.append({
                    'id_control': row[0],
                    'resultado': row[1] if row[1] else '',
                    'comentario': row[2] if row[2] else '',
                    'resultado_txt': row[3] if row[3] else '',
                    'usuario': row[4] if row[4] else '',
                    'fecha_registro': fecha_str,
                    'puesto': row[8] if row[8] else '',
                    'descripcion_control': row[9] if row[9] else '',
                    'accion': row[6] if row[6] else ''
                })
            
            print(f"✅ {len(controles)} registros de histórico obtenidos para armario {armario}")
            
            return jsonify({
                'success': True,
                'armario': armario,
                'nombre_pauta': nombre_pauta,
                'id_pedido': id_pedido,
                'controles': controles
            })
    
    except Exception as e:
        print(f"❌ Error obteniendo histórico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/correcciones/actualizar-resultado/<int:id_pedido>/<int:id_control>', methods=['POST'])
def actualizar_resultado_control(id_pedido, id_control):
    """Actualiza el resultado de un control (OK/NOK)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Body vacío'}), 400
        
        resultado = data.get('resultado', '').strip().upper()
        resultado_txt = data.get('resultado_txt', '').strip()
        comentario = data.get('comentario', '').strip()
        
        print(f"📝 Actualizando resultado - ID_Pedido: {id_pedido}, ID_Control: {id_control}, Resultado: {resultado}, Comentario recibido: '{comentario}'")
        
        # Validar resultado
        if resultado not in ['OK', 'NOK']:
            return jsonify({'success': False, 'message': 'Resultado debe ser OK o NOK'}), 400
        
        # 🆕 Si el resultado es OK, vaciar el comentario
        if resultado == 'OK':
            comentario = ''
            print(f"   ℹ️ Resultado OK → Comentario vaciado")
        else:
            # Si es NOK, mantener el comentario proporcionado (o vacío si no hay)
            comentario = comentario if comentario else ''
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Verificar que el registro existe
            cursor.execute("""
                SELECT ID_DatosUser FROM [Digitalizacion].[PE].[DatosUser]
                WHERE ID_Pedido = ? AND ID_Control = ?
            """, (id_pedido, id_control))
            
            registro = cursor.fetchone()
            
            if registro:
                # Actualizar registro existente
                id_datos_user = registro[0]
                cursor.execute("""
                    UPDATE [Digitalizacion].[PE].[DatosUser]
                    SET Resultado = ?, Resultado_txt = ?, Comentario = ?, FechaRegistro = GETDATE()
                    WHERE ID_DatosUser = ?
                """, (resultado, resultado_txt, comentario, id_datos_user))
                print(f"   ✅ UPDATE: ID_DatosUser={id_datos_user}, Comentario={comentario}")
            else:
                # Crear nuevo registro
                cursor.execute("""
                    INSERT INTO [Digitalizacion].[PE].[DatosUser]
                    (ID_Pedido, ID_Control, Resultado, Resultado_txt, Comentario, FechaRegistro, User)
                    VALUES (?, ?, ?, ?, ?, GETDATE(), ?)
                """, (id_pedido, id_control, resultado, resultado_txt, comentario, 'CORRECCIONES'))
                print(f"   ✅ Nuevo registro creado")
            
            conn.commit()
            
            print(f"✅ Resultado actualizado correctamente")
            
            return jsonify({
                'success': True,
                'message': 'Resultado actualizado correctamente'
            })
    
    except Exception as e:
        print(f"❌ Error actualizando resultado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/correcciones/eliminar', methods=['POST'])
def eliminar_controles():
    """Elimina registros de controles seleccionados"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Body vacío'}), 400
        
        id_pedido = data.get('id_pedido')
        id_controles = data.get('id_controles', [])
        
        if not id_pedido or not id_controles:
            return jsonify({'success': False, 'message': 'ID_Pedido e ID_Controles son requeridos'}), 400
        
        print(f"🗑️ Eliminando controles - ID_Pedido: {id_pedido}, Controles: {id_controles}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            eliminados = 0
            
            for id_control in id_controles:
                try:
                    # Eliminar de DatosUser
                    cursor.execute("""
                        DELETE FROM [Digitalizacion].[PE].[DatosUser]
                        WHERE ID_Pedido = ? AND ID_Control = ?
                    """, (id_pedido, id_control))
                    
                    if cursor.rowcount > 0:
                        eliminados += 1
                        print(f"   ✅ Eliminado registro: Pedido {id_pedido} - Control {id_control}")
                
                except Exception as e:
                    print(f"   ⚠️ Error eliminando control {id_control}: {e}")
                    continue
            
            conn.commit()
            
            print(f"✅ {eliminados} registros eliminados")
            
            return jsonify({
                'success': True,
                'message': f'{eliminados} registros eliminados correctamente',
                'eliminados': eliminados
            })
    
    except Exception as e:
        print(f"❌ Error eliminando controles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ====================================================================================
# 🆕 ENDPOINTS PARA GESTIÓN DE DATOS DEL ARMARIO (EDICIÓN Y ELIMINACIÓN)
# ====================================================================================

@app.route('/api/correcciones/pedido/<int:id_pedido>', methods=['GET'])
def get_pedido_correcciones(id_pedido):
    """Obtiene datos del pedido para modal Editar Armario en Correcciones"""
    try:
        print(f"🔍 [Correcciones] Obtener pedido ID: {id_pedido}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                SELECT ID_Pedido, Armario, NumPedido, Referencia, Nombre_Pauta
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))

            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404

            nombre_pauta = row[4] or ''
            referencias_validas = []

            if nombre_pauta:
                cursor.execute("""
                    SELECT TOP 1 ID_Pauta
                    FROM [Digitalizacion].[PE].[Pautas]
                    WHERE Nombre_Pauta = ?
                    ORDER BY ID_Pauta DESC
                """, (nombre_pauta,))

                pauta_row = cursor.fetchone()
                if pauta_row:
                    cursor.execute("""
                        SELECT Referencia
                        FROM [Digitalizacion].[PE].[Pauta_Referencias]
                        WHERE ID_Pauta = ?
                        ORDER BY Referencia ASC
                    """, (pauta_row[0],))
                    referencias_validas = [ref_row[0] for ref_row in cursor.fetchall() if ref_row[0]]

            pedido = {
                'id_pedido': row[0],
                'armario': row[1] or '',
                'num_pedido': row[2] or '',
                'referencia': row[3] or '',
                'nombre_pauta': nombre_pauta,
                'referencias_validas': referencias_validas
            }

            return jsonify({'success': True, 'pedido': pedido})

    except Exception as e:
        print(f"❌ Error [Correcciones] get pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/correcciones/pedido/<int:id_pedido>', methods=['PUT'])
def actualizar_pedido_correcciones(id_pedido):
    """Actualiza Armario, NumPedido y Referencia para Correcciones"""
    try:
        data = request.get_json() or {}
        armario = (data.get('armario') or '').strip()
        num_pedido = (data.get('num_pedido') or '').strip()
        referencia = (data.get('referencia') or '').strip()

        if not armario:
            return jsonify({'success': False, 'message': 'El campo Armario es obligatorio'}), 400

        if not referencia:
            return jsonify({'success': False, 'message': 'La referencia es obligatoria'}), 400

        print(f"📝 [Correcciones] Actualizar pedido {id_pedido}: Armario='{armario}', NumPedido='{num_pedido}', Referencia='{referencia}'")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()

            # Validar existencia
            cursor.execute("""
                SELECT Nombre_Pauta, Referencia
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))

            pedido_row = cursor.fetchone()
            if not pedido_row:
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404

            nombre_pauta = pedido_row[0] or ''
            referencia_actual = pedido_row[1] or ''

            # Validar unicidad de Armario (NumPedido sí puede repetirse)
            cursor.execute("""
                SELECT TOP 1 ID_Pedido
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE Armario = ? AND ID_Pedido <> ?
            """, (armario, id_pedido))

            duplicado = cursor.fetchone()
            if duplicado:
                return jsonify({
                    'success': False,
                    'message': f'El armario "{armario}" ya existe en otro pedido (ID: {duplicado[0]}).'
                }), 409

            if nombre_pauta:
                cursor.execute("""
                    SELECT TOP 1 ID_Pauta
                    FROM [Digitalizacion].[PE].[Pautas]
                    WHERE Nombre_Pauta = ?
                    ORDER BY ID_Pauta DESC
                """, (nombre_pauta,))
                pauta_row = cursor.fetchone()

                if pauta_row:
                    cursor.execute("""
                        SELECT COUNT(*)
                        FROM [Digitalizacion].[PE].[Pauta_Referencias]
                        WHERE ID_Pauta = ? AND Referencia = ?
                    """, (pauta_row[0], referencia))
                    referencia_valida = cursor.fetchone()[0] > 0

                    if not referencia_valida and referencia != referencia_actual:
                        return jsonify({
                            'success': False,
                            'message': f'La referencia "{referencia}" no está configurada para la pauta "{nombre_pauta}".'
                        }), 400

            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Pedido]
                SET Armario = ?, NumPedido = ?, Referencia = ?
                WHERE ID_Pedido = ?
            """, (armario, num_pedido, referencia, id_pedido))

            conn.commit()

            return jsonify({
                'success': True,
                'message': 'Pedido actualizado correctamente',
                'pedido': {
                    'id_pedido': id_pedido,
                    'armario': armario,
                    'num_pedido': num_pedido,
                    'referencia': referencia
                }
            })

    except Exception as e:
        print(f"❌ Error [Correcciones] actualizar pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/gestion/checklist-documentos', methods=['GET'])
def get_gestion_checklist_documentos():
    """Obtiene los registros de Checklist_Documentos para la vista de Pedidos Checklist."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [ID], [NumPedido], [NumArmario], [NombreArchivo], [Fecha_Modificacion_Archivo], [Fecha_Registro]
                FROM [Digitalizacion].[PE].[Checklist_Documentos]
                ORDER BY [Fecha_Modificacion_Archivo] DESC, [NumPedido] ASC
            """)
            rows = cursor.fetchall()
            docs = []
            for r in rows:
                fecha_mod = ''
                if r[4]:
                    try:
                        fecha_mod = r[4].strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        fecha_mod = str(r[4])
                fecha_reg = ''
                if r[5]:
                    try:
                        fecha_reg = r[5].strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        fecha_reg = str(r[5])
                docs.append({
                    'id': r[0],
                    'num_pedido': r[1] or '',
                    'num_armario': r[2] or '',
                    'nombre_archivo': r[3] or '',
                    'fecha_modificacion': fecha_mod,
                    'fecha_registro': fecha_reg
                })
            return jsonify({'success': True, 'docs': docs, 'total': len(docs)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/checklist-documentos/<int:doc_id>', methods=['DELETE'])
def delete_gestion_checklist_documento(doc_id):
    """Elimina un registro de Checklist_Documentos por su ID y también el certificado de pintura asociado por NumArmario."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [NombreArchivo], [NumArmario] FROM [Digitalizacion].[PE].[Checklist_Documentos] WHERE [ID] = ?
            """, (doc_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            nombre_archivo = row[0]
            num_armario    = row[1]

            # Eliminar certificados de pintura asociados al mismo NumArmario
            cert_eliminados = 0
            if num_armario:
                cursor.execute("""
                    SELECT COUNT(*) FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos]
                    WHERE [NumArmario] = ?
                """, (num_armario,))
                cert_eliminados = cursor.fetchone()[0]
                if cert_eliminados > 0:
                    cursor.execute("""
                        DELETE FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos]
                        WHERE [NumArmario] = ?
                    """, (num_armario,))
                    print(f'🗑️ [Gestión] Certificados_Pintura eliminados para NumArmario={num_armario}: {cert_eliminados} registro(s)')

            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Checklist_Documentos] WHERE [ID] = ?
            """, (doc_id,))
            conn.commit()
            print(f'🗑️ [Gestión] Checklist_Documentos ID={doc_id} eliminado: {nombre_archivo}')

            msg = f'Registro "{nombre_archivo}" eliminado'
            if cert_eliminados > 0:
                msg += f' (y {cert_eliminados} certificado{"s" if cert_eliminados > 1 else ""} de pintura asociado{"s" if cert_eliminados > 1 else ""})'
            return jsonify({'success': True, 'message': msg, 'cert_eliminados': cert_eliminados})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/pedidos-join', methods=['GET'])
def get_gestion_pedidos_join():
    """
    JOIN entre Checklist_Documentos y Pedido_Cantidad_SemanaEntrega_Cache por NumPedido.
    Devuelve filas con la lista de armarios e incluye trazabilidades calculadas
    usando el GAP configurado en config.json.
    """
    try:
        # Leer GAP de configuración
        cfg = _read_config()
        gap = int(cfg.get('gap_trazabilidad_semanas', 0))

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()

            # --- Cargar mapa de trazabilidad: {(semana, ano): set(DATOs)} ---
            cursor.execute("""
                SELECT [numsemana], [ano], [DATO]
                FROM [Digitalizacion].[PE].[vw_LotesChapaProduccion]
                WHERE [numsemana] IS NOT NULL AND [ano] IS NOT NULL
            """)
            traz_map = {}   # (semana:int, ano:int) -> set(dato)
            anos_con_semana53 = set()  # años que tienen semana 53 registrada
            for tr in cursor.fetchall():
                sem = int(tr[0]) if tr[0] is not None else None
                ano = int(tr[1]) if tr[1] is not None else None
                dato = str(tr[2] or '').strip()
                if sem is None or ano is None:
                    continue
                key = (sem, ano)
                traz_map.setdefault(key, set())
                if dato:
                    traz_map[key].add(dato)
                if sem == 53:
                    anos_con_semana53.add(ano)

            cursor.execute("""
                SELECT [DATO]
                FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos]
                WHERE [DATO] IS NOT NULL AND LTRIM(RTRIM([DATO])) <> ''
            """)
            certificados_chapa_existentes = {
                str(row[0]).strip()
                for row in cursor.fetchall()
                if row[0] is not None and str(row[0]).strip()
            }

            def _semanas_objetivo(semana_entrega_str, gap_semanas):
                """Devuelve lista de (semana, ano) que corresponden al target."""
                if not semana_entrega_str or gap_semanas == 0:
                    return []
                # Formato esperado: "16/2026" o "16 2026" etc.
                import re as _re
                m = _re.search(r'(\d+)[^\d]+(\d{4})', semana_entrega_str)
                if not m:
                    return []
                sem = int(m.group(1))
                ano = int(m.group(2))
                target_sem = sem - gap_semanas
                target_ano = ano
                while target_sem <= 0:
                    target_ano -= 1
                    max_sem = 53 if target_ano in anos_con_semana53 else 52
                    target_sem = max_sem + target_sem
                return [(target_sem, target_ano)]

            # --- Obtener pedidos-join ---
            cursor.execute("""
                SELECT
                    c.[NumPedido],
                    c.[NumArmario],
                    c.[Fecha_Modificacion_Archivo],
                    p.[Cantidad],
                    p.[SEMANAENTREGA],
                    p.[FechaRefresco],
                    cp.[NombreArchivo] AS CertPinturaArchivo
                FROM [Digitalizacion].[PE].[Checklist_Documentos] c
                LEFT JOIN [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache] p
                    ON c.[NumPedido] = p.[NumPedido]
                LEFT JOIN (
                    SELECT [NumArmario], MIN([NombreArchivo]) AS [NombreArchivo]
                    FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos]
                    GROUP BY [NumArmario]
                ) cp ON cp.[NumArmario] = c.[NumArmario]
                ORDER BY c.[NumPedido] ASC, c.[Fecha_Modificacion_Archivo] DESC
            """)
            rows = cursor.fetchall()
            items = []
            for r in rows:
                fecha_mod = ''
                if r[2]:
                    try:
                        fecha_mod = r[2].strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        fecha_mod = str(r[2])
                fecha_ref = ''
                fecha_ref_sort = ''
                if r[5]:
                    try:
                        fecha_ref = r[5].strftime('%d/%m/%Y %H:%M')
                        fecha_ref_sort = r[5].strftime('%Y%m%d%H%M')
                    except Exception:
                        fecha_ref = str(r[5])
                cantidad = None
                if r[3] is not None:
                    try:
                        cantidad = int(r[3])
                    except Exception:
                        cantidad = r[3]

                semana_entrega = str(r[4]).strip() if r[4] else ''

                # Calcular trazabilidades con GAP
                trazabilidades = []
                if gap > 0 and semana_entrega:
                    for key in _semanas_objetivo(semana_entrega, gap):
                        datos = traz_map.get(key, set())
                        trazabilidades.extend(datos)
                    trazabilidades = sorted(set(trazabilidades))

                certificados_chapa_encontrados = [
                    trazabilidad
                    for trazabilidad in trazabilidades
                    if trazabilidad in certificados_chapa_existentes
                ]
                certificados_chapa_faltantes = [
                    trazabilidad
                    for trazabilidad in trazabilidades
                    if trazabilidad not in certificados_chapa_existentes
                ]

                items.append({
                    'num_pedido': r[0] or '',
                    'num_armario': r[1] or '',
                    'fecha_modificacion': fecha_mod,
                    'cantidad': cantidad,
                    'semana_entrega': semana_entrega,
                    'fecha_refresco': fecha_ref,
                    'fecha_refresco_sort': fecha_ref_sort,
                    'trazabilidades': trazabilidades,
                    'cert_pintura': str(r[6]).strip() if r[6] else '',
                    'certificados_chapa_encontrados': certificados_chapa_encontrados,
                    'certificados_chapa_faltantes': certificados_chapa_faltantes
                })
            return jsonify({'success': True, 'items': items, 'total': len(items), 'gap': gap})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/refrescar-checklist', methods=['POST'])
def refrescar_checklist_documentos():
    """
    Escanea la carpeta Checklist (ID_Ruta=3 en Rutas_Gestion) buscando archivos
    PDF modificados DESPUÉS del último registro en Checklist_Documentos.
    Inserta los nuevos archivos en la tabla.
    Formato esperado del nombre: "Pedido 202603753_ Armario 76260499.pdf"
    """
    import re
    import os

    request_started_at = time.perf_counter()

    try:
        print("🔄 [Gestión] Iniciando refresco de Checklist_Documentos...")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()

            # 1. Obtener la ruta configurada para Checklist (ID_Ruta = 3)
            cursor.execute("""
                SELECT [Ruta], [Descripcion_Ruta]
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                WHERE [ID_Ruta] = 3
            """)
            ruta_row = cursor.fetchone()
            if not ruta_row or not ruta_row[0]:
                return jsonify({'success': False, 'message': 'No hay ruta configurada para Checklist (ID_Ruta=3)'}), 400

            ruta_carpeta = ruta_row[0].strip()
            print(f"   📁 Ruta Checklist: {ruta_carpeta}")

            if not os.path.exists(ruta_carpeta):
                return jsonify({'success': False, 'message': f'La carpeta no existe o no es accesible: {ruta_carpeta}'}), 400

            # 2. Obtener el SET de nombres de archivo ya registrados en BD
            cursor.execute("""
                SELECT [NombreArchivo]
                FROM [Digitalizacion].[PE].[Checklist_Documentos]
            """)
            nombres_omitir = {row[0] for row in cursor.fetchall()}
            print(f"   📋 Archivos ya en BD: {len(nombres_omitir)}")

            # 2b. Cargar capacidades de la cache: {NumPedido: Cantidad}
            cursor.execute("""
                SELECT [NumPedido], [Cantidad]
                FROM [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache]
            """)
            cache_capacidad = {str(row[0]): (row[1] if row[1] is not None else 0) for row in cursor.fetchall()}
            print(f"   📊 Pedidos en cache: {len(cache_capacidad)}")

            # 3. Escanear archivos PDF de la carpeta
            patron = re.compile(r'Pedido\s+(\d+)_\s*Armario\s+(\d+)', re.IGNORECASE)
            nuevos = 0
            omitidos = 0
            sin_patron = 0
            pedido_no_existe = 0

            for nombre_archivo in os.listdir(ruta_carpeta):
                if not nombre_archivo.lower().endswith('.pdf'):
                    continue

                # Omitir si ya está registrado
                if nombre_archivo in nombres_omitir:
                    omitidos += 1
                    continue

                ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)

                try:
                    ts = os.path.getmtime(ruta_completa)
                    from datetime import datetime
                    fecha_mod = datetime.fromtimestamp(ts)
                except Exception:
                    omitidos += 1
                    continue

                # Extraer NumPedido y NumArmario del nombre
                m = patron.search(nombre_archivo)
                if not m:
                    sin_patron += 1
                    print(f"   ⚠️ Sin patrón reconocible: {nombre_archivo}")
                    continue

                num_pedido = m.group(1)
                num_armario = m.group(2)

                # Validar que el NumPedido existe en la cache
                if num_pedido not in cache_capacidad:
                    pedido_no_existe += 1
                    print(f"   ❌ Pedido no encontrado en cache: {num_pedido} ({nombre_archivo})")
                    continue

                # Insertar en la tabla
                cursor.execute("""
                    INSERT INTO [Digitalizacion].[PE].[Checklist_Documentos]
                        ([NumPedido], [NumArmario], [NombreArchivo], [Ruta_Completa], [Fecha_Modificacion_Archivo])
                    VALUES (?, ?, ?, ?, ?)
                """, (num_pedido, num_armario, nombre_archivo, ruta_completa, fecha_mod))
                nuevos += 1

            conn.commit()

            elapsed = time.perf_counter() - request_started_at
            print(f"✅ [Gestión] Refresco completado en {elapsed:.2f}s — Nuevos: {nuevos}, Omitidos: {omitidos}, Sin patrón: {sin_patron}, Pedido no existe: {pedido_no_existe}")

            return jsonify({
                'success': True,
                'nuevos': nuevos,
                'omitidos': omitidos,
                'sin_patron': sin_patron,
                'pedido_no_existe': pedido_no_existe,
                'archivos_en_bd': len(nombres_omitir),
                'message': f'{nuevos} nuevos archivos registrados'
            })

    except Exception as e:
        print(f"❌ Error [Gestión] refrescando checklist tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/gestion/certificados-pintura', methods=['GET'])
def get_gestion_certificados_pintura():
    """Obtiene los registros de Certificados_Pintura_Documentos."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT cp.[ID], cp.[NumArmario], cp.[NombreArchivo],
                       cp.[Fecha_Modificacion_Archivo], cp.[Fecha_Registro],
                       ISNULL(MAX(cd.[NumPedido]), '') AS NumPedido
                FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos] cp
                LEFT JOIN [Digitalizacion].[PE].[Checklist_Documentos] cd
                    ON cd.[NumArmario] = cp.[NumArmario]
                GROUP BY cp.[ID], cp.[NumArmario], cp.[NombreArchivo],
                         cp.[Fecha_Modificacion_Archivo], cp.[Fecha_Registro]
                ORDER BY cp.[Fecha_Modificacion_Archivo] DESC, cp.[NumArmario] ASC
            """)
            rows = cursor.fetchall()
            docs = []
            for r in rows:
                docs.append({
                    'id': r[0],
                    'num_armario': str(r[1]) if r[1] is not None else '',
                    'nombre_archivo': str(r[2]) if r[2] is not None else '',
                    'fecha_modificacion': r[3].strftime('%d/%m/%Y %H:%M') if r[3] else '',
                    'fecha_modificacion_iso': r[3].isoformat() if r[3] else '',
                    'fecha_registro': r[4].strftime('%d/%m/%Y %H:%M') if r[4] else '',
                    'num_pedido': str(r[5]) if r[5] else '',
                })
            return jsonify({'success': True, 'docs': docs, 'total': len(docs)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/certificados-pintura/faltantes/export', methods=['GET'])
def export_gestion_certificados_pintura_faltantes():
    """Exporta a Excel los armarios de Checklist_Documentos sin certificado de pintura."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    cd.[NumArmario],
                    ISNULL(MAX(cd.[NumPedido]), '') AS NumPedido,
                    MAX(cd.[Fecha_Modificacion_Archivo]) AS UltimaFechaChecklist
                FROM [Digitalizacion].[PE].[Checklist_Documentos] cd
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos] cp
                    WHERE cp.[NumArmario] = cd.[NumArmario]
                )
                GROUP BY cd.[NumArmario]
                ORDER BY MAX(cd.[Fecha_Modificacion_Archivo]) DESC
            """)
            rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sin Certificado Pintura'

        encabezados = ['NumPedido', 'NumArmario', 'Fecha Checklist']
        ws.append(encabezados)

        header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_num in range(1, len(encabezados) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal='center')

        for r in rows:
            fecha_str = r[2].strftime('%d/%m/%Y %H:%M') if r[2] else ''
            ws.append([str(r[1]) if r[1] else '', str(r[0]) if r[0] else '', fecha_str])

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f'armarios_sin_certificado_pintura_{timestamp}.xlsx'
        filepath  = os.path.join(RUTA_PDFS, filename)
        wb.save(filepath)

        print(f'✅ [Gestión] Excel faltantes generado: {filepath}')
        return send_file(filepath, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl no está instalado'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/certificados-pintura/faltantes', methods=['GET'])
def get_gestion_certificados_pintura_faltantes():
    """Devuelve armarios de Checklist_Documentos que NO tienen certificado de pintura."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    cd.[NumArmario],
                    ISNULL(MAX(cd.[NumPedido]), '') AS NumPedido,
                    MAX(cd.[Fecha_Modificacion_Archivo]) AS UltimaFechaChecklist
                FROM [Digitalizacion].[PE].[Checklist_Documentos] cd
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos] cp
                    WHERE cp.[NumArmario] = cd.[NumArmario]
                )
                GROUP BY cd.[NumArmario]
                ORDER BY MAX(cd.[Fecha_Modificacion_Archivo]) DESC
            """)
            rows = cursor.fetchall()
            faltantes = []
            for r in rows:
                faltantes.append({
                    'num_armario': str(r[0]) if r[0] is not None else '',
                    'num_pedido':  str(r[1]) if r[1] else '',
                    'fecha_checklist': r[2].strftime('%d/%m/%Y %H:%M') if r[2] else '',
                    'fecha_checklist_iso': r[2].isoformat() if r[2] else '',
                })
            return jsonify({'success': True, 'faltantes': faltantes, 'total': len(faltantes)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/certificados-pintura/<int:doc_id>', methods=['DELETE'])
def delete_gestion_certificado_pintura(doc_id):
    """Elimina un registro de Certificados_Pintura_Documentos por su ID."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("SELECT [NombreArchivo] FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos] WHERE [ID] = ?", (doc_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            nombre_archivo = row[0]
            cursor.execute("DELETE FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos] WHERE [ID] = ?", (doc_id,))
            conn.commit()
            print(f'🗑️ [Gestión] Certificados_Pintura ID={doc_id} eliminado: {nombre_archivo}')
            return jsonify({'success': True, 'message': f'Registro "{nombre_archivo}" eliminado correctamente'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/refrescar-certificados-pintura', methods=['POST'])
def refrescar_certificados_pintura():
    """
    Escanea la carpeta 'Ruta certificados pintura' de Rutas_Gestion buscando archivos
    cuyo nombre (sin extensión) coincida con algún NumArmario en Checklist_Documentos.
    Solo inserta los que aún no están en Certificados_Pintura_Documentos.
    """
    import re as _re
    import traceback
    request_started_at = time.perf_counter()
    try:
        print("🔄 [Gestión] Iniciando refresco de Certificados_Pintura_Documentos...")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()

            # 1. Obtener la ruta configurada para Certificados Pintura
            #    Busca por ID_Ruta=2 primero (descripción exacta "Ruta certificados pintura"),
            #    y como fallback por LIKE en descripción
            cursor.execute("""
                SELECT [ID_Ruta], [Ruta], [Descripcion_Ruta]
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                WHERE [ID_Ruta] = 2
                   OR LOWER([Descripcion_Ruta]) LIKE '%certificados pintura%'
                ORDER BY CASE WHEN [ID_Ruta] = 2 THEN 0 ELSE 1 END
            """)
            ruta_row = cursor.fetchone()
            if not ruta_row or not ruta_row[1]:
                # Mostrar qué rutas hay disponibles para ayudar al diagnóstico
                cursor.execute("SELECT [ID_Ruta], [Descripcion_Ruta], [Ruta] FROM [Digitalizacion].[PE].[Rutas_Gestion]")
                rutas_disponibles = [(r[0], r[1], r[2]) for r in cursor.fetchall()]
                print(f"   ⚠️ Rutas disponibles en Rutas_Gestion:")
                for rd in rutas_disponibles:
                    print(f"      ID={rd[0]} | Desc='{rd[1]}' | Ruta='{rd[2]}'")
                return jsonify({
                    'success': False,
                    'message': f'No hay ruta configurada para Certificados Pintura (ID_Ruta=2). Rutas disponibles: {[r[1] for r in rutas_disponibles]}'
                }), 400

            id_ruta   = ruta_row[0]
            ruta_carpeta = ruta_row[1].strip()
            desc_ruta = ruta_row[2]
            print(f"   📁 Ruta Certificados Pintura (ID_Ruta={id_ruta}): {ruta_carpeta}")

            if not os.path.exists(ruta_carpeta):
                return jsonify({'success': False, 'message': f'La carpeta no existe o no es accesible: {ruta_carpeta}'}), 400

            # 2. Obtener el SET de NumArmarios que existen en Checklist_Documentos
            cursor.execute("SELECT DISTINCT [NumArmario] FROM [Digitalizacion].[PE].[Checklist_Documentos]")
            armarios_validos = {str(row[0]).strip() for row in cursor.fetchall() if row[0]}
            print(f"   📋 Armarios válidos (Checklist): {len(armarios_validos)}")

            if not armarios_validos:
                return jsonify({'success': True, 'nuevos': 0, 'omitidos': 0, 'sin_armario': 0,
                                'message': 'No hay armarios en Checklist_Documentos. Refresca primero esa tabla.'}), 200

            # 3. Obtener el SET de NombreArchivo ya registrados (para evitar duplicados)
            cursor.execute("SELECT [NombreArchivo] FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos]")
            nombres_omitir = {row[0] for row in cursor.fetchall()}
            print(f"   📋 Archivos ya en BD: {len(nombres_omitir)}")

            # 4. Escanear archivos en la carpeta con os.scandir (una sola llamada al SO,
            #    DirEntry cachea stat → evita isfile/getmtime adicionales por red)
            from datetime import datetime as _dt
            nuevos = 0
            omitidos = 0
            sin_armario = 0  # archivos cuyo stem no coincide con ningún armario
            batch_insertar = []  # acumulamos filas para insertar en bloque

            print(f"   🔍 [Gestión] Listando archivos en carpeta pintura...")
            with os.scandir(ruta_carpeta) as it:
                for entry in it:
                    if not entry.is_file(follow_symlinks=False):
                        continue

                    nombre_archivo = entry.name
                    stem = os.path.splitext(nombre_archivo)[0].strip()

                    # Comprobar si el stem coincide con algún armario válido
                    if stem not in armarios_validos:
                        sin_armario += 1
                        continue

                    # Omitir si el nombre de archivo ya está registrado
                    if nombre_archivo in nombres_omitir:
                        omitidos += 1
                        continue

                    try:
                        ts = entry.stat(follow_symlinks=False).st_mtime
                        fecha_mod = _dt.fromtimestamp(ts)
                    except Exception:
                        omitidos += 1
                        continue

                    batch_insertar.append((stem, nombre_archivo, entry.path, fecha_mod))

            print(f"   📂 [Gestión] Archivos a insertar: {len(batch_insertar)}, sin_armario: {sin_armario}, omitidos: {omitidos}")

            if batch_insertar:
                cursor.executemany("""
                    INSERT INTO [Digitalizacion].[PE].[Certificados_Pintura_Documentos]
                        ([NumArmario], [NombreArchivo], [Ruta_Completa], [Fecha_Modificacion_Archivo])
                    VALUES (?, ?, ?, ?)
                """, batch_insertar)
                nuevos = len(batch_insertar)

            conn.commit()

            elapsed = time.perf_counter() - request_started_at
            print(f"✅ [Gestión] Certificados refresco completado en {elapsed:.2f}s — Nuevos: {nuevos}, Omitidos: {omitidos}, Sin armario: {sin_armario}")

            return jsonify({
                'success': True,
                'nuevos': nuevos,
                'omitidos': omitidos,
                'sin_armario': sin_armario,
                'message': f'{nuevos} nuevos certificados registrados'
            })

    except Exception as e:
        print(f"❌ Error [Gestión] refrescando certificados pintura tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/gestion/certificados-chapa', methods=['GET'])
def get_gestion_certificados_chapa():
    """Obtiene los registros de Certificados_Chapa_Documentos."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [ID], [DATO], [NombreArchivo], [RutaRelativa],
                       [Fecha_Modificacion_Archivo], [Fecha_Registro]
                FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos]
                ORDER BY [Fecha_Modificacion_Archivo] DESC, [DATO] ASC
            """)
            rows = cursor.fetchall()
            docs = []
            for r in rows:
                docs.append({
                    'id': r[0],
                    'dato': str(r[1]) if r[1] is not None else '',
                    'nombre_archivo': str(r[2]) if r[2] is not None else '',
                    'ruta_relativa': str(r[3]) if r[3] is not None else '',
                    'fecha_modificacion': r[4].strftime('%d/%m/%Y %H:%M') if r[4] else '',
                    'fecha_modificacion_iso': r[4].isoformat() if r[4] else '',
                    'fecha_registro': r[5].strftime('%d/%m/%Y %H:%M') if r[5] else '',
                })
            return jsonify({'success': True, 'docs': docs, 'total': len(docs)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/certificados-chapa/<int:doc_id>', methods=['DELETE'])
def delete_gestion_certificado_chapa(doc_id):
    """Elimina un registro de Certificados_Chapa_Documentos por su ID."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("SELECT [NombreArchivo] FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos] WHERE [ID] = ?", (doc_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Registro no encontrado'}), 404
            nombre_archivo = row[0]
            cursor.execute("DELETE FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos] WHERE [ID] = ?", (doc_id,))
            conn.commit()
            print(f'🗑️ [Gestión] Certificados_Chapa ID={doc_id} eliminado: {nombre_archivo}')
            return jsonify({'success': True, 'message': f'Registro "{nombre_archivo}" eliminado correctamente'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/refrescar-certificados-chapa', methods=['POST'])
def refrescar_certificados_chapa():
    """
    Escanea recursivamente la carpeta de Recepción (ID_Ruta=1) buscando archivos.
    Solo inserta los que sean más recientes que MAX(Fecha_Modificacion_Archivo) de la tabla
    o que no estén ya registrados. DATO = stem del nombre de archivo.
    """
    request_started_at = time.perf_counter()
    try:
        print("🔄 [Gestión] Iniciando refresco de Certificados_Chapa_Documentos...")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()

            # 1. Obtener TODAS las rutas configuradas para Chapa
            cursor.execute("""
                SELECT [ID_Ruta], [Ruta], [Descripcion_Ruta]
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                WHERE [Descripcion_Ruta] = 'Ruta certificados de chapa'
                  AND [Ruta] IS NOT NULL AND LTRIM(RTRIM([Ruta])) <> ''
                ORDER BY [ID_Ruta]
            """)
            rutas_chapa = [(r[0], r[1].strip(), r[2]) for r in cursor.fetchall()]
            if not rutas_chapa:
                return jsonify({
                    'success': False,
                    'message': 'No hay rutas configuradas para Certificados Chapa (Descripcion_Ruta = "Ruta certificados de chapa")'
                }), 400

            rutas_accesibles = [r for r in rutas_chapa if os.path.exists(r[1])]
            if not rutas_accesibles:
                rutas_str = '; '.join(r[1] for r in rutas_chapa)
                return jsonify({'success': False, 'message': f'Ninguna carpeta es accesible: {rutas_str}'}), 400

            for r in rutas_chapa:
                accesible = '✅' if os.path.exists(r[1]) else '❌ (inaccesible)'
                print(f"   📁 Ruta Chapa (ID={r[0]}) {accesible}: {r[1]}")

            # 2. Obtener el SET de rutas completas ya registradas (RutaRelativa = ruta absoluta completa)
            #    También cargamos NombreArchivo para dedup por nombre cuando RutaRelativa esté vacía (registros legados)
            cursor.execute("""
                SELECT [RutaRelativa], [NombreArchivo]
                FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos]
            """)
            rutas_existentes   = set()   # rutas absolutas completas (registros nuevos)
            nombres_existentes = set()   # nombres de archivo (fallback para registros legados sin ruta)
            for row in cursor.fetchall():
                ruta_bd   = (row[0] or '').strip()
                nombre_bd = (row[1] or '').strip()
                if ruta_bd and os.path.sep in ruta_bd:
                    # Registro nuevo: RutaRelativa contiene ruta absoluta
                    rutas_existentes.add(ruta_bd.lower())
                else:
                    # Registro legado: solo tenemos el nombre
                    nombres_existentes.add(nombre_bd)
            print(f"   📋 Ya en BD — por ruta: {len(rutas_existentes)}, por nombre (legado): {len(nombres_existentes)}")

            # 3. Escanear recursivamente todas las rutas accesibles
            from pathlib import Path as _Path
            from datetime import datetime as _dt
            nuevos = 0
            omitidos = 0

            for id_ruta_chapa, ruta_carpeta, _ in rutas_accesibles:
                print(f"   🔍 Escaneando ruta ID={id_ruta_chapa}: {ruta_carpeta}")
                for root, dirs, files in os.walk(ruta_carpeta):
                    for nombre_archivo in files:
                        ruta_completa = os.path.join(root, nombre_archivo)

                        # Saltar si ya está registrado (por ruta completa o por nombre legado)
                        if ruta_completa.lower() in rutas_existentes:
                            omitidos += 1
                            continue
                        if nombre_archivo in nombres_existentes:
                            omitidos += 1
                            continue

                        if not os.path.isfile(ruta_completa):
                            continue

                        dato = _Path(nombre_archivo).stem

                        try:
                            ts = os.path.getmtime(ruta_completa)
                            fecha_mod = _dt.fromtimestamp(ts)
                        except Exception:
                            omitidos += 1
                            continue

                        try:
                            cursor.execute("""
                                INSERT INTO [Digitalizacion].[PE].[Certificados_Chapa_Documentos]
                                    ([DATO], [NombreArchivo], [RutaRelativa], [Fecha_Modificacion_Archivo])
                                VALUES (?, ?, ?, ?)
                            """, (dato, nombre_archivo, ruta_completa, fecha_mod))
                            nuevos += 1
                            rutas_existentes.add(ruta_completa.lower())
                        except Exception as e_ins:
                            omitidos += 1
                            continue

            conn.commit()

            elapsed = time.perf_counter() - request_started_at
            print(f"✅ [Gestión] Certificados Chapa refresco completado en {elapsed:.2f}s — Nuevos: {nuevos}, Omitidos: {omitidos}")

            return jsonify({
                'success': True,
                'nuevos': nuevos,
                'omitidos': omitidos,
                'message': f'{nuevos} nuevos certificados de chapa registrados'
            })

    except Exception as e:
        print(f"❌ Error [Gestión] refrescando certificados chapa: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


# ====================================================================================
# 🆕 ENDPOINT PARA EXPORTAR DOCUMENTACIÓN DE UN PEDIDO COMPLETO (VERDE)
# ====================================================================================

@app.route('/api/gestion/exportar-documentacion', methods=['POST'])
def exportar_documentacion_pedido():
    """
    Copia toda la documentación de un pedido 100 % verde a la
    carpeta destino configurada (Descripcion_Ruta = 'Ruta destino documentación').
    Crea subcarpeta {ruta_destino}\\{NumPedido} y copia:
      - Todos los archivos Checklist (Ruta_Completa de Checklist_Documentos)
      - Todos los cert. pintura (Ruta_Completa de Certificados_Pintura_Documentos)
      - Todos los cert. chapa    (RutaRelativa de Certificados_Chapa_Documentos)
    Solo permite ejecutar si checklist, pintura y chapa están 100 % OK.
    Al completarse con éxito registra la exportación en PE.Exportaciones_Documentacion.
    """
    import shutil
    import re as _re

    try:
        data = request.get_json() or {}
        num_pedido = (data.get('num_pedido') or '').strip()
        if not num_pedido:
            return jsonify({'success': False, 'message': 'num_pedido es obligatorio'}), 400

        print(f"📤 [Exportar] Iniciando exportación para NumPedido={num_pedido}")

        cfg = _read_config()
        gap = int(cfg.get('gap_trazabilidad_semanas', 0))

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()

            # ── 1. Ruta destino ───────────────────────────────────────────────
            cursor.execute("""
                SELECT [Ruta]
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                WHERE [Descripcion_Ruta] = 'Ruta destino documentación'
                  AND [Ruta] IS NOT NULL AND LTRIM(RTRIM([Ruta])) <> ''
            """)
            ruta_row = cursor.fetchone()
            if not ruta_row:
                return jsonify({'success': False, 'message': 'No hay ruta destino configurada (Descripcion_Ruta = "Ruta destino documentación")'}), 400
            ruta_destino = ruta_row[0].strip()
            if not os.path.exists(ruta_destino):
                return jsonify({'success': False, 'message': f'La ruta destino no existe o no es accesible: {ruta_destino}'}), 400

            # ── 2. Datos del pedido: armarios + archivos checklist ────────────
            cursor.execute("""
                SELECT [NumArmario], [Ruta_Completa], [NombreArchivo]
                FROM [Digitalizacion].[PE].[Checklist_Documentos]
                WHERE [NumPedido] = ?
            """, (num_pedido,))
            checklist_rows = cursor.fetchall()
            if not checklist_rows:
                return jsonify({'success': False, 'message': f'No hay registros de checklist para NumPedido={num_pedido}'}), 404

            num_armarios = [r[0] for r in checklist_rows]

            # ── 3. Datos de ERP: cantidad y semana de entrega ─────────────────
            cursor.execute("""
                SELECT [Cantidad], [SEMANAENTREGA]
                FROM [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache]
                WHERE [NumPedido] = ?
            """, (num_pedido,))
            erp_row = cursor.fetchone()
            cantidad_erp = int(erp_row[0]) if erp_row and erp_row[0] is not None else None
            semana_entrega = str(erp_row[1]).strip() if erp_row and erp_row[1] else ''

            # Validar checklist: nº armarios == cantidad ERP
            if cantidad_erp is None or len(checklist_rows) != cantidad_erp:
                return jsonify({
                    'success': False,
                    'message': f'Checklist no está completo: {len(checklist_rows)} armarios vs {cantidad_erp} esperados'
                }), 400

            # ── 4. Certificados pintura ────────────────────────────────────────
            placeholders = ','.join('?' * len(num_armarios))
            cursor.execute(f"""
                SELECT [NumArmario], [Ruta_Completa]
                FROM [Digitalizacion].[PE].[Certificados_Pintura_Documentos]
                WHERE [NumArmario] IN ({placeholders})
            """, tuple(num_armarios))
            pintura_rows = cursor.fetchall()
            pintura_por_armario = {r[0]: r[1] for r in pintura_rows}

            armarios_sin_pintura = [a for a in num_armarios if a not in pintura_por_armario]
            if armarios_sin_pintura:
                return jsonify({
                    'success': False,
                    'message': f'Faltan certificados de pintura para armarios: {armarios_sin_pintura}'
                }), 400

            # ── 5. Trazabilidades y certificados chapa ────────────────────────
            # Cargar mapa de trazabilidad
            cursor.execute("""
                SELECT [numsemana], [ano], [DATO]
                FROM [Digitalizacion].[PE].[vw_LotesChapaProduccion]
                WHERE [numsemana] IS NOT NULL AND [ano] IS NOT NULL
            """)
            traz_map = {}
            anos_con_semana53 = set()
            for tr in cursor.fetchall():
                sem = int(tr[0]) if tr[0] is not None else None
                ano = int(tr[1]) if tr[1] is not None else None
                dato = str(tr[2] or '').strip()
                if sem is None or ano is None:
                    continue
                traz_map.setdefault((sem, ano), set())
                if dato:
                    traz_map[(sem, ano)].add(dato)
                if sem == 53:
                    anos_con_semana53.add(ano)

            def _semanas_obj(semana_entrega_str, gap_semanas):
                if not semana_entrega_str or gap_semanas == 0:
                    return []
                m = _re.search(r'(\d+)[^\d]+(\d{4})', semana_entrega_str)
                if not m:
                    return []
                sem = int(m.group(1))
                ano = int(m.group(2))
                ts = sem - gap_semanas
                ta = ano
                while ts <= 0:
                    ta -= 1
                    ts = (53 if ta in anos_con_semana53 else 52) + ts
                return [(ts, ta)]

            trazabilidades = []
            if gap > 0 and semana_entrega:
                for key in _semanas_obj(semana_entrega, gap):
                    trazabilidades.extend(traz_map.get(key, set()))
                trazabilidades = sorted(set(trazabilidades))

            if not trazabilidades:
                return jsonify({
                    'success': False,
                    'message': f'No hay trazabilidad de fabricación para NumPedido={num_pedido}. '
                               f'Comprueba que la semana de entrega está configurada y el GAP es mayor que 0.'
                }), 400

            # Buscar archivos de chapa (todos los que tienen ruta en BD, sin filtrar por existencia aquí)
            chapa_rutas = []  # lista de (dato, ruta_absoluta)
            if trazabilidades:
                pl2 = ','.join('?' * len(trazabilidades))
                cursor.execute(f"""
                    SELECT [DATO], [RutaRelativa]
                    FROM [Digitalizacion].[PE].[Certificados_Chapa_Documentos]
                    WHERE [DATO] IN ({pl2})
                      AND [RutaRelativa] IS NOT NULL
                      AND LTRIM(RTRIM([RutaRelativa])) <> ''
                """, tuple(trazabilidades))
                chapa_rows_db = cursor.fetchall()
                dato_a_ruta = {}
                for row in chapa_rows_db:
                    dato = str(row[0] or '').strip()
                    ruta = (row[1] or '').strip()
                    if dato and ruta:
                        dato_a_ruta[dato] = ruta

                # Validar que todos los DATOs tienen ruta registrada en BD
                faltantes_chapa = [t for t in trazabilidades if t not in dato_a_ruta]
                if faltantes_chapa:
                    return jsonify({
                        'success': False,
                        'message': f'Faltan certificados de chapa en BD: {faltantes_chapa}'
                    }), 400

                chapa_rutas = [(dato, ruta) for dato, ruta in dato_a_ruta.items()]
                print(f"   🔩 [Exportar] Chapa: {len(chapa_rutas)} archivos encontrados en BD")

        # ── 6. Crear subcarpetas destino y copiar archivos ────────────────────
        carpeta_pedido   = os.path.join(ruta_destino, num_pedido)
        carpeta_checklist = os.path.join(carpeta_pedido, 'Checklist')
        carpeta_pintura   = os.path.join(carpeta_pedido, 'Certificados_Pintura')
        carpeta_chapa     = os.path.join(carpeta_pedido, 'Certificados_Chapa')
        for carpeta in (carpeta_checklist, carpeta_pintura, carpeta_chapa):
            os.makedirs(carpeta, exist_ok=True)
        print(f"   📁 Carpeta destino: {carpeta_pedido}")

        copiados = []
        errores  = []

        def _copiar(src, subcarpeta, etiqueta):
            if not src:
                errores.append(f'{etiqueta}: ruta vacía')
                return
            if not os.path.exists(src):
                errores.append(f'{etiqueta}: archivo no encontrado ({src})')
                return
            try:
                dst = os.path.join(subcarpeta, os.path.basename(src))
                if os.path.exists(dst) and not os.path.samefile(src, dst):
                    base, ext = os.path.splitext(os.path.basename(src))
                    import time as _t
                    dst = os.path.join(subcarpeta, f'{base}_{int(_t.time())}{ext}')
                shutil.copy2(src, dst)
                copiados.append(os.path.relpath(dst, carpeta_pedido))
                print(f"   ✅ Copiado [{etiqueta}]: {os.path.relpath(dst, carpeta_pedido)}")
            except Exception as ex:
                errores.append(f'{etiqueta}: {ex}')

        # Checklist
        for row in checklist_rows:
            ruta_cl = (row[1] or '').strip()
            _copiar(ruta_cl, carpeta_checklist, f'Checklist armario {row[0]}')

        # Pintura
        for armario, ruta_p in pintura_por_armario.items():
            _copiar((ruta_p or '').strip(), carpeta_pintura, f'Pintura armario {armario}')

        # Chapa
        for dato, ruta_ch in chapa_rutas:
            _copiar(ruta_ch, carpeta_chapa, f'Chapa {dato}')

        print(f"✅ [Exportar] NumPedido={num_pedido}: {len(copiados)} archivos copiados, {len(errores)} errores")

        # ── 7. Registrar exportación en BD (upsert) ─────────────────────────
        if len(errores) == 0:
            try:
                with ConexionODBC('Digitalizacion') as conn2:
                    if conn2:
                        cur2 = conn2.cursor()
                        # Crear tabla si no existe
                        cur2.execute("""
                            IF NOT EXISTS (
                                SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                                WHERE TABLE_SCHEMA = 'PE'
                                  AND TABLE_NAME   = 'Exportaciones_Documentacion'
                            )
                            CREATE TABLE [Digitalizacion].[PE].[Exportaciones_Documentacion] (
                                [ID]               INT IDENTITY(1,1) PRIMARY KEY,
                                [NumPedido]        NVARCHAR(50)  NOT NULL,
                                [FechaExportacion] DATETIME      NOT NULL DEFAULT GETDATE(),
                                [CarpetaDestino]   NVARCHAR(500) NULL,
                                CONSTRAINT UQ_ExportDoc_NumPedido UNIQUE ([NumPedido])
                            )
                        """)
                        # Upsert: si ya existe actualizar fecha, si no insertar
                        cur2.execute("""
                            IF EXISTS (SELECT 1 FROM [Digitalizacion].[PE].[Exportaciones_Documentacion] WHERE [NumPedido] = ?)
                                UPDATE [Digitalizacion].[PE].[Exportaciones_Documentacion]
                                SET [FechaExportacion] = GETDATE(), [CarpetaDestino] = ?
                                WHERE [NumPedido] = ?
                            ELSE
                                INSERT INTO [Digitalizacion].[PE].[Exportaciones_Documentacion]
                                    ([NumPedido], [FechaExportacion], [CarpetaDestino])
                                VALUES (?, GETDATE(), ?)
                        """, (num_pedido, carpeta_pedido, num_pedido, num_pedido, carpeta_pedido))
                        conn2.commit()
                        print(f"   📝 [Exportar] Marca de exportación guardada para NumPedido={num_pedido}")
            except Exception as ex_bd:
                print(f"   ⚠️ [Exportar] No se pudo guardar marca en BD: {ex_bd}")

        return jsonify({
            'success': len(errores) == 0,
            'message': f'{len(copiados)} archivos exportados a {carpeta_pedido}' + (f' — {len(errores)} errores' if errores else ''),
            'carpeta': carpeta_pedido,
            'copiados': copiados,
            'errores': errores
        }), 200 if len(errores) == 0 else 207

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500


@app.route('/api/gestion/pedidos-cache', methods=['GET'])
def get_gestion_pedidos_cache():
    """Devuelve todos los registros de Pedido_Cantidad_SemanaEntrega_Cache para la vista ERP."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [NumPedido], [Cantidad], [SEMANAENTREGA], [FechaRefresco]
                FROM [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache]
                ORDER BY [FechaRefresco] DESC
            """)
            pedidos = []
            for r in cursor.fetchall():
                fecha_raw = r[3]
                fecha_str = fecha_raw.strftime('%d/%m/%Y %H:%M') if fecha_raw else ''
                fecha_sort = fecha_raw.strftime('%Y-%m-%dT%H:%M:%S') if fecha_raw else ''
                pedidos.append({
                    'num_pedido':        r[0] or '',
                    'cantidad':          r[1],
                    'semana_entrega':    r[2] or '',
                    'fecha_refresco':    fecha_str,
                    'fecha_refresco_sort': fecha_sort,
                })
            return jsonify({'success': True, 'pedidos': pedidos, 'total': len(pedidos)})
    except Exception as e:
        print(f"❌ Error en get_gestion_pedidos_cache: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/exportaciones-documentacion', methods=['GET'])
def get_exportaciones_documentacion():
    """Devuelve todos los pedidos marcados como exportados."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            # Crear tabla si no existe
            cursor.execute("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.TABLES
                    WHERE TABLE_SCHEMA = 'PE' AND TABLE_NAME = 'Exportaciones_Documentacion'
                )
                CREATE TABLE [Digitalizacion].[PE].[Exportaciones_Documentacion] (
                    [ID]               INT IDENTITY(1,1) PRIMARY KEY,
                    [NumPedido]        NVARCHAR(50)  NOT NULL,
                    [FechaExportacion] DATETIME      NOT NULL DEFAULT GETDATE(),
                    [CarpetaDestino]   NVARCHAR(500) NULL,
                    CONSTRAINT UQ_ExportDoc_NumPedido UNIQUE ([NumPedido])
                )
            """)
            conn.commit()
            cursor.execute("""
                SELECT [NumPedido], [FechaExportacion], [CarpetaDestino]
                FROM [Digitalizacion].[PE].[Exportaciones_Documentacion]
                ORDER BY [FechaExportacion] DESC
            """)
            exportaciones = []
            for r in cursor.fetchall():
                exportaciones.append({
                    'num_pedido': r[0],
                    'fecha': r[1].strftime('%d/%m/%Y %H:%M') if r[1] else '',
                    'carpeta': r[2] or ''
                })
            return jsonify({'success': True, 'exportaciones': exportaciones})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/gestion/exportaciones-documentacion/<num_pedido>', methods=['DELETE'])
def delete_exportacion_documentacion(num_pedido):
    """Elimina la marca de exportado de un pedido."""
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Exportaciones_Documentacion]
                WHERE [NumPedido] = ?
            """, (num_pedido,))
            conn.commit()
            print(f"🗑️ [Exportaciones] Marca eliminada para NumPedido={num_pedido}")
            return jsonify({'success': True, 'message': f'Marca eliminada para pedido {num_pedido}'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    """Obtiene pedidos cacheados con cantidad, semana de entrega y fecha de refresco."""
    request_started_at = time.perf_counter()

    try:
        print("🔍 [Gestión] Obteniendo pedidos cacheados...")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión a base de datos'}), 500

            cursor = conn.cursor()
            try:
                cursor.timeout = 10
            except Exception:
                pass

            cursor.execute("""
                SELECT TOP (1000)
                    [NumPedido],
                    [Cantidad],
                    [SEMANAENTREGA],
                    [FechaRefresco]
                FROM [Digitalizacion].[PE].[Pedido_Cantidad_SemanaEntrega_Cache]
                ORDER BY [FechaRefresco] DESC, [NumPedido] ASC
            """)

            pedidos = []
            for row in cursor.fetchall():
                fecha_refresco = row[3]
                cantidad = None

                if row[1] is not None:
                    try:
                        cantidad_float = float(row[1])
                        cantidad = int(cantidad_float) if cantidad_float.is_integer() else cantidad_float
                    except Exception:
                        cantidad = row[1]

                pedidos.append({
                    'num_pedido': str(row[0]).strip() if row[0] is not None else '',
                    'cantidad': cantidad,
                    'semana_entrega': str(row[2]).strip() if row[2] is not None else '',
                    'fecha_refresco': fecha_refresco.strftime('%d/%m/%Y %H:%M:%S') if isinstance(fecha_refresco, datetime) else (str(fecha_refresco) if fecha_refresco else ''),
                    'fecha_refresco_sort': fecha_refresco.isoformat() if isinstance(fecha_refresco, datetime) else (str(fecha_refresco) if fecha_refresco else '')
                })

            print(f"✅ [Gestión] {len(pedidos)} pedidos cacheados obtenidos en {time.perf_counter() - request_started_at:.2f}s")

            return jsonify({
                'success': True,
                'pedidos': pedidos,
                'total': len(pedidos)
            })

    except Exception as e:
        print(f"❌ Error [Gestión] obteniendo pedidos cacheados tras {time.perf_counter() - request_started_at:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error del servidor: {str(e)}'}), 500

@app.route('/api/pedido/<int:id_pedido>', methods=['GET'])
def get_pedido_datos(id_pedido):
    """Obtiene los datos del Pedido para edición"""
    try:
        print(f"🔍 Obteniendo datos del Pedido ID: {id_pedido}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    ID_Pedido,
                    Armario,
                    Fecha,
                    Referencia,
                    Comentarios,
                    Cerrado,
                    Nombre_Pauta,
                    NumPedido
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))
            
            row = cursor.fetchone()
            
            if not row:
                print(f"⚠️ Pedido {id_pedido} no encontrado")
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            # Convertir fecha a string si es datetime
            fecha_str = ''
            if row[2]:
                if isinstance(row[2], datetime):
                    fecha_str = row[2].strftime('%Y-%m-%d')
                else:
                    fecha_str = str(row[2])
            
            pedido = {
                'id_pedido': row[0],
                'armario': row[1] or '',
                'fecha': fecha_str,
                'referencia': row[3] or '',
                'comentarios': row[4] or '',
                'cerrado': row[5] if row[5] is not None else 0,
                'nombre_pauta': row[6] or '',
                'num_pedido': row[7] or ''
            }
            
            print(f"✅ Datos del Pedido obtenidos: {pedido['armario']}")
            
            return jsonify({
                'success': True,
                'pedido': pedido
            })
    
    except Exception as e:
        print(f"❌ Error obteniendo datos del Pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/pedido/<int:id_pedido>', methods=['PUT'])
def actualizar_pedido_datos(id_pedido):
    """Actualiza los datos del Pedido - DESHABILITADO POR AHORA"""
    return jsonify({'success': False, 'message': 'Esta funcionalidad está deshabilitada por mantenimiento'}), 501




@app.route('/api/pedido/<int:id_pedido>', methods=['DELETE'])
def eliminar_pedido(id_pedido):
    """Elimina un Pedido y todos sus datos asociados"""
    try:
        print(f"🗑️ Eliminando Pedido ID: {id_pedido}")
        
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            
            cursor = conn.cursor()
            
            # Obtener información del pedido antes de eliminarlo
            cursor.execute("""
                SELECT Armario, NumPedido 
                FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))
            
            pedido_info = cursor.fetchone()
            if not pedido_info:
                print(f"⚠️ Pedido {id_pedido} no encontrado")
                return jsonify({'success': False, 'message': 'Pedido no encontrado'}), 404
            
            armario_nombre = pedido_info[0]
            num_pedido = pedido_info[1]
            
            print(f"   Armario: {armario_nombre}")
            print(f"   NumPedido: {num_pedido}")
            
            # Paso 1: Contar registros asociados en DatosUser
            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[DatosUser]
                WHERE ID_Pedido = ?
            """, (id_pedido,))
            
            count_datos_user = cursor.fetchone()[0]
            print(f"   📊 Registros en DatosUser: {count_datos_user}")
            
            # Paso 2: Eliminar registros de DatosUser
            if count_datos_user > 0:
                cursor.execute("""
                    DELETE FROM [Digitalizacion].[PE].[DatosUser]
                    WHERE ID_Pedido = ?
                """, (id_pedido,))
                
                print(f"   ✅ {count_datos_user} registros eliminados de DatosUser")
            
            # Paso 3: Eliminar el Pedido
            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Pedido]
                WHERE ID_Pedido = ?
            """, (id_pedido,))
            
            conn.commit()
            
            print(f"✅ Pedido {id_pedido} ({armario_nombre}) eliminado completamente")
            
            return jsonify({
                'success': True,
                'message': f'Armario "{armario_nombre}" y sus {count_datos_user} controles eliminados correctamente'
            })
    
    except Exception as e:
        print(f"❌ Error eliminando Pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ====================================================================================
# ====================================================================================
# ENDPOINT PARA TRAZABILIDAD DE PEDIDOS (vw_LotesChapaProduccion)
# ====================================================================================

@app.route('/api/gestion/trazabilidad-cache', methods=['GET'])
def get_gestion_trazabilidad_cache():
    """
    Devuelve numsemana, ano y DATO desde [PE].[vw_LotesChapaProduccion].
    """
    request_started_at = time.perf_counter()
    try:
        print("🔍 [Gestión] Obteniendo trazabilidad desde vw_LotesChapaProduccion...")
        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500
            cursor = conn.cursor()
            cursor.execute("""
                SELECT [numsemana], [ano], [DATO]
                FROM [Digitalizacion].[PE].[vw_LotesChapaProduccion]
                ORDER BY [ano] DESC, [numsemana] DESC
            """)
            rows = cursor.fetchall()
            items = []
            for r in rows:
                items.append({
                    'numsemana': r[0],
                    'ano':       r[1],
                    'dato':      str(r[2] or ''),
                })
            print(f"✅ [Gestión] {len(items)} filas de trazabilidad en {time.perf_counter() - request_started_at:.2f}s")
            return jsonify({'success': True, 'items': items, 'total': len(items)})
    except Exception as e:
        import traceback
        print(f"❌ Error obteniendo trazabilidad: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ====================================================================================
# ENDPOINTS PARA CONFIGURACIÓN GENERAL (config.json)
# ====================================================================================

_CONFIG_FILE = os.path.join(
    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else BASE_DIR,
    'config.json' if getattr(sys, 'frozen', False) else os.path.join('api', 'config.json')
)

# Clave que se almacena en Rutas_Gestion como Descripcion_Ruta
_CFG_KEY_GAP = 'GAP trazabilidad semanas'


def _read_config():
    """
    Lee la configuración desde BD (Rutas_Gestion) para claves conocidas,
    y complementa con config.json para el resto.
    Siempre devuelve al menos {}.
    """
    cfg = {}
    # 1. Leer JSON como base (puede no existir en el servidor → {})
    try:
        if os.path.exists(_CONFIG_FILE):
            with open(_CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
    except Exception:
        pass

    # 2. Sobreescribir con los valores de BD (tienen prioridad)
    try:
        with ConexionODBC('Digitalizacion') as conn:
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT [Descripcion_Ruta], [Ruta]
                    FROM [Digitalizacion].[PE].[Rutas_Gestion]
                    WHERE [Descripcion_Ruta] = ?
                """, (_CFG_KEY_GAP,))
                row = cursor.fetchone()
                if row and row[1] is not None:
                    try:
                        cfg['gap_trazabilidad_semanas'] = int(str(row[1]).strip())
                    except ValueError:
                        pass
    except Exception:
        pass

    return cfg


def _write_config(data):
    """
    Escribe la configuración:
    - gap_trazabilidad_semanas → BD (Rutas_Gestion, upsert)
    - resto de claves          → config.json
    """
    # 1. Persistir GAP en BD
    if 'gap_trazabilidad_semanas' in data:
        try:
            valor = str(int(data['gap_trazabilidad_semanas']))
            with ConexionODBC('Digitalizacion') as conn:
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT COUNT(*) FROM [Digitalizacion].[PE].[Rutas_Gestion]
                        WHERE [Descripcion_Ruta] = ?
                    """, (_CFG_KEY_GAP,))
                    existe = cursor.fetchone()[0] > 0
                    if existe:
                        cursor.execute("""
                            UPDATE [Digitalizacion].[PE].[Rutas_Gestion]
                            SET [Ruta] = ?
                            WHERE [Descripcion_Ruta] = ?
                        """, (valor, _CFG_KEY_GAP))
                    else:
                        cursor.execute("""
                            INSERT INTO [Digitalizacion].[PE].[Rutas_Gestion]
                                ([Descripcion_Ruta], [Ruta])
                            VALUES (?, ?)
                        """, (_CFG_KEY_GAP, valor))
                    conn.commit()
                    print(f'✅ GAP trazabilidad guardado en BD: {valor} semanas')
        except Exception as e:
            print(f'⚠️ No se pudo guardar GAP en BD: {e}')

    # 2. Guardar el resto en JSON (si es posible escribir)
    try:
        with open(_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # En el servidor puede que no haya permisos de escritura junto al EXE


@app.route('/api/config', methods=['GET'])
def get_config():
    """Devuelve la configuración general (config.json)."""
    try:
        return jsonify({'success': True, 'config': _read_config()})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/config', methods=['PUT'])
def update_config():
    """Actualiza uno o varios campos de la configuración general."""
    try:
        data = request.get_json() or {}
        cfg = _read_config()
        cfg.update(data)
        _write_config(cfg)
        print(f'✅ config.json actualizado: {data}')
        return jsonify({'success': True, 'config': cfg})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ====================================================================================
# ENDPOINTS PARA RUTAS GESTIÓN
# ====================================================================================

@app.route('/api/rutas-gestion', methods=['GET'])

def get_rutas_gestion():
    """Obtiene todas las rutas de gestión configuradas"""
    try:
        print("🔍 Obteniendo rutas de gestión...")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    [ID_Ruta],
                    [Descripcion_Ruta],
                    [Ruta],
                    [Fecha_Modificacion],
                    [Usuario_Modificacion]
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                ORDER BY [ID_Ruta]
            """)

            rutas = []
            for row in cursor.fetchall():
                fecha_str = ''
                if row[3]:
                    try:
                        fecha_str = row[3].strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        fecha_str = str(row[3])
                rutas.append({
                    'id_ruta': row[0],
                    'descripcion_ruta': row[1] or '',
                    'ruta': row[2] or '',
                    'fecha_modificacion': fecha_str,
                    'usuario_modificacion': row[4] or ''
                })

            print(f"✅ {len(rutas)} rutas obtenidas")
            return jsonify({'success': True, 'rutas': rutas})

    except Exception as e:
        print(f"❌ Error obteniendo rutas de gestión: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/rutas-gestion/<int:id_ruta>', methods=['PUT'])
def actualizar_ruta_gestion(id_ruta):
    """Actualiza la ruta de un registro de Rutas_Gestion"""
    try:
        data = request.get_json() or {}
        ruta = (data.get('ruta') or '').strip()
        usuario = (data.get('usuario') or '').strip()

        print(f"📝 Actualizando ruta ID {id_ruta}: '{ruta}' por '{usuario}'")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()

            cursor.execute("""
                SELECT COUNT(*)
                FROM [Digitalizacion].[PE].[Rutas_Gestion]
                WHERE ID_Ruta = ?
            """, (id_ruta,))

            if cursor.fetchone()[0] == 0:
                return jsonify({'success': False, 'message': 'Ruta no encontrada'}), 404

            cursor.execute("""
                UPDATE [Digitalizacion].[PE].[Rutas_Gestion]
                SET [Ruta] = ?,
                    [Fecha_Modificacion] = GETDATE(),
                    [Usuario_Modificacion] = ?
                WHERE ID_Ruta = ?
            """, (ruta if ruta else None, usuario if usuario else None, id_ruta))

            conn.commit()

            print(f"✅ Ruta {id_ruta} actualizada correctamente")
            return jsonify({'success': True, 'message': 'Ruta actualizada correctamente'})

    except Exception as e:
        print(f"❌ Error actualizando ruta {id_ruta}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/rutas-gestion', methods=['POST'])
def crear_ruta_gestion():
    """Crea una nueva fila en Rutas_Gestion (usado para añadir rutas adicionales de chapa)"""
    try:
        data = request.get_json() or {}
        descripcion = (data.get('descripcion_ruta') or '').strip()
        ruta = (data.get('ruta') or '').strip()
        usuario = (data.get('usuario') or '').strip()

        if not descripcion:
            return jsonify({'success': False, 'message': 'descripcion_ruta es obligatorio'}), 400

        print(f"📝 Creando nueva ruta: descripcion='{descripcion}', ruta='{ruta}'")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO [Digitalizacion].[PE].[Rutas_Gestion]
                    ([Descripcion_Ruta], [Ruta], [Fecha_Modificacion], [Usuario_Modificacion])
                OUTPUT INSERTED.ID_Ruta
                VALUES (?, ?, GETDATE(), ?)
            """, (descripcion, ruta if ruta else None, usuario if usuario else None))
            new_id = cursor.fetchone()[0]
            conn.commit()
            print(f"✅ Nueva ruta creada ID={new_id}")
            return jsonify({'success': True, 'id_ruta': new_id, 'message': 'Ruta creada correctamente'})

    except Exception as e:
        print(f"❌ Error creando ruta: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/rutas-gestion/<int:id_ruta>', methods=['DELETE'])
def eliminar_ruta_gestion(id_ruta):
    """Elimina una fila de Rutas_Gestion. Solo permite eliminar filas de 'Ruta certificados de chapa'."""
    try:
        print(f"🗑️ Eliminando ruta ID={id_ruta}")

        with ConexionODBC('Digitalizacion') as conn:
            if not conn:
                return jsonify({'success': False, 'message': 'Error de conexión'}), 500

            cursor = conn.cursor()
            cursor.execute("""
                SELECT [Descripcion_Ruta] FROM [Digitalizacion].[PE].[Rutas_Gestion] WHERE ID_Ruta = ?
            """, (id_ruta,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Ruta no encontrada'}), 404

            if row[0] != 'Ruta certificados de chapa':
                return jsonify({'success': False, 'message': 'Solo se pueden eliminar rutas de tipo "Ruta certificados de chapa"'}), 403

            cursor.execute("""
                DELETE FROM [Digitalizacion].[PE].[Rutas_Gestion] WHERE ID_Ruta = ?
            """, (id_ruta,))
            conn.commit()
            print(f"✅ Ruta ID={id_ruta} eliminada")
            return jsonify({'success': True, 'message': 'Ruta eliminada correctamente'})

    except Exception as e:
        print(f"❌ Error eliminando ruta {id_ruta}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    print("🚀 Iniciando aplicación Checklist Power...")
    print("📦 Versión: 2.0 | Header EMESA + Widget integrado")
    print(f"📁 Directorio base: {BASE_DIR}")
    print(f"📄 Templates: {RUTA_TEMPLATES}")
    print(f"🎨 Assets: {RUTA_ASSETS}")
    print(f"🖼️ Imágenes: {RUTA_IMAGENES}")
    print(f"📋 PDFs Generados (local): {RUTA_PDFS}")
    print(f"☁️ PDFs Compartida: {RUTA_PDFS_COMPARTIDA}")  # 🆕 Mostrar carpeta compartida
    print("=" * 50)
    
    # Verificar que las rutas existen
    if not os.path.exists(RUTA_TEMPLATES):
        print(f"⚠️ ADVERTENCIA: No se encuentra el directorio Templates: {RUTA_TEMPLATES}")
    if not os.path.exists(RUTA_ASSETS):
        print(f"⚠️ ADVERTENCIA: No se encuentra el directorio assets: {RUTA_ASSETS}")
    if not os.path.exists(RUTA_IMAGENES):
        print(f"⚠️ ADVERTENCIA: No se encuentra el directorio IMAGENES: {RUTA_IMAGENES}")
    if not os.path.exists(RUTA_PDFS):
        print(f"⚠️ ADVERTENCIA: No se encuentra el directorio PDFs_Generados: {RUTA_PDFS}")
    if not os.path.exists(RUTA_PDFS_COMPARTIDA):
        print(f"⚠️ ADVERTENCIA: No se encuentra el directorio compartido: {RUTA_PDFS_COMPARTIDA}")
    
    # 🔒 Cargar certificado CA-EMESA para HTTPS
    cert_file, key_file = generate_self_signed_cert()
    
    if cert_file and key_file:
        # Ejecutar con HTTPS usando ssl_context
        print(f"\n🔒 Iniciando servidor HTTPS en https://127.0.0.1:3007")
        print(f"🔒 Acceso local también disponible en https://localhost:3007")
        print(f"🌐 Acceso en red disponible en https://192.168.253.9:3007")
        print(f"✅  Certificado firmado por CA interna EMESA (sin advertencias si TIC distribuyó root_ca.cer).\n")
        try:
            # Importar ssl para mejor control
            import ssl
            context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
            context.load_cert_chain(cert_file, key_file)
            app.run(host='0.0.0.0', port=3007, ssl_context=context, debug=False, use_reloader=False, threaded=True)
        except Exception as e:
            print(f"⚠️ Error con SSL context: {e}")
            print(f"⚠️ Intentando con ssl_context simple...\n")
            app.run(host='0.0.0.0', port=3007, ssl_context=(cert_file, key_file), debug=False, use_reloader=False, threaded=True)
    else:
        # Fallback a HTTP si hay problema con certificado
        print(f"\n⚠️  Iniciando servidor HTTP en http://127.0.0.1:3007")
        print(f"⚠️  Acceso local también disponible en http://localhost:3007")
        print(f"🌐 Acceso en red disponible en http://192.168.253.9:3007")
        print(f"⚠️  El lector de QR NO funcionará desde direcciones IP sin HTTPS\n")
        app.run(host='0.0.0.0', port=3007, debug=False, use_reloader=False, threaded=True)
